"""
PyPSA Model Execution Routes
=============================

Handles PyPSA model configuration, execution, and real-time progress tracking.

Endpoints:
- POST /project/save-model-config - Save model configuration
- POST /project/run-pypsa-model - Execute PyPSA model
- GET /project/pypsa-model-progress - Server-Sent Events for real-time logs
- POST /project/stop-pypsa-model - Stop/cancel running model
- GET /project/pypsa-solver-logs - Stream solver log file in real-time
"""

from fastapi import APIRouter, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional, Dict, Any, List
import os
import json
import asyncio
import logging
from datetime import datetime
from pathlib import Path
import traceback
import sys
from io import StringIO
import signal
import psutil

router = APIRouter()
logger = logging.getLogger(__name__)

# Global variables for streaming logs and process management
current_log_buffer = []
model_status = {"running": False, "completed": False, "error": None, "pid": None}
current_solver_log_path = None


class MonthlyConstraints(BaseModel):
    """Monthly constraint configuration"""
    enabled: bool
    type: str
    cycles: int


class CoreSettings(BaseModel):
    """Core PyPSA settings"""
    committable: bool
    monthlyConstraints: MonthlyConstraints
    co2Constraints: bool


class EnergyManagement(BaseModel):
    """Energy management settings"""
    batteryCycle: bool
    storageDischarging: str
    runPypsaModelOn: str


class Optimization(BaseModel):
    """Optimization settings"""
    solver: str
    multiYearInvestment: str
    weightings: str


class AssetManagement(BaseModel):
    """Asset management settings"""
    generatorRetirements: bool
    storageRetirements: bool
    generatorCluster: bool


class AdvancedOptions(BaseModel):
    """Advanced option settings"""
    batteryCycleCost: bool
    rollingHorizon: bool


class Configuration(BaseModel):
    """Complete configuration model"""
    coreSettings: CoreSettings
    energyManagement: EnergyManagement
    optimization: Optimization
    assetManagement: AssetManagement
    advancedOptions: AdvancedOptions


class ModelConfigRequest(BaseModel):
    """Request model for saving configuration"""
    projectPath: str = Field(..., description="Project folder path")
    scenarioName: str = Field(..., description="Scenario name")
    baseYear: str = Field(..., description="Base year for simulation")
    lastUpdated: str = Field(..., description="Last update timestamp")
    configuration: Configuration


class RunModelRequest(BaseModel):
    """Request model for running PyPSA model"""
    projectPath: str = Field(..., description="Project folder path")
    scenarioName: str = Field(..., description="Scenario name")


# ============================================================================
# LOGGING UTILITIES
# ============================================================================

class StreamingLogger:
    """Custom logger that captures logs for streaming to frontend"""

    def __init__(self):
        self.logs = []
        self.string_io = StringIO()

    def log_buffer(self, log_entry: str):
        """Add a pre-formatted log entry directly to buffers (used by solver capture)"""
        self.logs.append(log_entry)
        current_log_buffer.append(log_entry)
        # Don't call print() here to avoid recursion when stdout is captured

    def log(self, level: str, message: str):
        """Add a log entry"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}"
        self.logs.append(log_entry)
        current_log_buffer.append(log_entry)
        print(log_entry)  # Also print to console

    def info(self, message: str):
        """Log info message"""
        self.log("INFO", message)

    def warning(self, message: str):
        """Log warning message"""
        self.log("WARNING", message)

    def error(self, message: str):
        """Log error message"""
        self.log("ERROR", message)

    def success(self, message: str):
        """Log success message"""
        self.log("SUCCESS", message)

    def get_logs(self) -> str:
        """Get all logs as formatted string"""
        return "\n".join(self.logs)


# ============================================================================
# CONFIGURATION MANAGEMENT
# ============================================================================

@router.post("/save-model-config")
async def save_model_config(request: ModelConfigRequest):
    """
    Save PyPSA model configuration to JSON file.

    Creates a configuration file in the project's inputs folder that will be
    used when running the PyPSA model.

    Args:
        request: Model configuration parameters

    Returns:
        dict: Success status and message

    Raises:
        HTTPException: 400 if validation fails, 500 on error
    """
    try:
        # Validate project path
        project_path = Path(request.projectPath)
        if not project_path.exists():
            raise HTTPException(
                status_code=400,
                detail=f"Project path does not exist: {request.projectPath}"
            )

        # Create inputs folder if it doesn't exist
        inputs_folder = project_path / "inputs"
        inputs_folder.mkdir(exist_ok=True)

        # Save configuration to JSON file
        config_file = inputs_folder / f"pypsa_config_{request.scenarioName}.json"

        config_data = {
            "projectPath": request.projectPath,
            "scenarioName": request.scenarioName,
            "baseYear": request.baseYear,
            "lastUpdated": request.lastUpdated,
            "configuration": request.configuration.dict()
        }

        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config_data, f, indent=2, ensure_ascii=False)

        logger.info(f"Configuration saved: {config_file}")

        return {
            "success": True,
            "message": f"Configuration saved successfully for scenario '{request.scenarioName}'",
            "configFile": str(config_file)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error saving configuration: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to save configuration: {str(e)}"
        )


# ============================================================================
# MODEL EXECUTION
# ============================================================================

@router.post("/run-pypsa-model")
async def run_pypsa_model(request: RunModelRequest):
    """
    Start PyPSA model execution.

    This endpoint initiates the model run in a background task. Use the
    /pypsa-model-progress endpoint to receive real-time logs via SSE.

    Args:
        request: Project path and scenario name

    Returns:
        dict: Success status and message

    Raises:
        HTTPException: 400 if validation fails, 409 if already running, 500 on error
    """
    global model_status, current_log_buffer

    try:
        # Check if model is already running
        if model_status["running"]:
            raise HTTPException(
                status_code=409,
                detail="A model is already running. Please wait for it to complete."
            )

        # Validate project path
        project_path = Path(request.projectPath)
        if not project_path.exists():
            raise HTTPException(
                status_code=400,
                detail=f"Project path does not exist: {request.projectPath}"
            )

        # Check if Master sheet exists (required for model execution)
        master_sheet = project_path / "inputs" / "Master sheet BAU.xlsx"
        if not master_sheet.exists():
            raise HTTPException(
                status_code=400,
                detail=f"Master sheet BAU.xlsx not found in inputs folder. Please ensure input data exists."
            )

        # Reset status
        model_status = {"running": True, "completed": False, "error": None}
        current_log_buffer = []

        # Start model execution in background
        asyncio.create_task(
            execute_pypsa_model(str(project_path), request.scenarioName)
        )

        return {
            "success": True,
            "message": f"Model execution started for scenario '{request.scenarioName}'",
            "projectPath": request.projectPath,
            "scenarioName": request.scenarioName
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting model: {str(e)}")
        model_status["running"] = False
        raise HTTPException(
            status_code=500,
            detail=f"Failed to start model execution: {str(e)}"
        )


async def execute_pypsa_model(project_folder: str, scenario_name: str):
    """
    Execute the PyPSA model with proper logging and process tracking.

    This function runs the complete PyPSA energy system model and streams
    logs to the frontend via the log buffer. Tracks PID for cancellation.

    Args:
        project_folder: Path to project folder
        scenario_name: Name of the scenario
    """
    global model_status
    stream_logger = StreamingLogger()

    try:
        stream_logger.info("="*80)
        stream_logger.info("PYPSA MODEL EXECUTION STARTED")
        stream_logger.info("="*80)
        stream_logger.info(f"Project Folder: {project_folder}")
        stream_logger.info(f"Scenario Name: {scenario_name}")

        # Track current process ID
        import os
        current_pid = os.getpid()
        model_status["pid"] = current_pid
        stream_logger.info(f"Process ID: {current_pid}")
        stream_logger.info("(Use 'Stop Model' button to cancel if needed)")
        stream_logger.info("")

        # Import PyPSA model function
        from models.pypsa_model_executor import run_pypsa_model_complete

        # Execute the model
        stream_logger.info("Initializing PyPSA model execution...")
        stream_logger.info("HiGHS solver logs will appear in real-time...")
        stream_logger.info("")

        result = await asyncio.to_thread(
            run_pypsa_model_complete,
            project_folder,
            scenario_name,
            stream_logger
        )

        if result["success"]:
            stream_logger.success("="*80)
            stream_logger.success("MODEL EXECUTION COMPLETED SUCCESSFULLY")
            stream_logger.success("="*80)
            stream_logger.info(f"Output folder: {result.get('output_folder', 'N/A')}")
            stream_logger.info(f"Total execution time: {result.get('execution_time', 'N/A')}")

            model_status["completed"] = True
            model_status["error"] = None
        else:
            stream_logger.error("="*80)
            stream_logger.error("MODEL EXECUTION FAILED")
            stream_logger.error("="*80)
            stream_logger.error(f"Error: {result.get('error', 'Unknown error')}")

            model_status["error"] = result.get('error', 'Unknown error')

    except Exception as e:
        error_msg = f"Fatal error during model execution: {str(e)}\n{traceback.format_exc()}"
        stream_logger.error(error_msg)
        model_status["error"] = str(e)

    finally:
        model_status["running"] = False
        model_status["pid"] = None
        stream_logger.info("Model execution finished.")


# ============================================================================
# SERVER-SENT EVENTS FOR REAL-TIME LOGS
# ============================================================================

@router.get("/pypsa-model-progress")
async def pypsa_model_progress():
    """
    Stream real-time model execution logs via Server-Sent Events (SSE).

    This endpoint provides live updates of the model execution progress.
    The frontend should connect to this endpoint using EventSource.

    Returns:
        StreamingResponse: SSE stream with log updates
    """
    async def event_generator():
        """Generate SSE events with log updates"""
        global model_status, current_log_buffer
        last_log_index = 0

        try:
            while True:
                # Check if there are new logs
                if last_log_index < len(current_log_buffer):
                    # Send new logs
                    new_logs = current_log_buffer[last_log_index:]
                    log_text = "\n".join(new_logs)

                    yield f"data: {json.dumps({'type': 'progress', 'log': log_text})}\n\n"

                    last_log_index = len(current_log_buffer)

                # Check if model has finished
                if not model_status["running"]:
                    if model_status["error"]:
                        yield f"data: {json.dumps({'type': 'end', 'status': 'failed', 'error': model_status['error']})}\n\n"
                    else:
                        yield f"data: {json.dumps({'type': 'end', 'status': 'completed'})}\n\n"
                    break

                # Wait before checking again
                await asyncio.sleep(0.5)

        except asyncio.CancelledError:
            logger.info("Client disconnected from progress stream")
        except Exception as e:
            logger.error(f"Error in event generator: {str(e)}")
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Disable buffering in nginx
        }
    )


@router.get("/model-status")
async def get_model_status():
    """
    Get current model execution status.

    Returns:
        dict: Current status (running, completed, error, pid)
    """
    return {
        "running": model_status["running"],
        "completed": model_status["completed"],
        "error": model_status["error"],
        "pid": model_status.get("pid"),
        "logCount": len(current_log_buffer)
    }


@router.post("/stop-pypsa-model")
async def stop_pypsa_model():
    """
    Stop/cancel the currently running PyPSA model.

    This endpoint gracefully terminates the running model execution by:
    1. Killing the process if PID is available
    2. Updating model status to cancelled
    3. Cleaning up resources

    Returns:
        dict: Success status and message

    Raises:
        HTTPException: 404 if no model is running, 500 on error
    """
    global model_status, current_log_buffer

    try:
        if not model_status["running"]:
            raise HTTPException(
                status_code=404,
                detail="No model is currently running"
            )

        pid = model_status.get("pid")

        if pid:
            try:
                # Try to terminate the process gracefully
                process = psutil.Process(pid)
                process.terminate()  # Send SIGTERM

                # Wait up to 5 seconds for graceful shutdown
                try:
                    process.wait(timeout=5)
                    logger.info(f"Process {pid} terminated gracefully")
                except psutil.TimeoutExpired:
                    # Force kill if it doesn't respond
                    process.kill()  # Send SIGKILL
                    logger.warning(f"Process {pid} force killed")

                current_log_buffer.append("\n⚠️  Model execution cancelled by user")

            except psutil.NoSuchProcess:
                logger.warning(f"Process {pid} not found (already terminated)")
            except Exception as e:
                logger.error(f"Error terminating process: {str(e)}")
                current_log_buffer.append(f"\n⚠️  Error stopping process: {str(e)}")

        # Update status
        model_status["running"] = False
        model_status["error"] = "Cancelled by user"
        model_status["pid"] = None

        current_log_buffer.append("\n❌ Model execution stopped")

        return {
            "success": True,
            "message": "Model execution cancelled successfully",
            "pid": pid
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error stopping model: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to stop model execution: {str(e)}"
        )


@router.get("/pypsa-solver-logs")
async def stream_solver_logs(
    projectPath: str = Query(..., description="Project folder path"),
    scenarioName: str = Query(..., description="Scenario name")
):
    """
    Stream HiGHS solver log file in real-time via SSE.

    This endpoint tails the solver log file and streams new lines as they're written.
    Crucial for monitoring long-running optimizations that can take hours.

    Args:
        projectPath: Project folder path
        scenarioName: Scenario name

    Returns:
        StreamingResponse: SSE stream with solver log updates
    """
    async def solver_log_generator():
        """Tail solver log file and stream updates"""
        try:
            # Find solver log file
            log_dir = Path(projectPath) / "results" / "pypsa_optimization" / scenarioName / "logs"
            solver_log_file = log_dir / f"{scenarioName}_solver.log"

            if not solver_log_file.exists():
                yield f"data: {json.dumps({'type': 'info', 'message': 'Waiting for solver to start...'})}\n\n"

                # Wait up to 30 seconds for log file to be created
                for _ in range(60):
                    await asyncio.sleep(0.5)
                    if solver_log_file.exists():
                        break
                else:
                    yield f"data: {json.dumps({'type': 'error', 'message': 'Solver log file not created'})}\n\n"
                    return

            # Tail the log file
            last_position = 0

            while model_status["running"]:
                try:
                    with open(solver_log_file, 'r') as f:
                        f.seek(last_position)
                        new_lines = f.read()

                        if new_lines:
                            yield f"data: {json.dumps({'type': 'log', 'content': new_lines})}\n\n"
                            last_position = f.tell()

                except Exception as e:
                    logger.error(f"Error reading solver log: {str(e)}")

                await asyncio.sleep(1)  # Check for updates every second

            # Final read after model stops
            try:
                with open(solver_log_file, 'r') as f:
                    f.seek(last_position)
                    final_lines = f.read()
                    if final_lines:
                        yield f"data: {json.dumps({'type': 'log', 'content': final_lines})}\n\n"
            except:
                pass

            yield f"data: {json.dumps({'type': 'end'})}\n\n"

        except asyncio.CancelledError:
            logger.info("Solver log stream cancelled")
        except Exception as e:
            logger.error(f"Error in solver log generator: {str(e)}")
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        solver_log_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


# ============================================================================
# CONFIGURATION RETRIEVAL AND APPLICATION
# ============================================================================

@router.get("/pypsa/settings")
async def get_pypsa_settings(projectPath: str = Query(..., description="Project root path")):
    """
    Retrieve PyPSA settings from template Excel file.

    This endpoint reads the PyPSA settings template file and returns all
    configuration tables for the user to select from.

    Args:
        projectPath: Project folder path

    Returns:
        dict: Settings tables with headers and data
    """
    try:
        # Look for settings template file in inputs folder
        project_path = Path(projectPath)
        inputs_folder = project_path / "inputs"

        # Try to find PyPSA settings template file
        settings_files = list(inputs_folder.glob("*settings*.xlsx")) + \
                        list(inputs_folder.glob("*template*.xlsx")) + \
                        list(inputs_folder.glob("PyPSA*.xlsx"))

        if not settings_files:
            # Return empty settings if no template found
            logger.warning(f"No PyPSA settings template found in {inputs_folder}")
            return {
                "success": True,
                "tables": {},
                "message": "No settings template file found"
            }

        # Use the first matching file
        settings_file = settings_files[0]
        logger.info(f"Reading PyPSA settings from: {settings_file}")

        # Read all sheets from Excel file
        import openpyxl
        workbook = openpyxl.load_workbook(settings_file, read_only=True, data_only=True)

        tables = {}

        for sheet_name in workbook.sheetnames:
            # Skip sheets that start with ~ (hidden sheets)
            if sheet_name.startswith('~'):
                continue

            worksheet = workbook[sheet_name]

            # Read headers (first row)
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            headers = [h for h in headers if h is not None]

            if not headers:
                continue

            # Read data rows
            data_rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Filter out completely empty rows
                if not any(cell is not None for cell in row):
                    continue

                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                    else:
                        row_dict[header] = None

                data_rows.append(row_dict)

            if data_rows:
                tables[sheet_name] = {
                    "headers": headers,
                    "data": data_rows
                }

        workbook.close()

        return {
            "success": True,
            "tables": tables,
            "file": str(settings_file),
            "sheet_count": len(tables)
        }

    except Exception as error:
        logger.error(f"Error reading PyPSA settings: {error}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to read PyPSA settings: {str(error)}"
        )


@router.post("/pypsa/apply-configuration")
async def apply_configuration(
    projectPath: str = Body(..., description="Project folder path"),
    scenarioName: str = Body(..., description="Scenario name")
):
    """
    Apply PyPSA configuration for a scenario.

    This endpoint validates the project and scenario, and prepares the
    configuration for model execution.

    Args:
        projectPath: Project folder path
        scenarioName: Name of the scenario

    Returns:
        dict: Success status and message
    """
    try:
        # Validate project path
        project_path = Path(projectPath)
        if not project_path.exists():
            raise HTTPException(
                status_code=400,
                detail=f"Project path does not exist: {projectPath}"
            )

        # Create scenario folder if it doesn't exist
        scenario_folder = project_path / "results" / "pypsa_optimization" / scenarioName
        scenario_folder.mkdir(parents=True, exist_ok=True)

        logger.info(f"Configuration applied for scenario '{scenarioName}'")

        return {
            "success": True,
            "message": f"Configuration applied successfully for scenario '{scenarioName}'",
            "scenarioName": scenarioName,
            "scenarioPath": str(scenario_folder)
        }

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error applying configuration: {error}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to apply configuration: {str(error)}"
        )
