"""
Demand Projection Page - COMPLETE IMPLEMENTATION
Full feature parity with React DemandProjection.jsx

Features:
- Dual view mode (Consolidated vs Sector-specific)
- Backend integration (sectors, data, color config)
- 5 tabs per view (Data Table, Area Chart, Stacked Bar, Line Chart, Correlation)
- Unit conversion (MWh, kWh, GWh, TWh)
- State persistence
- Configure Forecast modal
- Real-time SSE progress tracking
"""

from dash import html, dcc, callback, Input, Output, State, ALL, MATCH, callback_context, no_update
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
import pandas as pd
import numpy as np
import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.local_service import service as api
from utils.state_manager import StateManager, ConversionFactors, safe_numeric, safe_multiply


def layout(active_project=None):
    """
    Complete Demand Projection page with dual views
    """

    if not active_project:
        return dbc.Container([
            dbc.Alert([
                html.H4('⚠️ No Project Loaded', className='alert-heading'),
                html.P('Please load or create a project first to use demand forecasting.'),
                dbc.Button('Go to Projects', id={'type': 'nav-link', 'page': 'Load Project'},
                          color='primary')
            ], color='warning')
        ], className='p-4')

    return dbc.Container([
        # Configure Forecast Modal
        dbc.Modal([
            dbc.ModalHeader('⚙️ Configure Demand Forecast'),
            dbc.ModalBody([
                html.Div(id='configure-forecast-modal-content')
            ]),
            dbc.ModalFooter([
                dbc.Button('Cancel', id='cancel-forecast-btn', color='secondary', className='me-2'),
                dbc.Button('🚀 Start Forecasting', id='start-forecast-btn', color='success')
            ])
        ], id='configure-forecast-modal', is_open=False, size='xl'),

        # Progress Modal
        dbc.Modal([
            dbc.ModalHeader([
                html.Div([
                    html.H5('📊 Demand Forecasting in Progress', className='mb-0'),
                    dbc.Button('−', id='minimize-progress-modal', color='link',
                              className='ms-auto', style={'fontSize': '1.5rem'})
                ], style={'display': 'flex', 'width': '100%', 'alignItems': 'center'})
            ]),
            dbc.ModalBody([
                html.Div(id='forecast-progress-content')
            ]),
            dbc.ModalFooter([
                dbc.Button('Cancel Forecasting', id='cancel-forecasting-btn',
                          color='danger', outline=True, className='me-2'),
                dbc.Button('Close', id='close-progress-modal', color='secondary')
            ])
        ], id='forecast-progress-modal', is_open=False, size='lg'),

        # Floating process indicator (when modal minimized)
        html.Div(
            id='floating-process-indicator',
            style={'display': 'none'}
        ),

        # COMPACT HEADER - React-style single-row layout
        html.Div([
            # View Toggle (Compact inline buttons)
            html.Div([
                dbc.ButtonGroup([
                    dbc.Button(
                        [html.I(className='bi bi-bar-chart-fill me-1'), 'Consolidated View'],
                        id='consolidated-view-btn',
                        size='sm',
                        color='primary',
                        outline=False,
                        style={'fontSize': '0.875rem', 'fontWeight': '600'}
                    ),
                    dbc.Button(
                        [html.I(className='bi bi-sliders me-1'), 'Sector View'],
                        id='sector-view-btn',
                        size='sm',
                        color='primary',
                        outline=True,
                        style={'fontSize': '0.875rem', 'fontWeight': '600'}
                    )
                ], style={'backgroundColor': 'rgba(226, 232, 240, 0.7)', 'padding': '0.125rem',
                         'borderRadius': '0.375rem', 'border': '1px solid rgba(203, 213, 225, 0.5)'})
            ], style={'marginRight': '1rem'}),

            # Unit Selector (Compact inline)
            html.Div([
                html.Label('Unit', style={'fontSize': '0.875rem', 'fontWeight': '600',
                                         'color': '#475569', 'marginRight': '0.5rem',
                                         'marginBottom': '0'}),
                dcc.Dropdown(
                    id='consolidated-unit-selector',
                    options=[
                        {'label': 'kWh', 'value': 'kwh'},
                        {'label': 'MWh', 'value': 'mwh'},
                        {'label': 'GWh', 'value': 'gwh'},
                        {'label': 'TWh', 'value': 'twh'}
                    ],
                    value='mwh',
                    clearable=False,
                    style={'width': '110px', 'fontSize': '0.875rem'}
                )
            ], style={'display': 'flex', 'alignItems': 'center', 'marginRight': '1rem'}),

            # Configure Forecast Button (Compact)
            dbc.Button(
                [html.I(className='bi bi-gear-fill me-2'), 'Configure Forecast'],
                id='open-configure-forecast-btn',
                size='sm',
                color='indigo',
                style={'fontSize': '0.875rem', 'fontWeight': '600',
                      'backgroundColor': '#4f46e5', 'border': 'none',
                      'boxShadow': '0 1px 2px 0 rgba(0, 0, 0, 0.05)'}
            )
        ], style={'display': 'flex', 'justifyContent': 'center', 'alignItems': 'center',
                 'marginBottom': '1rem', 'padding': '0.5rem',
                 'backgroundColor': '#f8fafc', 'borderRadius': '0.5rem'}),

        # Consolidated View Content
        html.Div([

            # Tabs
            dbc.Tabs([
                dbc.Tab(
                    html.Div(id='consolidated-data-table-content', className='p-3'),
                    label='📋 Data Table',
                    tab_id='table'
                ),
                dbc.Tab(
                    html.Div(id='consolidated-area-chart-content', className='p-3'),
                    label='📊 Area Chart',
                    tab_id='area'
                ),
                dbc.Tab(
                    html.Div(id='consolidated-stacked-bar-content', className='p-3'),
                    label='📈 Stacked Bar Chart',
                    tab_id='stacked'
                ),
                dbc.Tab(
                    html.Div(id='consolidated-line-chart-content', className='p-3'),
                    label='📉 Line Chart',
                    tab_id='line'
                ),
            ], id='consolidated-tabs', active_tab='table', className='mb-3')
        ], id='consolidated-view-content', style={'display': 'block'}),

        # Sector-Specific View Content
        html.Div([
            # Horizontal Scrollable Sector Pills (React-style)
            html.Div([
                html.Div(
                    id='sector-pills-container',
                    style={
                        'overflowX': 'auto',
                        'whiteSpace': 'nowrap',
                        'padding': '0.5rem',
                        'backgroundColor': '#ffffff',
                        'borderRadius': '0.5rem',
                        'border': '1px solid rgba(226, 232, 240, 0.8)',
                        'boxShadow': '0 1px 2px 0 rgba(0, 0, 0, 0.05)',
                        'marginBottom': '0.75rem'
                    },
                    children=[]  # Will be populated by callback
                )
            ]),

            # Hidden dropdown for backward compatibility with callbacks
            dcc.Dropdown(
                id='sector-selector',
                options=[],
                value=None,
                clearable=False,
                style={'display': 'none'}
            ),
            dcc.Dropdown(
                id='sector-unit-selector',
                options=[
                    {'label': 'MWh', 'value': 'mwh'},
                    {'label': 'kWh', 'value': 'kwh'},
                    {'label': 'GWh', 'value': 'gwh'},
                    {'label': 'TWh', 'value': 'twh'}
                ],
                value='mwh',
                clearable=False,
                style={'display': 'none'}
            ),

            # Tabs
            dbc.Tabs([
                dbc.Tab(
                    html.Div(id='sector-data-table-content', className='p-3'),
                    label='📋 Data Table',
                    tab_id='table'
                ),
                dbc.Tab(
                    html.Div(id='sector-line-chart-content', className='p-3'),
                    label='📉 Line Chart',
                    tab_id='line'
                ),
                dbc.Tab(
                    html.Div(id='sector-correlation-content', className='p-3'),
                    label='🔗 Correlation',
                    tab_id='correlation'
                ),
            ], id='sector-tabs', active_tab='table', className='mb-3')
        ], id='sector-view-content', style={'display': 'none'}),

        # Loading overlay
        dcc.Loading(
            id='demand-projection-loading',
            type='circle',
            children=html.Div(id='loading-trigger')
        ),

        # Hidden stores for state management
        dcc.Store(id='demand-projection-state', storage_type='session', data=StateManager.create_demand_state()),
        dcc.Store(id='sectors-store', data=[]),
        dcc.Store(id='consolidated-data-store', data=None),
        dcc.Store(id='sector-data-store', data=None),
        dcc.Store(id='color-config-store', data={}),
        dcc.Store(id='forecast-process-state', data=None),

        # NEW: Stores for enhanced forecast configuration (React parity)
        dcc.Store(id='existing-scenarios-store', data=[]),  # List of existing scenario names
        dcc.Store(id='sector-metadata-store', data={}),  # Row counts, correlation data per sector

        # SSE control store for real-time forecast progress (React parity)
        dcc.Store(id='forecast-sse-control', data={'action': 'idle', 'url': ''}),

        # Interval for polling forecast progress
        dcc.Interval(id='forecast-progress-interval', interval=1000, disabled=True),  # 1 second

        # Hidden div for SSE clientside callback output
        html.Div(id='forecast-sse-status', style={'display': 'none'}),

        # Hidden divs for callback outputs (referenced by forecast_callbacks.py)
        html.Div(id='forecast-execution-status', style={'display': 'none'}),
        html.Div(id='sectors-list-preview', style={'display': 'none'}),

        # SSE Handler Script for Demand Forecasting (React parity)
        html.Script('''
            // Global EventSource for demand forecasting
            window.demandForecastEventSource = null;

            // Initialize dash_clientside namespace
            if (typeof window.dash_clientside === 'undefined') {
                window.dash_clientside = {};
            }
            if (typeof window.dash_clientside.forecast_sse === 'undefined') {
                window.dash_clientside.forecast_sse = {};
            }

            // Define clientside callback handler
            window.dash_clientside.forecast_sse.handle_sse = function(sse_control, current_state) {
                    if (!sse_control || !current_state) {
                        return [window.dash_clientside.no_update, 'idle'];
                    }

                    const action = sse_control.action;
                    const url = sse_control.url;

                    // Close existing connection if any
                    if (action === 'stop' || action === 'start') {
                        if (window.demandForecastEventSource) {
                            window.demandForecastEventSource.close();
                            window.demandForecastEventSource = null;
                        }
                    }

                    // Start new SSE connection
                    if (action === 'start' && url) {
                        const eventSource = new EventSource(url);
                        window.demandForecastEventSource = eventSource;

                        // Connection opened
                        eventSource.onopen = function() {
                            console.log('Forecast SSE connection opened');
                        };

                        // Progress event
                        eventSource.addEventListener('progress', function(event) {
                            try {
                                const data = JSON.parse(event.data);
                                const newState = {...current_state};

                                if (data.total_sectors) {
                                    newState.total_sectors = data.total_sectors;
                                }

                                newState.progress = data.progress || 0;
                                newState.message = data.message || 'Processing...';

                                if (data.step && data.message && data.sector) {
                                    newState.logs = [...(newState.logs || []), {
                                        type: 'progress',
                                        text: '(' + data.sector + ') - ' + data.message
                                    }];
                                }

                                // Update store
                                if (window.dash_clientside && window.dash_clientside.set_props) {
                                    window.dash_clientside.set_props('forecast-process-state', {data: newState});
                                }
                            } catch (e) {
                                console.error('Error parsing progress event:', e);
                            }
                        });

                        // Sector completed event
                        eventSource.addEventListener('sector_completed', function(event) {
                            try {
                                const data = JSON.parse(event.data);
                                const newState = {...current_state};

                                newState.current_sector = (newState.current_sector || 0) + 1;
                                newState.logs = [...(newState.logs || []), {
                                    type: 'success',
                                    text: 'Sector \\'' + data.sector + '\\' processed successfully.'
                                }];

                                // Update store
                                if (window.dash_clientside && window.dash_clientside.set_props) {
                                    window.dash_clientside.set_props('forecast-process-state', {data: newState});
                                }
                            } catch (e) {
                                console.error('Error parsing sector_completed event:', e);
                            }
                        });

                        // Sector failed event
                        eventSource.addEventListener('sector_failed', function(event) {
                            try {
                                const data = JSON.parse(event.data);
                                const newState = {...current_state};

                                newState.logs = [...(newState.logs || []), {
                                    type: 'error',
                                    text: 'Sector \\'' + data.sector + '\\' failed: ' + data.error
                                }];

                                // Update store
                                if (window.dash_clientside && window.dash_clientside.set_props) {
                                    window.dash_clientside.set_props('forecast-process-state', {data: newState});
                                }
                            } catch (e) {
                                console.error('Error parsing sector_failed event:', e);
                            }
                        });

                        // End event
                        eventSource.addEventListener('end', function(event) {
                            try {
                                const data = JSON.parse(event.data);
                                const newState = {...current_state};

                                if (data.status === 'completed') {
                                    newState.status = 'completed';
                                    newState.progress = 100;
                                    newState.message = 'Forecast process finished!';

                                    if (data.result) {
                                        const result = data.result;
                                        const successful = result.successful_sectors || 0;
                                        const failed = result.failed_sectors || 0;
                                        newState.logs = [...(newState.logs || []), {
                                            type: 'success',
                                            text: '✅ Forecast completed: ' + successful + ' sectors successful, ' + failed + ' failed.'
                                        }];
                                    } else {
                                        newState.logs = [...(newState.logs || []), {
                                            type: 'success',
                                            text: '✅ Forecast process completed successfully.'
                                        }];
                                    }
                                } else {
                                    newState.status = 'failed';
                                    newState.logs = [...(newState.logs || []), {
                                        type: 'error',
                                        text: '❌ Forecast failed: ' + (data.error || 'An unknown error occurred.')
                                    }];
                                }

                                eventSource.close();

                                // Update store
                                if (window.dash_clientside && window.dash_clientside.set_props) {
                                    window.dash_clientside.set_props('forecast-process-state', {data: newState});
                                }
                            } catch (e) {
                                console.error('Error parsing end event:', e);
                            }
                        });

                        // Error handler
                        eventSource.onerror = function() {
                            console.error('Forecast SSE connection error');
                            eventSource.close();
                            window.demandForecastEventSource = null;
                        };

                        return [window.dash_clientside.no_update, 'connected'];
                    }

                    return [window.dash_clientside.no_update, 'idle'];
            };
        ''')

    ], fluid=True, style={'padding': '2rem'})


# ============================================================================
# SSE CLIENTSIDE CALLBACK REGISTRATION
# ============================================================================

from dash import clientside_callback, ClientsideFunction

# Register clientside callback for forecast SSE handling
clientside_callback(
    ClientsideFunction(
        namespace='forecast_sse',
        function_name='handle_sse'
    ),
    Output('forecast-process-state', 'data', allow_duplicate=True),
    Output('forecast-sse-status', 'children'),
    Input('forecast-sse-control', 'data'),
    State('forecast-process-state', 'data'),
    prevent_initial_call=True
)


# ============================================================================
# CALLBACKS - VIEW TOGGLING
# ============================================================================

@callback(
    Output('consolidated-view-content', 'style'),
    Output('sector-view-content', 'style'),
    Output('consolidated-view-btn', 'outline'),
    Output('sector-view-btn', 'outline'),
    Output('demand-projection-state', 'data', allow_duplicate=True),
    Input('consolidated-view-btn', 'n_clicks'),
    Input('sector-view-btn', 'n_clicks'),
    State('demand-projection-state', 'data'),
    prevent_initial_call=True
)
def toggle_view_mode(consolidated_clicks, sector_clicks, current_state):
    """Toggle between Consolidated and Sector-specific views"""
    ctx = callback_context
    if not ctx.triggered:
        return no_update, no_update, no_update, no_update, no_update

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]

    if button_id == 'consolidated-view-btn':
        # Show consolidated view
        updated_state = StateManager.merge_state(current_state, {'isConsolidated': True})
        return (
            {'display': 'block'},
            {'display': 'none'},
            False,  # Consolidated button solid
            True,   # Sector button outline
            updated_state
        )
    else:
        # Show sector view
        updated_state = StateManager.merge_state(current_state, {'isConsolidated': False})
        return (
            {'display': 'none'},
            {'display': 'block'},
            True,   # Consolidated button outline
            False,  # Sector button solid
            updated_state
        )


# ============================================================================
# CALLBACKS - BACKEND INTEGRATION
# ============================================================================

@callback(
    Output('sectors-store', 'data'),
    Output('sector-selector', 'options'),
    Output('sector-pills-container', 'children'),
    Output('color-config-store', 'data'),
    Input('active-project-store', 'data')
)
def load_project_sectors(active_project):
    """Load sectors and color configuration when project loads"""
    if not active_project or not active_project.get('path'):
        return [], [], [], {}

    try:
        # Fetch sectors
        sectors_response = api.get_sectors(active_project['path'])
        sectors = sectors_response.get('sectors', [])

        # Create dropdown options (for backward compatibility)
        sector_options = [{'label': sector, 'value': idx}
                         for idx, sector in enumerate(sectors)]

        # Create horizontal scrollable sector pills (React-style)
        sector_pills = []
        for idx, sector in enumerate(sectors):
            pill = dbc.Button(
                sector,
                id={'type': 'sector-pill', 'index': idx},
                size='sm',
                outline=True if idx != 0 else False,
                color='primary' if idx == 0 else 'primary',
                style={
                    'marginRight': '0.25rem',
                    'fontSize': '0.875rem',
                    'fontWeight': '600',
                    'whiteSpace': 'nowrap',
                    'border': '2px solid transparent' if idx == 0 else '2px solid #cbd5e1',
                    'backgroundColor': '#4f46e5' if idx == 0 else '#ffffff',
                    'color': '#ffffff' if idx == 0 else '#475569'
                },
                n_clicks=0
            )
            sector_pills.append(pill)

        # Fetch color configuration
        try:
            color_response = api.get_color_settings(active_project['path'])
            colors = color_response.get('colors', {})
        except:
            colors = {}

        return sectors, sector_options, sector_pills, colors

    except Exception as e:
        print(f"Error loading sectors: {e}")
        return [], [], [], {}


@callback(
    Output('consolidated-data-store', 'data'),
    Input('sectors-store', 'data'),
    Input('demand-projection-state', 'data'),
    State('active-project-store', 'data'),
    prevent_initial_call=True
)
def load_consolidated_data(sectors, state, active_project):
    """Load consolidated data when in consolidated view"""
    if not active_project or not sectors:
        return None

    if not state or not state.get('isConsolidated'):
        return no_update

    try:
        response = api.get_consolidated_electricity(
            active_project['path'],
            sectors=sectors
        )
        return response.get('data', [])
    except Exception as e:
        print(f"Error loading consolidated data: {e}")
        return None


@callback(
    Output('sector-selector', 'value', allow_duplicate=True),
    Output('sector-pills-container', 'children', allow_duplicate=True),
    Input({'type': 'sector-pill', 'index': ALL}, 'n_clicks'),
    State('sectors-store', 'data'),
    State('sector-selector', 'value'),
    prevent_initial_call=True
)
def handle_sector_pill_click(n_clicks_list, sectors, current_sector_idx):
    """Handle sector pill clicks and update styling"""
    ctx = callback_context
    if not ctx.triggered or not sectors:
        return no_update, no_update

    # Find which pill was clicked
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]
    if not triggered_id or triggered_id == '':
        return no_update, no_update

    # Parse the clicked sector index
    import json
    try:
        pill_id = json.loads(triggered_id)
        clicked_idx = pill_id['index']
    except:
        return no_update, no_update

    # Update sector pills with new active state
    sector_pills = []
    for idx, sector in enumerate(sectors):
        is_active = (idx == clicked_idx)
        pill = dbc.Button(
            sector,
            id={'type': 'sector-pill', 'index': idx},
            size='sm',
            outline=not is_active,
            color='primary',
            style={
                'marginRight': '0.25rem',
                'fontSize': '0.875rem',
                'fontWeight': '600',
                'whiteSpace': 'nowrap',
                'border': '2px solid transparent' if is_active else '2px solid #cbd5e1',
                'backgroundColor': '#4f46e5' if is_active else '#ffffff',
                'color': '#ffffff' if is_active else '#475569'
            },
            n_clicks=0
        )
        sector_pills.append(pill)

    return clicked_idx, sector_pills


@callback(
    Output('sector-data-store', 'data'),
    Input('sector-selector', 'value'),
    State('sectors-store', 'data'),
    State('active-project-store', 'data'),
    State('demand-projection-state', 'data'),
    prevent_initial_call=True
)
def load_sector_data(sector_idx, sectors, active_project, state):
    """Load sector-specific data"""
    if sector_idx is None or not sectors or not active_project:
        return None

    if state and state.get('isConsolidated'):
        return no_update

    try:
        sector_name = sectors[sector_idx]
        response = api.extract_sector_data(
            active_project['path'],
            sector_name
        )
        return response.get('data', [])
    except Exception as e:
        print(f"Error loading sector data: {e}")
        return None


# ============================================================================
# CALLBACKS - DATA TABLE TAB (Consolidated)
# ============================================================================

@callback(
    Output('consolidated-data-table-content', 'children'),
    Input('consolidated-data-store', 'data'),
    Input('consolidated-unit-selector', 'value'),
    State('sectors-store', 'data')
)
def render_consolidated_data_table(data, unit, sectors):
    """Render consolidated data table with sticky header and first column"""
    if not data or not sectors:
        return dbc.Alert('No data available. Please load project data.', color='info')

    try:
        # Convert data to DataFrame
        df = pd.DataFrame(data)

        # Apply unit conversion
        factor = ConversionFactors.FACTORS.get(unit, 1)
        numeric_cols = [col for col in df.columns if col != 'Year']

        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: safe_multiply(x, factor))

        # Create custom table with sticky header and first column
        # Table header
        header_cells = [
            html.Th(
                col,
                style={
                    'position': 'sticky' if i == 0 else 'static',
                    'left': '0' if i == 0 else 'auto',
                    'top': '0',
                    'zIndex': '20' if i == 0 else '10',
                    'backgroundColor': '#f8fafc',
                    'fontWeight': '600',
                    'borderBottom': '2px solid #e2e8f0',
                    'padding': '0.5rem',
                    'fontSize': '0.875rem',
                    'whiteSpace': 'nowrap'
                }
            )
            for i, col in enumerate(df.columns)
        ]

        # Table rows
        table_rows = []
        for idx, row in df.iterrows():
            cells = [
                html.Td(
                    f'{row[col]:.2f}' if isinstance(row[col], (int, float)) and col != 'Year' else str(row[col]),
                    style={
                        'position': 'sticky' if i == 0 else 'static',
                        'left': '0' if i == 0 else 'auto',
                        'zIndex': '10' if i == 0 else '1',
                        'backgroundColor': '#ffffff' if i == 0 else 'transparent',
                        'padding': '0.5rem',
                        'fontSize': '0.875rem',
                        'fontWeight': '700' if i == 0 else '600',
                        'whiteSpace': 'nowrap',
                        'borderBottom': '1px solid #f1f5f9'
                    },
                    className='table-hover'
                )
                for i, col in enumerate(df.columns)
            ]
            table_rows.append(html.Tr(cells))

        table = html.Table(
            [
                html.Thead(html.Tr(header_cells)),
                html.Tbody(table_rows)
            ],
            style={
                'width': '100%',
                'borderCollapse': 'collapse',
                'fontSize': '0.875rem'
            },
            className='table table-sm table-striped table-hover'
        )

        return html.Div([
            html.Div([
                html.H5(f'Consolidated Electricity Demand ({ConversionFactors.get_label(unit)})',
                       className='mb-3')
            ]),
            html.Div(
                table,
                style={
                    'overflowX': 'auto',
                    'overflowY': 'auto',
                    'maxHeight': '78vh',
                    'border': '1px solid #e2e8f0',
                    'borderRadius': '0.375rem'
                }
            ),
            html.P(
                f'Showing data in {ConversionFactors.get_label(unit)}. Total sectors: {len(sectors)}',
                className='text-muted mt-3 mb-0',
                style={'fontSize': '0.875rem'}
            )
        ])

    except Exception as e:
        return dbc.Alert(f'Error rendering table: {str(e)}', color='danger')


# ============================================================================
# CALLBACKS - AREA CHART TAB (Consolidated)
# ============================================================================

@callback(
    Output('consolidated-area-chart-content', 'children'),
    Input('consolidated-data-store', 'data'),
    Input('consolidated-unit-selector', 'value'),
    State('sectors-store', 'data'),
    State('color-config-store', 'data')
)
def render_consolidated_area_chart(data, unit, sectors, colors):
    """Render consolidated area chart (stacked)"""
    if not data or not sectors:
        return dbc.Alert('No data available for chart.', color='info')

    try:
        df = pd.DataFrame(data)

        # Apply unit conversion
        factor = ConversionFactors.FACTORS.get(unit, 1)

        # Create figure
        fig = go.Figure()

        # Add traces for each sector (stacked area)
        for sector in sectors:
            if sector in df.columns:
                fig.add_trace(go.Scatter(
                    x=df['Year'],
                    y=df[sector].apply(lambda x: safe_multiply(x, factor)),
                    name=sector,
                    mode='lines',
                    stackgroup='one',
                    fillcolor=colors.get(sector, '#ccc'),
                    line=dict(width=0.5, color=colors.get(sector, '#ccc')),
                    hovertemplate=f'{sector}<br>Year: %{{x}}<br>Demand: %{{y:.2f}} {ConversionFactors.get_label(unit)}<extra></extra>'
                ))

        fig.update_layout(
            title=f'Consolidated Demand Forecast - Stacked Area ({ConversionFactors.get_label(unit)})',
            xaxis_title='Year',
            yaxis_title=f'Electricity Demand ({ConversionFactors.get_label(unit)})',
            hovermode='x unified',
            legend=dict(orientation='h', y=-0.2),
            height=500,
            template='plotly_white'
        )

        return dcc.Graph(figure=fig, config={'displayModeBar': True})

    except Exception as e:
        return dbc.Alert(f'Error rendering chart: {str(e)}', color='danger')


# ============================================================================
# CALLBACKS - STACKED BAR CHART TAB (Consolidated)
# ============================================================================

@callback(
    Output('consolidated-stacked-bar-content', 'children'),
    Input('consolidated-data-store', 'data'),
    Input('consolidated-unit-selector', 'value'),
    State('sectors-store', 'data'),
    State('color-config-store', 'data')
)
def render_consolidated_stacked_bar(data, unit, sectors, colors):
    """Render consolidated stacked bar chart"""
    if not data or not sectors:
        return dbc.Alert('No data available for chart.', color='info')

    try:
        df = pd.DataFrame(data)

        # Apply unit conversion
        factor = ConversionFactors.FACTORS.get(unit, 1)

        # Create figure
        fig = go.Figure()

        # Add bars for each sector (stacked)
        for sector in sectors:
            if sector in df.columns:
                fig.add_trace(go.Bar(
                    x=df['Year'],
                    y=df[sector].apply(lambda x: safe_multiply(x, factor)),
                    name=sector,
                    marker_color=colors.get(sector, '#ccc'),
                    hovertemplate=f'{sector}<br>Year: %{{x}}<br>Demand: %{{y:.2f}} {ConversionFactors.get_label(unit)}<extra></extra>'
                ))

        fig.update_layout(
            title=f'Consolidated Demand Forecast - Stacked Bar ({ConversionFactors.get_label(unit)})',
            xaxis_title='Year',
            yaxis_title=f'Electricity Demand ({ConversionFactors.get_label(unit)})',
            barmode='stack',
            hovermode='x unified',
            legend=dict(orientation='h', y=-0.2),
            height=500,
            template='plotly_white'
        )

        return dcc.Graph(figure=fig, config={'displayModeBar': True})

    except Exception as e:
        return dbc.Alert(f'Error rendering chart: {str(e)}', color='danger')


# ============================================================================
# CALLBACKS - LINE CHART TAB (Consolidated)
# ============================================================================

@callback(
    Output('consolidated-line-chart-content', 'children'),
    Input('consolidated-data-store', 'data'),
    Input('consolidated-unit-selector', 'value'),
    State('sectors-store', 'data'),
    State('color-config-store', 'data')
)
def render_consolidated_line_chart(data, unit, sectors, colors):
    """Render consolidated line chart (all sectors as separate lines)"""
    if not data or not sectors:
        return dbc.Alert('No data available for chart.', color='info')

    try:
        df = pd.DataFrame(data)

        # Apply unit conversion
        factor = ConversionFactors.FACTORS.get(unit, 1)

        # Create figure
        fig = go.Figure()

        # Add line for each sector
        for sector in sectors:
            if sector in df.columns:
                fig.add_trace(go.Scatter(
                    x=df['Year'],
                    y=df[sector].apply(lambda x: safe_multiply(x, factor)),
                    name=sector,
                    mode='lines+markers',
                    line=dict(width=2, color=colors.get(sector, '#ccc')),
                    marker=dict(size=6),
                    hovertemplate=f'{sector}<br>Year: %{{x}}<br>Demand: %{{y:.2f}} {ConversionFactors.get_label(unit)}<extra></extra>'
                ))

        fig.update_layout(
            title=f'Consolidated Demand Forecast - Line Chart ({ConversionFactors.get_label(unit)})',
            xaxis_title='Year',
            yaxis_title=f'Electricity Demand ({ConversionFactors.get_label(unit)})',
            hovermode='x unified',
            legend=dict(orientation='h', y=-0.2),
            height=500,
            template='plotly_white'
        )

        return dcc.Graph(figure=fig, config={'displayModeBar': True})

    except Exception as e:
        return dbc.Alert(f'Error rendering chart: {str(e)}', color='danger')


# ============================================================================
# CALLBACKS - SECTOR DATA TABLE TAB
# ============================================================================

@callback(
    Output('sector-data-table-content', 'children'),
    Input('sector-data-store', 'data'),
    Input('sector-unit-selector', 'value'),
    State('sector-selector', 'value'),
    State('sectors-store', 'data')
)
def render_sector_data_table(data, unit, sector_idx, sectors):
    """Render sector-specific data table with sticky header and first column"""
    if not data or sector_idx is None or not sectors:
        return dbc.Alert('No data available. Please select a sector.', color='info')

    try:
        sector_name = sectors[sector_idx]

        # Convert data to DataFrame
        df = pd.DataFrame(data)

        # Apply unit conversion ONLY to Electricity columns
        factor = ConversionFactors.FACTORS.get(unit, 1)

        # Find electricity columns (case-insensitive, exact match or contains 'electricity')
        electricity_cols = []
        for col in df.columns:
            col_lower = str(col).lower()
            # Match: exact 'electricity' or contains 'electricity' (but not 'year')
            if col_lower == 'electricity' or ('electricity' in col_lower and 'year' not in col_lower):
                electricity_cols.append(col)

        # Apply conversion to all electricity columns
        for col in electricity_cols:
            # Convert to numeric first, handling non-numeric values
            df[col] = pd.to_numeric(df[col], errors='coerce')
            # Apply unit conversion factor
            df[col] = df[col] * factor  # Direct multiplication (NaN-safe)

        # Create custom table with sticky header and first column
        # Table header
        header_cells = [
            html.Th(
                col,
                style={
                    'position': 'sticky' if i == 0 else 'static',
                    'left': '0' if i == 0 else 'auto',
                    'top': '0',
                    'zIndex': '20' if i == 0 else '10',
                    'backgroundColor': '#f8fafc',
                    'fontWeight': '600',
                    'borderBottom': '2px solid #e2e8f0',
                    'padding': '0.5rem',
                    'fontSize': '0.875rem',
                    'whiteSpace': 'nowrap'
                }
            )
            for i, col in enumerate(df.columns)
        ]

        # Table rows
        table_rows = []
        for idx, row in df.iterrows():
            cells = []
            for i, col in enumerate(df.columns):
                value = row[col]
                # Format value: 2 decimals for floats, no decimals for Year
                if col.lower() in ['year']:
                    formatted_value = str(int(value)) if pd.notna(value) else ''
                elif pd.notna(value) and isinstance(value, (int, float)):
                    formatted_value = f'{value:.2f}'
                else:
                    formatted_value = str(value) if pd.notna(value) else ''

                cells.append(html.Td(
                    formatted_value,
                    style={
                        'position': 'sticky' if i == 0 else 'static',
                        'left': '0' if i == 0 else 'auto',
                        'zIndex': '10' if i == 0 else '1',
                        'backgroundColor': '#ffffff' if i == 0 else 'transparent',
                        'padding': '0.5rem',
                        'fontSize': '0.875rem',
                        'fontWeight': '700' if i == 0 else '600',
                        'whiteSpace': 'nowrap',
                        'borderBottom': '1px solid #f1f5f9'
                    }
                ))
            table_rows.append(html.Tr(cells))

        table = html.Table(
            [
                html.Thead(html.Tr(header_cells)),
                html.Tbody(table_rows)
            ],
            style={
                'width': '100%',
                'borderCollapse': 'collapse',
                'fontSize': '0.875rem'
            },
            className='table table-sm table-striped table-hover'
        )

        return html.Div([
            html.Div([
                html.H5(f'{sector_name} - Demand Data ({ConversionFactors.get_label(unit)})',
                       className='mb-3')
            ]),
            html.Div(
                table,
                style={
                    'overflowX': 'auto',
                    'overflowY': 'auto',
                    'maxHeight': '78vh',
                    'border': '1px solid #e2e8f0',
                    'borderRadius': '0.375rem'
                }
            ),
            html.P(
                f'Showing data in {ConversionFactors.get_label(unit)}',
                className='text-muted mt-3 mb-0',
                style={'fontSize': '0.875rem'}
            )
        ])

    except Exception as e:
        return dbc.Alert(f'Error rendering table: {str(e)}', color='danger')


# ============================================================================
# CALLBACKS - SECTOR LINE CHART TAB
# ============================================================================

@callback(
    Output('sector-line-chart-content', 'children'),
    Input('sector-data-store', 'data'),
    Input('sector-unit-selector', 'value'),
    State('sector-selector', 'value'),
    State('sectors-store', 'data'),
    State('color-config-store', 'data')
)
def render_sector_line_chart(data, unit, sector_idx, sectors, colors):
    """Render sector-specific line chart - ONLY Electricity data"""
    if not data or sector_idx is None or not sectors:
        return dbc.Alert('No data available for chart.', color='info')

    try:
        sector_name = sectors[sector_idx]
        df = pd.DataFrame(data)

        # Apply unit conversion ONLY to Electricity column
        factor = ConversionFactors.FACTORS.get(unit, 1)

        # Find Electricity column (case-insensitive)
        electricity_col = None
        for col in df.columns:
            if col.lower() == 'electricity':
                electricity_col = col
                break

        if not electricity_col:
            return dbc.Alert('No Electricity data found for this sector.', color='warning')

        # Get Year column (case-insensitive)
        year_col = 'Year' if 'Year' in df.columns else 'year'

        # Create figure with single line for Electricity
        fig = go.Figure()

        fig.add_trace(go.Scatter(
            x=df[year_col],
            y=df[electricity_col].apply(lambda x: safe_multiply(x, factor)),
            name='Electricity Demand',
            mode='lines+markers',
            line=dict(width=3, color=colors.get(sector_name, '#4f46e5')),
            marker=dict(size=8),
            hovertemplate=f'{sector_name}<br>Year: %{{x}}<br>Demand: %{{y:.2f}} {ConversionFactors.get_label(unit)}<extra></extra>'
        ))

        fig.update_layout(
            title=f'{sector_name} - Electricity Demand ({ConversionFactors.get_label(unit)})',
            xaxis_title='Year',
            yaxis_title=f'Electricity Demand ({ConversionFactors.get_label(unit)})',
            hovermode='x unified',
            height=500,
            template='plotly_white',
            showlegend=False
        )

        return dcc.Graph(figure=fig, config={'displayModeBar': True})

    except Exception as e:
        return dbc.Alert(f'Error rendering chart: {str(e)}', color='danger')


# ============================================================================
# CALLBACKS - SECTOR CORRELATION TAB
# ============================================================================

@callback(
    Output('sector-correlation-content', 'children'),
    Input('sector-data-store', 'data'),
    State('sector-selector', 'value'),
    State('sectors-store', 'data'),
    State('active-project-store', 'data')
)
def render_sector_correlation(data, sector_idx, sectors, active_project):
    """Render sector correlation analysis"""
    if not data or sector_idx is None or not sectors or not active_project:
        return dbc.Alert('No data available for correlation analysis.', color='info')

    try:
        sector_name = sectors[sector_idx]

        # Fetch correlation data from backend
        correlation_response = api.get_sector_correlation(
            active_project['path'],
            sector_name
        )

        correlation_matrix_raw = correlation_response.get('correlation_matrix_raw', {})
        drivers = correlation_response.get('drivers', [])

        if not correlation_matrix_raw or not drivers:
            return dbc.Alert('No correlation data available for this sector.', color='info')

        # Create correlation heatmap
        fig = go.Figure(data=go.Heatmap(
            z=correlation_matrix_raw.get('values', []),
            x=drivers,
            y=drivers,
            colorscale='RdBu',
            zmid=0,
            text=correlation_matrix_raw.get('values', []),
            texttemplate='%{text:.3f}',
            textfont={"size": 10},
            colorbar=dict(title='Correlation')
        ))

        fig.update_layout(
            title=f'{sector_name} - Correlation Matrix',
            xaxis_title='Drivers',
            yaxis_title='Drivers',
            height=600,
            template='plotly_white'
        )

        # Summary statistics
        summary_stats = html.Div([
            html.H6('Correlation Summary', className='mt-4 mb-3'),
            dbc.Table([
                html.Thead([
                    html.Tr([
                        html.Th('Driver Pair'),
                        html.Th('Correlation'),
                        html.Th('Strength')
                    ])
                ]),
                html.Tbody([
                    html.Tr([
                        html.Td(f"{pair['driver1']} - {pair['driver2']}"),
                        html.Td(f"{pair['value']:.3f}"),
                        html.Td(
                            html.Span(
                                pair['strength'],
                                className=f"badge bg-{pair['badge_color']}"
                            )
                        )
                    ]) for pair in correlation_response.get('top_correlations', [])
                ])
            ], bordered=True, striped=True, hover=True, size='sm')
        ])

        return html.Div([
            dcc.Graph(figure=fig, config={'displayModeBar': True}),
            summary_stats
        ])

    except Exception as e:
        # If correlation endpoint doesn't exist or fails, show graceful fallback
        return dbc.Alert([
            html.H5('Correlation Analysis', className='alert-heading'),
            html.P('Correlation analysis will be available once forecast models are run.'),
            html.Hr(),
            html.P('Run a forecast first using the "Configure Forecast" button.', className='mb-0')
        ], color='info')


# Initialize sector selector with first sector
@callback(
    Output('sector-selector', 'value'),
    Input('sector-selector', 'options'),
    State('sector-selector', 'value'),
    prevent_initial_call=True
)
def initialize_sector_selector(options, current_value):
    """Set first sector as default"""
    if options and current_value is None:
        return 0
    return no_update


# ============================================================================
# CALLBACKS - CONFIGURE FORECAST MODAL
# ============================================================================

# NEW: Fetch scenarios and sector metadata when opening configure modal
@callback(
    Output('existing-scenarios-store', 'data'),
    Output('sector-metadata-store', 'data'),
    Input('open-configure-forecast-btn', 'n_clicks'),
    State('active-project-store', 'data'),
    State('sectors-store', 'data'),
    prevent_initial_call=True
)
def fetch_configure_modal_data(n_clicks, active_project, sectors):
    """
    Fetch existing scenarios and sector metadata (correlation, row counts)
    This matches React's useEffect behavior when modal opens
    """
    if not n_clicks or not active_project or not sectors:
        return no_update, no_update

    try:
        # Fetch existing scenarios for duplicate check
        scenarios_response = api.get_scenarios(active_project['path'])
        existing_scenarios = scenarios_response.get('scenarios', [])

        # Validate which sectors have data (filters out empty sectors)
        validation_response = api.validate_sectors_with_data(active_project['path'], sectors)
        valid_sectors = validation_response.get('valid_sectors', sectors) if validation_response.get('success') else sectors
        invalid_sectors = validation_response.get('invalid_sectors', []) if validation_response.get('success') else []

        # Log validation results
        if invalid_sectors:
            logger.warning(f"Skipping {len(invalid_sectors)} sectors with no data: {[s['sector'] for s in invalid_sectors]}")

        # Fetch metadata for each VALID sector (correlation data and row counts)
        sector_metadata = {}
        for sector in valid_sectors:
            try:
                # Extract sector data to get row count
                data_response = api.extract_sector_data(active_project['path'], sector)
                if data_response.get('success'):
                    rows = data_response.get('data', [])
                    row_count = len(rows)

                    # Get correlation analysis for MLR parameters
                    corr_response = api.get_sector_correlation(active_project['path'], sector)
                    correlations = []
                    mlr_params = []

                    if corr_response.get('success'):
                        # Extract correlation with electricity (excluding Year, Electricity itself)
                        corr_matrix = corr_response.get('correlation_matrix', {})
                        electricity_corr = corr_matrix.get('Electricity', {})

                        # Create list of parameters sorted by correlation strength
                        for param, corr_value in electricity_corr.items():
                            if param.lower() not in ['year', 'electricity']:
                                correlations.append({
                                    'variable': param,
                                    'correlation': abs(corr_value)  # Use absolute value
                                })

                        # Sort by correlation strength (highest first)
                        correlations.sort(key=lambda x: x['correlation'], reverse=True)
                        mlr_params = [c['variable'] for c in correlations]

                    sector_metadata[sector] = {
                        'row_count': row_count,
                        'correlations': correlations,
                        'mlr_params': mlr_params,  # Available MLR parameters
                        'max_wam_window': max(3, row_count - 2)  # React formula: rowCount - 2
                    }
                else:
                    # Try to get available indicators from Economic_Indicators sheet
                    try:
                        import openpyxl
                        inputs_dir = os.path.join(active_project['path'], 'inputs')
                        excel_path = os.path.join(inputs_dir, 'input_demand_file.xlsx')
                        if os.path.exists(excel_path):
                            wb = openpyxl.load_workbook(excel_path, data_only=True)
                            econ_sheet = wb['Economic_Indicators'] if 'Economic_Indicators' in wb.sheetnames else None
                            if econ_sheet:
                                # Get column names excluding Year
                                headers = [cell.value for cell in next(econ_sheet.iter_rows(min_row=1, max_row=1)) if cell.value]
                                available_params = [h for h in headers if str(h).lower() not in ['year']]
                            else:
                                available_params = []  # No fallback indicators if sheet missing
                            wb.close()
                        else:
                            available_params = []
                    except:
                        available_params = []  # Empty list if error

                    # Default values if data extraction fails
                    sector_metadata[sector] = {
                        'row_count': 10,
                        'correlations': [],
                        'mlr_params': available_params,  # Dynamic fallback or empty
                        'max_wam_window': 8
                    }

            except Exception as e:
                print(f"Error fetching metadata for sector {sector}: {e}")
                # Try to get available indicators from Economic_Indicators sheet
                try:
                    import openpyxl
                    inputs_dir = os.path.join(active_project['path'], 'inputs')
                    excel_path = os.path.join(inputs_dir, 'input_demand_file.xlsx')
                    if os.path.exists(excel_path):
                        wb = openpyxl.load_workbook(excel_path, data_only=True)
                        econ_sheet = wb['Economic_Indicators'] if 'Economic_Indicators' in wb.sheetnames else None
                        if econ_sheet:
                            headers = [cell.value for cell in next(econ_sheet.iter_rows(min_row=1, max_row=1)) if cell.value]
                            available_params = [h for h in headers if str(h).lower() not in ['year']]
                        else:
                            available_params = []
                        wb.close()
                    else:
                        available_params = []
                except:
                    available_params = []

                sector_metadata[sector] = {
                    'row_count': 10,
                    'correlations': [],
                    'mlr_params': available_params,  # Dynamic fallback or empty
                    'max_wam_window': 8
                }

        return existing_scenarios, sector_metadata

    except Exception as e:
        print(f"Error fetching configure modal data: {e}")
        import traceback
        traceback.print_exc()
        return [], {}


@callback(
    Output('configure-forecast-modal', 'is_open'),
    Output('configure-forecast-modal-content', 'children'),
    Input('open-configure-forecast-btn', 'n_clicks'),
    Input('cancel-forecast-btn', 'n_clicks'),
    Input('start-forecast-btn', 'n_clicks'),
    State('configure-forecast-modal', 'is_open'),
    State('active-project-store', 'data'),
    State('sectors-store', 'data'),
    State('existing-scenarios-store', 'data'),
    State('sector-metadata-store', 'data'),
    prevent_initial_call=True
)
def toggle_configure_modal(open_clicks, cancel_clicks, start_clicks, is_open, active_project, sectors, existing_scenarios, sector_metadata):
    """Toggle configure forecast modal and render content"""
    ctx = callback_context
    if not ctx.triggered:
        return no_update, no_update

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]

    # Close modal if cancel or start button clicked
    if button_id in ['cancel-forecast-btn', 'start-forecast-btn']:
        return False, no_update

    # Open modal and render content
    if button_id == 'open-configure-forecast-btn':
        if not active_project or not sectors:
            return True, dbc.Alert('Please load a project first.', color='warning')

        # Initialize sector_metadata if None
        if not sector_metadata:
            sector_metadata = {}

        # Render configure form matching React version exactly
        modal_content = html.Div([
            # Section A: Basic Configuration (3-column grid)
            html.Div([
                html.H5('📋 Basic Configuration', className='mb-3', style={'fontWeight': '600', 'color': '#0f172a'}),
                dbc.Row([
                    # Scenario Name with duplicate warning
                    dbc.Col([
                        dbc.Label('Scenario Name *', className='fw-semibold', style={'fontSize': '0.875rem', 'color': '#475569'}),
                        dbc.Input(
                            id='forecast-scenario-name',
                            type='text',
                            placeholder='Project_Demand_V1',
                            value='Project_Demand_V1',
                            className='mb-1',
                            style={'fontSize': '0.875rem'}
                        ),
                        # Duplicate warning (hidden by default, shown via callback)
                        html.Div(
                            id='scenario-name-warning',
                            children=[],
                            className='mt-1',
                            style={'minHeight': '20px'}
                        )
                    ], width=4),

                    # Projection Year
                    dbc.Col([
                        dbc.Label('Projection Year *', className='fw-semibold', style={'fontSize': '0.875rem', 'color': '#475569'}),
                        dbc.Input(
                            id='forecast-target-year',
                            type='number',
                            placeholder='2050',
                            min=2025,
                            max=2100,
                            className='mb-3',
                            style={'fontSize': '0.875rem'}
                        )
                    ], width=4),

                    # Exclude COVID Years
                    dbc.Col([
                        dbc.Label('Data Options', className='fw-semibold', style={'fontSize': '0.875rem', 'color': '#475569'}),
                        dbc.Checklist(
                            id='forecast-exclude-covid',
                            options=[
                                {'label': ' Exclude COVID-19 Years (FY 2021-2023)', 'value': 'exclude_covid'}
                            ],
                            value=['exclude_covid'],  # Default: checked
                            className='mt-2',
                            style={'fontSize': '0.875rem'}
                        )
                    ], width=4)
                ], className='mb-4')
            ], className='p-4', style={
                'backgroundColor': '#f8fafc',
                'borderRadius': '0.75rem',
                'marginBottom': '1.5rem',
                'border': '1px solid #e2e8f0'
            }),

            # Section B: Sector-wise Forecast Configuration Table
            html.Div([
                html.H5('⚙️ Sector-wise Forecast Configuration', className='mb-3', style={'fontWeight': '600', 'color': '#0f172a'}),
                html.P('Configure forecasting models and parameters for each sector',
                       className='text-muted mb-3',
                       style={'fontSize': '0.875rem'}),

                # Warning about filtered sectors (if any)
                html.Div([
                    dbc.Alert([
                        html.I(className='bi bi-info-circle me-2'),
                        html.Span(f"Note: {len(sectors) - len(sector_metadata)} sectors have been filtered out due to missing or empty data. Only sectors with valid data are shown below.")
                    ], color='info', className='mb-3', style={'fontSize': '0.875rem'})
                ]) if len(sector_metadata) < len(sectors) else html.Div(),

                # Table
                html.Div([
                    # Table header
                    html.Div([
                        html.Div('Sector / Category', style={
                            'width': '20%',
                            'fontWeight': '600',
                            'padding': '0.875rem',
                            'fontSize': '0.875rem',
                            'color': '#475569',
                            'textTransform': 'uppercase',
                            'letterSpacing': '0.05em'
                        }),
                        html.Div('Forecasting Models', style={
                            'width': '25%',
                            'fontWeight': '600',
                            'padding': '0.875rem',
                            'fontSize': '0.875rem',
                            'color': '#475569',
                            'textTransform': 'uppercase',
                            'letterSpacing': '0.05em'
                        }),
                        html.Div('MLR Input Parameters', style={
                            'width': '40%',
                            'fontWeight': '600',
                            'padding': '0.875rem',
                            'fontSize': '0.875rem',
                            'color': '#475569',
                            'textTransform': 'uppercase',
                            'letterSpacing': '0.05em'
                        }),
                        html.Div('Years for WAM', style={
                            'width': '15%',
                            'fontWeight': '600',
                            'padding': '0.875rem',
                            'fontSize': '0.875rem',
                            'color': '#475569',
                            'textTransform': 'uppercase',
                            'letterSpacing': '0.05em'
                        })
                    ], style={
                        'display': 'flex',
                        'backgroundColor': '#f1f5f9',
                        'borderBottom': '2px solid #cbd5e1',
                        'borderRadius': '0.5rem 0.5rem 0 0'
                    }),

                    # Table rows - one per sector (only valid sectors with data)
                    html.Div([
                        html.Div([
                            # Column 1: Sector Name
                            html.Div(
                                sector,
                                style={
                                    'width': '20%',
                                    'padding': '0.875rem',
                                    'fontWeight': '600',
                                    'fontSize': '0.875rem',
                                    'color': '#0f172a'
                                }
                            ),

                            # Column 2: Forecasting Models (multi-select)
                            html.Div([
                                dbc.Checklist(
                                    id={'type': 'sector-models', 'sector': idx},
                                    options=[
                                        {'label': ' SLR', 'value': 'SLR'},
                                        {'label': ' MLR', 'value': 'MLR'},
                                        {'label': ' WAM', 'value': 'WAM'}
                                    ],
                                    value=['SLR', 'MLR', 'WAM'],  # Default: All selected
                                    inline=True,
                                    className='mb-0',
                                    style={'fontSize': '0.875rem'}
                                )
                            ], style={'width': '25%', 'padding': '0.875rem'}),

                            # Column 3: MLR Parameters (conditional dropdown) - DYNAMIC from correlation
                            html.Div(
                                dcc.Dropdown(
                                    id={'type': 'mlr-params', 'sector': idx},
                                    options=[
                                        {'label': param, 'value': param}
                                        for param in sector_metadata.get(sector, {}).get('mlr_params', [])
                                    ],
                                    value=sector_metadata.get(sector, {}).get('mlr_params', []),  # AUTO-SELECT all
                                    multi=True,
                                    placeholder='Select parameters...',
                                    style={'fontSize': '0.875rem'}
                                ),
                                style={'width': '40%', 'padding': '0.875rem'},
                                id={'type': 'mlr-params-container', 'sector': idx}
                            ),

                            # Column 4: WAM Years (conditional select) - DYNAMIC from row count
                            html.Div(
                                dcc.Dropdown(
                                    id={'type': 'wam-years', 'sector': idx},
                                    options=[
                                        {'label': str(i), 'value': i}
                                        for i in range(3, sector_metadata.get(sector, {}).get('max_wam_window', 10) + 1)
                                    ],
                                    value=3,
                                    clearable=False,
                                    style={'fontSize': '0.875rem', 'width': '90px'}
                                ),
                                style={'width': '15%', 'padding': '0.875rem'},
                                id={'type': 'wam-years-container', 'sector': idx}
                            )
                        ], style={
                            'display': 'flex',
                            'alignItems': 'center',
                            'borderBottom': '1px solid #f1f5f9',
                            'backgroundColor': '#ffffff' if idx % 2 == 0 else '#f8fafc',
                            'transition': 'background-color 0.15s ease'
                        })
                        for idx, sector in enumerate(sector_metadata.keys())  # ONLY show sectors with valid data
                    ], style={'maxHeight': '400px', 'overflowY': 'auto'})
                ], style={
                    'border': '1px solid #e2e8f0',
                    'borderRadius': '0.5rem',
                    'overflow': 'hidden',
                    'boxShadow': '0 1px 2px 0 rgba(0, 0, 0, 0.05)'
                })
            ], className='p-4', style={
                'backgroundColor': 'white',
                'borderRadius': '0.75rem',
                'border': '1px solid #e2e8f0'
            })
        ], style={'maxHeight': '70vh', 'overflowY': 'auto'})

        return True, modal_content

    return no_update, no_update


# NEW: Callback to show duplicate scenario warning (React parity)
@callback(
    Output('scenario-name-warning', 'children'),
    Input('forecast-scenario-name', 'value'),
    State('existing-scenarios-store', 'data'),
    prevent_initial_call=True
)
def check_scenario_name_duplicate(scenario_name, existing_scenarios):
    """
    Check if scenario name already exists and show warning
    Matches React behavior: warning only, doesn't block
    """
    if not scenario_name or not existing_scenarios:
        return []

    scenario_name_lower = scenario_name.strip().lower()
    is_duplicate = any(
        existing.lower() == scenario_name_lower
        for existing in existing_scenarios
    )

    if is_duplicate:
        return dbc.Alert([
            html.I(className='bi bi-exclamation-triangle me-2'),
            html.Span('This scenario name already exists. If you continue, the previous results will be replaced.')
        ], color='warning', className='py-2 px-3 mb-0', style={'fontSize': '0.8rem'})
    else:
        return html.Small(
            f'Default name "Project_Demand_V1" is pre-filled — you can rename if needed.',
            className='text-muted',
            style={'fontSize': '0.75rem'}
        )


# ============================================================================
# CALLBACKS - START FORECASTING
# ============================================================================

@callback(
    Output('forecast-progress-modal', 'is_open', allow_duplicate=True),
    Output('forecast-process-state', 'data', allow_duplicate=True),
    Output('forecast-progress-interval', 'disabled', allow_duplicate=True),
    Input('start-forecast-btn', 'n_clicks'),
    State('forecast-scenario-name', 'value'),
    State('forecast-target-year', 'value'),
    State('forecast-exclude-covid', 'value'),
    State({'type': 'sector-models', 'sector': ALL}, 'value'),
    State({'type': 'mlr-params', 'sector': ALL}, 'value'),
    State({'type': 'wam-years', 'sector': ALL}, 'value'),
    State('active-project-store', 'data'),
    State('sectors-store', 'data'),
    prevent_initial_call=True
)
def start_forecasting(n_clicks, scenario_name, target_year, exclude_covid,
                     sector_models_list, mlr_params_list, wam_years_list,
                     active_project, sectors):
    """Start forecasting process with real-time polling"""
    print("\n" + "="*80)
    print("[DEBUG] start_forecasting callback triggered")
    print(f"[DEBUG] scenario_name: {scenario_name}, target_year: {target_year}")
    print("="*80 + "\n")

    if not n_clicks:
        print("[DEBUG] No clicks, returning no_update")
        return no_update, no_update, no_update

    # Validation
    if not scenario_name or not target_year:
        print("[DEBUG] Missing scenario_name or target_year")
        return no_update, no_update, no_update

    try:
        # Prepare sector configurations
        sector_configs = []
        for idx, sector in enumerate(sectors):
            models = sector_models_list[idx] if idx < len(sector_models_list) else ['SLR', 'MLR', 'WAM']
            mlr_params = mlr_params_list[idx] if idx < len(mlr_params_list) else []
            wam_years = wam_years_list[idx] if idx < len(wam_years_list) else 3

            if models:  # Only include sectors with at least one model selected
                # CRITICAL FIX: Extract sector data before starting forecast (matching FastAPI)
                sector_data = []
                try:
                    print(f"[DEBUG] Extracting data for sector: {sector}")
                    sector_data_response = api.extract_sector_data(active_project['path'], sector)
                    print(f"[DEBUG] Sector '{sector}' response: success={sector_data_response.get('success')}, data_length={len(sector_data_response.get('data', []))}")

                    if sector_data_response.get('success'):
                        sector_data = sector_data_response.get('data', [])
                        if not sector_data:
                            print(f"[WARNING] Sector '{sector}' has empty data array despite success=True")
                    else:
                        error_msg = sector_data_response.get('error', 'Unknown error')
                        print(f"[ERROR] Failed to extract sector '{sector}': {error_msg}")

                except Exception as e:
                    print(f"[ERROR] Exception extracting data for sector {sector}: {e}")
                    import traceback
                    traceback.print_exc()

                # Skip sector if no data
                if not sector_data:
                    print(f"[WARNING] Skipping sector '{sector}' - no data available")
                    continue

                sector_config = {
                    'name': sector,
                    'models': models,
                    'data': sector_data  # CRITICAL: Include actual sector data
                }
                print(f"[DEBUG] Sector '{sector}' config ready with {len(sector_data)} data rows")

                # Add MLR parameters if MLR is selected
                if 'MLR' in models and mlr_params:
                    sector_config['mlr_parameters'] = mlr_params

                # Add WAM window if WAM is selected
                if 'WAM' in models:
                    sector_config['wam_window'] = wam_years

                sector_configs.append(sector_config)

        if not sector_configs:
            return no_update, no_update, no_update

        # Prepare forecast configuration matching React structure
        forecast_config = {
            'scenario_name': scenario_name.strip(),
            'target_year': int(target_year),
            'exclude_covid_years': 'exclude_covid' in (exclude_covid or []),
            'sectors': sector_configs
        }

        # Start forecast via API
        response = api.start_demand_forecast(
            active_project['path'],
            forecast_config
        )

        # Initialize process state for real-time tracking (matching React format)
        from datetime import datetime
        timestamp = datetime.now().strftime('%H:%M:%S')

        process_state = {
            'process_id': response.get('process_id'),
            'status': 'running',
            'progress': 0,
            'message': 'Starting forecast...',
            'total_sectors': len(sector_configs),
            'current_sector': 0,
            'logs': [{
                'type': 'info',
                'text': 'Forecast process started',
                'timestamp': timestamp
            }]
        }

        print(f"[DEBUG] Forecast started successfully, process_id: {response.get('process_id')}")
        print(f"[DEBUG] Opening modal and enabling interval polling")

        # Open progress modal and enable interval for polling
        return True, process_state, False  # False = enable interval

    except Exception as e:
        print(f"[ERROR] Error starting forecast: {e}")
        import traceback
        traceback.print_exc()
        return no_update, no_update, no_update


# ============================================================================
# POLLING CALLBACK - Real-time progress updates
# ============================================================================

@callback(
    Output('forecast-process-state', 'data', allow_duplicate=True),
    Output('forecast-progress-interval', 'disabled', allow_duplicate=True),
    Input('forecast-progress-interval', 'n_intervals'),
    State('forecast-process-state', 'data'),
    prevent_initial_call=True
)
def poll_forecast_progress(n_intervals, process_state):
    """Poll for forecast progress updates from SSE queue"""
    print(f"[DEBUG] poll_forecast_progress: n_intervals={n_intervals}, process_state={process_state}")

    if not process_state or not process_state.get('process_id'):
        print("[DEBUG] No process_state or process_id, disabling interval")
        return no_update, True  # Disable interval

    # Check if process is still running
    status = process_state.get('status', 'running')
    if status in ['completed', 'failed', 'cancelled']:
        print(f"[DEBUG] Process status is {status}, disabling interval")
        return no_update, True  # Disable interval

    try:
        # Import the SSE queue
        from services.local_service import forecast_sse_queue

        # Read all available events from queue (non-blocking)
        events = []
        while not forecast_sse_queue.empty():
            try:
                event = forecast_sse_queue.get_nowait()
                events.append(event)
                print(f"[DEBUG] Got SSE event: {event}")
            except:
                break

        if not events:
            print("[DEBUG] No new events in queue")
            return no_update, False  # Keep polling

        # Process events and update state (matching React format with timestamps)
        new_state = dict(process_state)
        from datetime import datetime

        for event in events:
            event_type = event.get('type')
            print(f"[DEBUG] Processing event type: {event_type}")
            timestamp = datetime.now().strftime('%H:%M:%S')

            if event_type == 'progress':
                new_state['progress'] = event.get('progress', 0)
                new_state['message'] = event.get('message', 'Processing...')
                new_state['sector'] = event.get('sector', '')

                # Add to logs with timestamp
                if 'logs' not in new_state:
                    new_state['logs'] = []

                sector = event.get('sector', 'Overall')
                message = event.get('message', '')
                if sector and message:
                    new_state['logs'].append({
                        'type': 'progress',
                        'text': f'({sector}) - {message}',
                        'timestamp': timestamp
                    })

            elif event_type == 'sector_completed':
                sector = event.get('sector', 'Unknown')
                if 'logs' not in new_state:
                    new_state['logs'] = []
                new_state['logs'].append({
                    'type': 'success',
                    'text': f"Sector '{sector}' processed successfully.",
                    'timestamp': timestamp
                })
                # Update current sector count
                new_state['current_sector'] = (new_state.get('current_sector', 0) + 1)

            elif event_type == 'sector_failed':
                sector = event.get('sector', 'Unknown')
                error = event.get('error', 'Unknown error')
                if 'logs' not in new_state:
                    new_state['logs'] = []
                new_state['logs'].append({
                    'type': 'error',
                    'text': f"Sector '{sector}' failed: {error}",
                    'timestamp': timestamp
                })

            elif event_type == 'end':
                status = event.get('status', 'completed')
                new_state['status'] = status
                new_state['progress'] = 100 if status == 'completed' else new_state.get('progress', 0)

                if 'logs' not in new_state:
                    new_state['logs'] = []

                if status == 'completed':
                    result = event.get('result', {})
                    successful = result.get('successful_sectors', 0)
                    failed = result.get('failed_sectors', 0)
                    new_state['logs'].append({
                        'type': 'success',
                        'text': f'✅ Forecast completed: {successful} sectors successful, {failed} failed.',
                        'timestamp': timestamp
                    })
                    return new_state, True  # Disable interval
                else:
                    error = event.get('error', 'Unknown error')
                    new_state['logs'].append({
                        'type': 'error',
                        'text': f'❌ Forecast failed: {error}',
                        'timestamp': timestamp
                    })
                    return new_state, True  # Disable interval

        print(f"[DEBUG] Returning updated state with {len(new_state.get('logs', []))} logs")
        return new_state, False  # Keep polling

    except Exception as e:
        print(f"[ERROR] Error in poll_forecast_progress: {e}")
        import traceback
        traceback.print_exc()
        return no_update, False  # Keep polling despite error


# ============================================================================
# CALLBACKS - RENDER PROGRESS CONTENT
# ============================================================================

@callback(
    Output('forecast-progress-content', 'children'),
    Input('forecast-process-state', 'data')
)
def render_forecast_progress(process_state):
    """
    Render forecast progress content in modal (matching React NotificationModal)
    React reference: frontend/src/components/common/NotificationModal.jsx
    """
    if not process_state:
        return html.Div([
            dbc.Spinner(color='primary', size='sm', spinner_class_name='me-2'),
            html.Span('Initializing forecast...', className='text-muted')
        ], style={'padding': '1rem', 'textAlign': 'center'})

    status = process_state.get('status', 'running')
    progress = process_state.get('progress', 0)
    message = process_state.get('message', 'Processing...')
    logs = process_state.get('logs', [])
    total_sectors = process_state.get('total_sectors', 0)
    current_sector = process_state.get('current_sector', 0)

    # Header with status badge (React-style)
    if status == 'running':
        status_icon = '🔄'
        status_color = 'primary'
        status_text = 'IN PROGRESS'
    elif status == 'completed':
        status_icon = '✅'
        status_color = 'success'
        status_text = 'COMPLETED'
    elif status == 'failed':
        status_icon = '❌'
        status_color = 'danger'
        status_text = 'FAILED'
    elif status == 'cancelled':
        status_icon = '⚠️'
        status_color = 'warning'
        status_text = 'CANCELLED'
    else:
        status_icon = 'ℹ️'
        status_color = 'info'
        status_text = 'PROCESSING'

    # React-style header with status and task counter
    header_section = html.Div([
        html.Div([
            html.Div([
                html.Span(status_icon, style={'fontSize': '1.25rem', 'marginRight': '0.5rem'}),
                html.Strong(status_text, style={'fontSize': '0.875rem', 'color': '#0f172a'})
            ], style={'display': 'flex', 'alignItems': 'center'}),
            html.Div([
                html.Span(f'{current_sector} / {total_sectors}',
                         style={'fontSize': '0.875rem', 'fontWeight': '600', 'color': '#64748b'}),
                html.Span(' sectors', style={'fontSize': '0.75rem', 'color': '#94a3b8', 'marginLeft': '0.25rem'})
            ], style={'display': 'flex', 'alignItems': 'center'})
        ], style={
            'display': 'flex',
            'justifyContent': 'space-between',
            'alignItems': 'center',
            'marginBottom': '1rem',
            'padding': '0.75rem 1rem',
            'backgroundColor': '#f8fafc',
            'borderRadius': '0.5rem',
            'border': '1px solid #e2e8f0'
        })
    ])

    # Progress bar with percentage (React-style)
    progress_section = html.Div([
        html.Div([
            html.Span(f'{int(progress)}%',
                     style={'fontSize': '0.75rem', 'fontWeight': '600', 'color': '#475569'}),
            html.Span(message,
                     style={'fontSize': '0.75rem', 'color': '#64748b', 'fontStyle': 'italic'})
        ], style={
            'display': 'flex',
            'justifyContent': 'space-between',
            'marginBottom': '0.5rem'
        }),
        dbc.Progress(
            value=progress,
            className='mb-3',
            color=status_color,
            striped=status == 'running',
            animated=status == 'running',
            style={'height': '8px', 'borderRadius': '4px'}
        )
    ])

    # Logs container with timestamps (React-style)
    log_items = []
    for log in logs:  # Show all logs (React doesn't limit to 20)
        log_type = log.get('type', 'info')
        log_text = log.get('text', '')
        timestamp = log.get('timestamp', '')

        # Icon and color mapping
        if log_type == 'error':
            icon = '❌'
            text_color = '#dc2626'
            bg_color = '#fef2f2'
        elif log_type == 'success':
            icon = '✅'
            text_color = '#16a34a'
            bg_color = '#f0fdf4'
        elif log_type == 'progress':
            icon = '🔄'
            text_color = '#3b82f6'
            bg_color = '#eff6ff'
        elif log_type == 'warning':
            icon = '⚠️'
            text_color = '#ea580c'
            bg_color = '#fff7ed'
        else:
            icon = 'ℹ️'
            text_color = '#64748b'
            bg_color = '#f8fafc'

        log_items.append(
            html.Div([
                # Timestamp
                html.Span(
                    timestamp,
                    style={
                        'fontSize': '0.6875rem',
                        'color': '#94a3b8',
                        'fontFamily': 'monospace',
                        'marginRight': '0.75rem',
                        'minWidth': '70px',
                        'display': 'inline-block'
                    }
                ),
                # Icon
                html.Span(icon, style={'marginRight': '0.5rem', 'fontSize': '0.875rem'}),
                # Log text
                html.Span(
                    log_text,
                    style={
                        'fontSize': '0.8125rem',
                        'color': text_color,
                        'lineHeight': '1.5'
                    }
                )
            ], style={
                'padding': '0.5rem 0.75rem',
                'marginBottom': '0.25rem',
                'backgroundColor': bg_color,
                'borderRadius': '0.375rem',
                'borderLeft': f'3px solid {text_color}',
                'display': 'flex',
                'alignItems': 'flex-start'
            })
        )

    # Logs container with React-style scrollbar
    logs_container = html.Div([
        html.Div([
            html.H6('Process Logs', style={
                'fontSize': '0.875rem',
                'fontWeight': '600',
                'color': '#0f172a',
                'marginBottom': '0.75rem'
            }),
            html.Span(f'{len(logs)} entries', style={
                'fontSize': '0.75rem',
                'color': '#94a3b8',
                'marginLeft': '0.5rem'
            })
        ], style={'display': 'flex', 'alignItems': 'center', 'marginBottom': '0.75rem'}),

        html.Div(
            log_items if log_items else [
                html.Div('No logs yet...', style={
                    'textAlign': 'center',
                    'color': '#94a3b8',
                    'fontSize': '0.875rem',
                    'padding': '2rem'
                })
            ],
            style={
                'maxHeight': '320px',
                'overflowY': 'auto',
                'padding': '0.5rem',
                'backgroundColor': '#ffffff',
                'borderRadius': '0.5rem',
                'border': '1px solid #e2e8f0',
                # Custom scrollbar styling
                'scrollbarWidth': 'thin',
                'scrollbarColor': '#cbd5e1 #f1f5f9'
            }
        )
    ])

    return html.Div([
        header_section,
        progress_section,
        logs_container
    ], style={'padding': '0.5rem'})


# ============================================================================
# CALLBACKS - PROGRESS MODAL CONTROLS
# ============================================================================

@callback(
    Output('forecast-progress-modal', 'is_open', allow_duplicate=True),
    Output('floating-process-indicator', 'style', allow_duplicate=True),
    Output('floating-process-indicator', 'children', allow_duplicate=True),
    Output('forecast-sse-control', 'data', allow_duplicate=True),
    Input('close-progress-modal', 'n_clicks'),
    Input('minimize-progress-modal', 'n_clicks'),
    State('forecast-process-state', 'data'),
    prevent_initial_call=True
)
def control_progress_modal(close_clicks, minimize_clicks, process_state):
    """Control progress modal visibility and floating indicator with SSE cleanup"""
    ctx = callback_context
    if not ctx.triggered:
        return no_update, no_update, no_update, no_update

    button_id = ctx.triggered[0]['prop_id'].split('.')[0]

    if button_id == 'close-progress-modal':
        # Close modal completely and stop SSE connection
        sse_control = {'action': 'stop', 'url': ''}
        return False, {'display': 'none'}, '', sse_control

    if button_id == 'minimize-progress-modal':
        # Minimize modal, show floating indicator (keep SSE running)
        if process_state and process_state.get('status') == 'running':
            floating_indicator = dbc.Card([
                dbc.CardBody([
                    html.Div([
                        html.Span('🔄 Forecasting in progress...', className='me-3'),
                        html.Span(f"{process_state.get('progress', 0)}%", className='badge bg-primary me-2'),
                        dbc.Button('Show', id='show-progress-modal', size='sm', color='link')
                    ], style={'display': 'flex', 'alignItems': 'center'})
                ], className='py-2 px-3')
            ], style={
                'position': 'fixed',
                'bottom': '20px',
                'right': '20px',
                'zIndex': '1050',
                'boxShadow': '0 4px 6px rgba(0,0,0,0.1)'
            })

            return False, {'display': 'block'}, floating_indicator, no_update

    return no_update, no_update, no_update, no_update


@callback(
    Output('forecast-progress-modal', 'is_open', allow_duplicate=True),
    Output('floating-process-indicator', 'style', allow_duplicate=True),
    Input('show-progress-modal', 'n_clicks'),
    prevent_initial_call=True
)
def show_progress_modal(n_clicks):
    """Show progress modal when floating indicator clicked"""
    if n_clicks:
        return True, {'display': 'none'}
    return no_update, no_update


# ============================================================================
# CALLBACKS - CANCEL FORECASTING
# ============================================================================

@callback(
    Output('forecast-progress-modal', 'is_open', allow_duplicate=True),
    Output('forecast-process-state', 'data', allow_duplicate=True),
    Output('forecast-progress-interval', 'disabled', allow_duplicate=True),
    Input('cancel-forecasting-btn', 'n_clicks'),
    State('forecast-process-state', 'data'),
    prevent_initial_call=True
)
def cancel_forecasting(n_clicks, process_state):
    """Cancel ongoing forecast process"""
    print("\n" + "="*80)
    print("[DEBUG] cancel_forecasting callback triggered")
    print(f"[DEBUG] n_clicks: {n_clicks}")
    print(f"[DEBUG] process_state: {process_state}")
    print("="*80 + "\n")

    if not n_clicks or not process_state:
        print("[DEBUG] Preventing update: n_clicks or process_state is None")
        raise PreventUpdate

    process_id = process_state.get('process_id')
    print(f"[DEBUG] Extracted process_id: {process_id}")

    if not process_id:
        print("[DEBUG] No process_id found in state")
        raise PreventUpdate

    try:
        print(f"[DEBUG] Calling api.cancel_forecast with process_id: {process_id}")
        # Call API to cancel process (new signature: only process_id)
        result = api.cancel_forecast(process_id)
        print(f"[DEBUG] Cancel forecast result: {result}")

        if result.get('success'):
            # Update state to cancelled
            updated_state = {
                **process_state,
                'status': 'cancelled'
            }
            print(f"[DEBUG] Cancellation successful, closing modal and updating state")
            print(f"[DEBUG] Updated state: {updated_state}")
            # Close modal, update state, stop interval
            return False, updated_state, True
        else:
            print(f"[DEBUG] Cancellation failed: {result.get('error')}")
            # Keep modal open on failure
            return no_update, no_update, no_update

    except Exception as e:
        print(f"[ERROR] Exception in cancel_forecasting callback: {e}")
        import traceback
        traceback.print_exc()
        return no_update, no_update, no_update


# ============================================================================
# UNIT CONVERSION STATE SYNC
# ============================================================================

@callback(
    Output('demand-projection-state', 'data', allow_duplicate=True),
    Input('consolidated-unit-selector', 'value'),
    Input('sector-unit-selector', 'value'),
    State('demand-projection-state', 'data'),
    prevent_initial_call=True
)
def sync_unit_state(consolidated_unit, sector_unit, current_state):
    """Sync unit selection to state"""
    ctx = callback_context
    if not ctx.triggered:
        return no_update

    trigger_id = ctx.triggered[0]['prop_id'].split('.')[0]

    if trigger_id == 'consolidated-unit-selector':
        return StateManager.merge_state(current_state, {
            'consolidated': {**current_state.get('consolidated', {}), 'selectedUnit': consolidated_unit}
        })
    elif trigger_id == 'sector-unit-selector':
        return StateManager.merge_state(current_state, {
            'sectorView': {**current_state.get('sectorView', {}), 'selectedUnit': sector_unit}
        })

    return no_update


# ============================================================================
# TAB STATE SYNC
# ============================================================================

@callback(
    Output('demand-projection-state', 'data', allow_duplicate=True),
    Input('consolidated-tabs', 'active_tab'),
    Input('sector-tabs', 'active_tab'),
    State('demand-projection-state', 'data'),
    prevent_initial_call=True
)
def sync_tab_state(consolidated_tab, sector_tab, current_state):
    """Sync active tab to state"""
    ctx = callback_context
    if not ctx.triggered:
        return no_update

    trigger_id = ctx.triggered[0]['prop_id'].split('.')[0]

    if trigger_id == 'consolidated-tabs':
        return StateManager.merge_state(current_state, {
            'consolidated': {**current_state.get('consolidated', {}), 'activeTab': consolidated_tab}
        })
    elif trigger_id == 'sector-tabs':
        return StateManager.merge_state(current_state, {
            'sectorView': {**current_state.get('sectorView', {}), 'activeTab': sector_tab}
        })

    return no_update
