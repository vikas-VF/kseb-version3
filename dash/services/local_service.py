"""
Local Service - Direct Business Logic Execution (No FastAPI)
Replaces api_client.py for fully self-contained Dash webapp
"""

import os
import json
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
import logging
import openpyxl
import subprocess
import threading
import queue
import time
import sys

# Add models directory to path for imports
models_path = os.path.join(os.path.dirname(__file__), '..', 'models')
if models_path not in sys.path:
    sys.path.insert(0, models_path)

# Add config and utils to path
config_path = os.path.join(os.path.dirname(__file__), '..', 'config')
utils_path = os.path.join(os.path.dirname(__file__), '..', 'utils')
if config_path not in sys.path:
    sys.path.insert(0, config_path)
if utils_path not in sys.path:
    sys.path.insert(0, utils_path)

# Import PyPSA network caching and analyzer
from network_cache import load_network_cached, get_cache_stats, invalidate_network_cache
from pypsa_analyzer import PyPSASingleNetworkAnalyzer

# Import model registry and validator
try:
    from model_registry import (
        export_model_registry,
        get_available_models_for_sector,
        calculate_wam_max_window,
        get_recommended_models,
        AVAILABLE_MODELS,
        DEFAULT_FORECAST_CONFIG
    )
    from forecast_validator import (
        validate_forecast_config,
        validate_sector_data as validate_sector_data_structure
    )
    MODEL_REGISTRY_AVAILABLE = True
except ImportError as e:
    logger.warning(f"Model registry/validator not available: {e}")
    MODEL_REGISTRY_AVAILABLE = False

logger = logging.getLogger(__name__)

# Global state for tracking forecast processes
forecast_processes = {}
forecast_sse_queue = queue.Queue()

# Global state for tracking PyPSA solver logs
pypsa_solver_sse_queue = queue.Queue()
pypsa_solver_processes = {}

# Global state for tracking load profile generation processes
profile_processes = {}
profile_sse_queue = queue.Queue()


# ==================== ERROR MESSAGE CATALOG ====================

ERROR_MESSAGES = {
    'project_not_found': {
        'message': 'Project folder not found',
        'hint': 'Please check the project path and try again. Ensure the folder exists and you have permission to access it.'
    },
    'excel_file_missing': {
        'message': 'Required Excel file not found',
        'hint': 'Make sure input_demand_file.xlsx exists in the inputs/ folder.'
    },
    'forecast_failed': {
        'message': 'Forecasting process failed',
        'hint': 'Check the data quality and try again. Ensure all required columns are present.'
    },
    'insufficient_data': {
        'message': 'Not enough data for forecasting',
        'hint': 'At least 5 years of historical data are required for reliable forecasting.'
    },
    'pypsa_network_error': {
        'message': 'PyPSA network file could not be loaded',
        'hint': 'The .nc file may be corrupted. Try running the optimization again.'
    },
    'process_timeout': {
        'message': 'Process exceeded maximum runtime',
        'hint': 'The operation took too long and was terminated. Try reducing the data size or simplifying the configuration.'
    },
    'process_not_found': {
        'message': 'Process not found',
        'hint': 'The process may have already completed or been cancelled.'
    },
    'cancellation_failed': {
        'message': 'Failed to cancel process',
        'hint': 'The process may have already completed. Please check the results.'
    }
}


def format_error(error_key: str, technical_details: str = None) -> Dict:
    """
    Format user-friendly error message

    Args:
        error_key: Key from ERROR_MESSAGES dict
        technical_details: Optional technical error details

    Returns:
        Dict with success=False, error message, hint, and optional technical details
    """
    error_info = ERROR_MESSAGES.get(error_key, {
        'message': 'An error occurred',
        'hint': 'Please try again. If the problem persists, contact support.'
    })

    return {
        'success': False,
        'error': error_info['message'],
        'hint': error_info['hint'],
        'technical': technical_details if logger.isEnabledFor(logging.DEBUG) else None
    }


def cleanup_process(process_id: str, process_type: str):
    """
    Clean up process state after completion/error/cancellation

    Args:
        process_id: Process identifier
        process_type: Type of process ('forecast', 'profile', 'pypsa')
    """
    global forecast_processes, profile_processes, pypsa_solver_processes

    try:
        if process_type == 'forecast':
            if process_id in forecast_processes:
                proc_info = forecast_processes[process_id]

                # Cancel watchdog timer if exists
                if 'timer' in proc_info and proc_info['timer']:
                    proc_info['timer'].cancel()

                # Remove from tracking
                del forecast_processes[process_id]
                logger.info(f"Cleaned up forecast process: {process_id}")

        elif process_type == 'profile':
            if process_id in profile_processes:
                proc_info = profile_processes[process_id]

                # Cancel watchdog timer if exists
                if 'timer' in proc_info and proc_info['timer']:
                    proc_info['timer'].cancel()

                del profile_processes[process_id]
                logger.info(f"Cleaned up profile process: {process_id}")

        elif process_type == 'pypsa':
            if process_id in pypsa_solver_processes:
                proc_info = pypsa_solver_processes[process_id]

                # Cancel watchdog timer if exists
                if 'timer' in proc_info and proc_info['timer']:
                    proc_info['timer'].cancel()

                del pypsa_solver_processes[process_id]
                logger.info(f"Cleaned up PyPSA process: {process_id}")

    except Exception as e:
        logger.error(f"Error during cleanup of {process_type} process {process_id}: {e}")


def create_process_with_timeout(cmd: list, process_id: str, process_type: str,
                                config: Dict, timeout_seconds: int = 3600) -> subprocess.Popen:
    """
    Create subprocess with watchdog timeout

    Args:
        cmd: Command list for Popen
        process_id: Unique process identifier
        process_type: 'forecast', 'profile', or 'pypsa'
        config: Process configuration dict
        timeout_seconds: Maximum runtime (default 3600 = 1 hour)

    Returns:
        subprocess.Popen object
    """
    global forecast_processes, profile_processes, pypsa_solver_processes

    # Start subprocess
    process = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        bufsize=1  # Line buffered
    )

    # Create watchdog timer
    def timeout_handler():
        """Handler called if process exceeds timeout"""
        if process.poll() is None:  # Still running
            logger.warning(f"Process {process_id} ({process_type}) timed out after {timeout_seconds}s, terminating...")

            try:
                process.kill()

                # Send timeout error to appropriate queue
                error_event = {
                    'type': 'error',
                    'status': 'timeout',
                    'message': f'Process exceeded timeout of {timeout_seconds} seconds and was terminated',
                    'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
                }

                if process_type == 'forecast':
                    forecast_sse_queue.put(error_event)
                elif process_type == 'pypsa':
                    pypsa_solver_sse_queue.put(error_event)

            except Exception as e:
                logger.error(f"Error in timeout handler: {e}")

    timer = threading.Timer(timeout_seconds, timeout_handler)
    timer.start()

    # Store process info
    proc_info = {
        'process': process,
        'timer': timer,
        'start_time': time.time(),
        'config': config,
        'timeout': timeout_seconds
    }

    if process_type == 'forecast':
        forecast_processes[process_id] = proc_info
    elif process_type == 'profile':
        profile_processes[process_id] = proc_info
    elif process_type == 'pypsa':
        pypsa_solver_processes[process_id] = proc_info

    return process


# ==================== EXCEL PROCESSING HELPER FUNCTIONS ====================

def find_sheet(workbook, sheet_name: str):
    """
    Find a sheet by case-insensitive name.

    Args:
        workbook: openpyxl workbook object
        sheet_name: Sheet name to find

    Returns:
        Worksheet if found, None otherwise
    """
    lower_case_name = sheet_name.lower()
    for name in workbook.sheetnames:
        if name.lower() == lower_case_name:
            return workbook[name]
    return None


def find_cell_position(worksheet, marker: str) -> Optional[Tuple[int, int]]:
    """
    Find a marker cell (like '~Consumption_Sectors' or '~Econometric_Parameters').

    Args:
        worksheet: openpyxl worksheet
        marker: Marker string to find (case-insensitive)

    Returns:
        Tuple of (row, col) if found, None otherwise
    """
    lower_marker = marker.lower()
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            if (isinstance(cell_value, str) and
                cell_value.strip().lower() == lower_marker):
                return (row_idx, col_idx)
    return None


def safe_float(value, default=0.0):
    """
    Safely convert value to float with fallback.

    Args:
        value: Value to convert
        default: Default value if conversion fails

    Returns:
        Float value or default
    """
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default


def safe_int(value, default=0):
    """
    Safely convert value to int with fallback.

    Args:
        value: Value to convert
        default: Default value if conversion fails

    Returns:
        Integer value or default
    """
    try:
        return int(float(value)) if value is not None else default
    except (ValueError, TypeError):
        return default


def is_solar_sector(sector_name: str) -> bool:
    """
    Check if a sector name contains 'solar' or 'Solar'.

    Args:
        sector_name: Name of the sector

    Returns:
        True if sector is a solar generation sector
    """
    return 'solar' in sector_name.lower()


def calculate_td_loss_percentage(target_year: int, loss_points: List[Dict]) -> float:
    """
    Calculate T&D loss percentage for a given year using linear interpolation.

    Args:
        target_year: Year to calculate loss for
        loss_points: List of dicts with 'year' and 'loss' keys

    Returns:
        Loss percentage as decimal (e.g., 0.10 for 10%)
    """
    DEFAULT_LOSS = 0.10  # 10%

    if not loss_points or len(loss_points) == 0:
        return DEFAULT_LOSS

    # Filter and sort points
    sorted_points = sorted(
        [p for p in loss_points if isinstance(p.get('year'), (int, float)) and isinstance(p.get('loss'), (int, float))],
        key=lambda p: p['year']
    )

    if len(sorted_points) == 0:
        return DEFAULT_LOSS
    if len(sorted_points) == 1:
        return sorted_points[0]['loss'] / 100

    first_point = sorted_points[0]
    last_point = sorted_points[-1]

    # Extrapolation
    if target_year <= first_point['year']:
        return first_point['loss'] / 100
    if target_year >= last_point['year']:
        return last_point['loss'] / 100

    # Interpolation
    for i in range(len(sorted_points) - 1):
        p1 = sorted_points[i]
        p2 = sorted_points[i + 1]

        if target_year >= p1['year'] and target_year <= p2['year']:
            if p2['year'] - p1['year'] == 0:
                return p1['loss'] / 100

            # Linear interpolation
            interpolated_loss = p1['loss'] + (target_year - p1['year']) * (p2['loss'] - p1['loss']) / (p2['year'] - p1['year'])
            return interpolated_loss / 100

    return DEFAULT_LOSS


class LocalService:
    """
    Local service that directly executes business logic without HTTP calls.
    Provides same interface as APIClient but uses direct function calls.
    """

    def __init__(self):
        pass

    # ==================== PROJECT MANAGEMENT ====================

    def create_project(self, name: str, location: str, description: str = '') -> Dict:
        """Create new project with folder structure"""
        try:
            import shutil
            from pathlib import Path

            project_path = os.path.join(location, name)

            # Create directory structure
            os.makedirs(project_path, exist_ok=True)
            os.makedirs(os.path.join(project_path, 'inputs'), exist_ok=True)
            os.makedirs(os.path.join(project_path, 'results', 'demand_forecasts'), exist_ok=True)
            os.makedirs(os.path.join(project_path, 'results', 'load_profiles'), exist_ok=True)
            os.makedirs(os.path.join(project_path, 'results', 'pypsa_optimization'), exist_ok=True)

            # Copy Excel template files to inputs folder (matching FastAPI behavior)
            template_dir = Path(__file__).parent.parent / 'input'
            inputs_dir = Path(project_path) / 'inputs'

            template_files = [
                'input_demand_file.xlsx',
                'load_curve_template.xlsx',
                'pypsa_input_template.xlsx'
            ]

            for template_file in template_files:
                src = template_dir / template_file
                dst = inputs_dir / template_file
                if src.exists():
                    shutil.copy2(src, dst)
                    logger.info(f"Copied template: {template_file} to {dst}")
                else:
                    logger.warning(f"Template file not found: {src}")

            # Create project.json
            project_meta = {
                'name': name,
                'path': project_path,
                'description': description,
                'created': pd.Timestamp.now().isoformat()
            }

            with open(os.path.join(project_path, 'project.json'), 'w') as f:
                json.dump(project_meta, f, indent=2)

            return {'success': True, 'project': project_meta}

        except Exception as e:
            logger.error(f"Error creating project: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def load_project(self, project_path: str) -> Dict:
        """Load existing project"""
        try:
            project_json = os.path.join(project_path, 'project.json')

            if not os.path.exists(project_json):
                return {'success': False, 'error': 'project.json not found'}

            with open(project_json, 'r') as f:
                project_meta = json.load(f)

            return {'success': True, 'project': project_meta}

        except Exception as e:
            logger.error(f"Error loading project: {e}")
            return {'success': False, 'error': str(e)}

    def check_directory(self, path: str) -> Dict:
        """Validate directory path"""
        exists = os.path.exists(path) and os.path.isdir(path)
        return {'exists': exists, 'is_directory': exists}

    # ==================== SECTORS ====================

    def get_sectors(self, project_path: str) -> Dict:
        """
        Get consumption sectors from Excel file using ~consumption_sectors marker.

        Reads the 'main' sheet and looks for the '~consumption_sectors' marker,
        then extracts all sector names listed below it (matching FastAPI logic).
        """
        try:
            # Look for input demand Excel file
            inputs_dir = os.path.join(project_path, 'inputs')
            excel_path = os.path.join(inputs_dir, 'input_demand_file.xlsx')

            if not os.path.exists(excel_path):
                logger.warning(f"input_demand_file.xlsx not found, using default sectors")
                return {'sectors': ['Residential', 'Commercial', 'Industrial', 'Agriculture', 'Public Lighting']}

            # Load workbook with openpyxl
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

            # Find 'main' sheet (case-insensitive)
            main_sheet = find_sheet(workbook, 'main')

            if not main_sheet:
                logger.warning("Sheet 'main' not found, using sheet-based sector detection")
                workbook.close()
                # Fallback to sheet names
                xls = pd.ExcelFile(excel_path)
                sectors = [sheet for sheet in xls.sheet_names
                          if sheet.lower() not in ['main', 'metadata', 'info', 'config', 'summary', 'economic_indicators']]
                return {'sectors': sectors}

            # Convert to list of rows
            rows = list(main_sheet.iter_rows(values_only=True))

            sectors = []
            start_index = -1

            # Find the marker '~consumption_sectors' (case-insensitive)
            for i, row in enumerate(rows):
                for cell_value in row:
                    if (isinstance(cell_value, str) and
                        cell_value.strip().lower() == '~consumption_sectors'):
                        start_index = i + 2  # Start reading 2 rows below the marker
                        break
                if start_index != -1:
                    break

            # Extract sectors starting from the marker
            if start_index != -1:
                for i in range(start_index, len(rows)):
                    cell = rows[i][0]  # First column
                    if not cell or str(cell).strip() == '':
                        break  # Stop at first empty cell

                    # Strip the "~" prefix if present
                    sector_name = str(cell).strip()
                    if sector_name.startswith('~'):
                        sector_name = sector_name[1:].strip()

                    sectors.append(sector_name)

            workbook.close()

            if not sectors:
                logger.warning("No sectors found under ~consumption_sectors marker, using defaults")
                return {'sectors': ['Residential', 'Commercial', 'Industrial', 'Agriculture', 'Public Lighting']}

            logger.info(f"Found {len(sectors)} sectors: {sectors}")
            return {'sectors': sectors}

        except Exception as e:
            logger.error(f"Error getting sectors: {e}")
            # Return default sectors
            return {'sectors': ['Residential', 'Commercial', 'Industrial', 'Agriculture', 'Public Lighting']}

    # ==================== EXCEL PARSING ====================

    def validate_sectors_with_data(self, project_path: str, sectors: list) -> Dict:
        """
        Validate which sectors have valid data (non-empty with Year and Electricity columns).
        Returns dict with valid and invalid sector lists.

        This helps filter out empty sectors before forecast configuration.
        """
        try:
            inputs_dir = os.path.join(project_path, 'inputs')
            excel_path = os.path.join(inputs_dir, 'input_demand_file.xlsx')

            if not os.path.exists(excel_path):
                return {
                    'success': False,
                    'error': 'input_demand_file.xlsx not found'
                }

            valid_sectors = []
            invalid_sectors = []

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            available_sheets = {sheet.lower(): sheet for sheet in workbook.sheetnames}

            for sector in sectors:
                sector_lower = sector.lower()

                # Check if sector sheet exists
                if sector_lower not in available_sheets:
                    invalid_sectors.append({
                        'sector': sector,
                        'reason': 'Sheet not found'
                    })
                    continue

                # Try to read the sheet
                try:
                    actual_sheet_name = available_sheets[sector_lower]
                    df = pd.read_excel(excel_path, sheet_name=actual_sheet_name)

                    # Validate: must have at least 2 rows and Year + Electricity columns
                    if df.empty or len(df) < 2:
                        invalid_sectors.append({
                            'sector': sector,
                            'reason': 'Empty or insufficient data (need at least 2 rows)'
                        })
                        continue

                    if 'Year' not in df.columns or 'Electricity' not in df.columns:
                        invalid_sectors.append({
                            'sector': sector,
                            'reason': 'Missing required columns (Year, Electricity)'
                        })
                        continue

                    # Check for non-null values
                    year_valid = df['Year'].notna().any()
                    elec_valid = df['Electricity'].notna().any()

                    if not year_valid or not elec_valid:
                        invalid_sectors.append({
                            'sector': sector,
                            'reason': 'No valid data in Year or Electricity column'
                        })
                        continue

                    # Sector is valid
                    valid_sectors.append(sector)

                except Exception as e:
                    invalid_sectors.append({
                        'sector': sector,
                        'reason': f'Error reading sheet: {str(e)}'
                    })

            workbook.close()

            logger.info(f"Sector validation: {len(valid_sectors)} valid, {len(invalid_sectors)} invalid")
            if invalid_sectors:
                logger.warning(f"Invalid sectors: {[s['sector'] for s in invalid_sectors]}")

            return {
                'success': True,
                'valid_sectors': valid_sectors,
                'invalid_sectors': invalid_sectors,
                'total_sectors': len(sectors),
                'valid_count': len(valid_sectors),
                'invalid_count': len(invalid_sectors)
            }

        except Exception as e:
            logger.error(f"Error validating sectors: {e}")
            return {
                'success': False,
                'error': str(e)
            }

    def extract_sector_data(self, project_path: str, sector: str) -> Dict:
        """
        Extract sector-specific data merged with economic indicators.

        Process (matching FastAPI logic):
        1. Read econometric parameters for the sector from 'main' sheet using ~Econometric_Parameters marker
        2. Extract Year and Electricity data from sector-specific sheet
        3. Merge with economic indicator values from 'Economic_Indicators' sheet
        """
        try:
            inputs_dir = os.path.join(project_path, 'inputs')
            excel_path = os.path.join(inputs_dir, 'input_demand_file.xlsx')

            if not os.path.exists(excel_path):
                return {'success': False, 'error': 'input_demand_file.xlsx not found'}

            # Load workbook with openpyxl
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            # Find sheets (case-insensitive)
            main_sheet = find_sheet(workbook, 'main')
            econ_sheet = find_sheet(workbook, 'Economic_Indicators')

            if not main_sheet:
                logger.warning("Sheet 'main' not found, returning sector data only")
                # Fallback to simple sector reading
                df = pd.read_excel(excel_path, sheet_name=sector)
                workbook.close()
                return {
                    'success': True,
                    'data': df.to_dict('records'),
                    'columns': df.columns.tolist()
                }

            if not econ_sheet:
                logger.warning("Sheet 'Economic_Indicators' not found, returning sector data only")
                # Fallback to simple sector reading
                df = pd.read_excel(excel_path, sheet_name=sector)
                workbook.close()
                return {
                    'success': True,
                    'data': df.to_dict('records'),
                    'columns': df.columns.tolist()
                }

            # 1. Get sector-specific economic parameters from Main sheet
            econ_param_marker = find_cell_position(main_sheet, '~Econometric_Parameters')

            if not econ_param_marker:
                logger.warning("Marker '~Econometric_Parameters' not found, returning sector data only")
                # Fallback to simple sector reading
                df = pd.read_excel(excel_path, sheet_name=sector)
                workbook.close()
                return {
                    'success': True,
                    'data': df.to_dict('records'),
                    'columns': df.columns.tolist()
                }

            marker_row, marker_col = econ_param_marker
            headers_row = marker_row + 1

            # Find the sector column
            sector_column = None
            max_col = main_sheet.max_column

            for col in range(marker_col, max_col + 1):
                cell_value = main_sheet.cell(row=headers_row, column=col).value
                if cell_value and str(cell_value).strip().lower() == sector.strip().lower():
                    sector_column = col
                    break

            if sector_column is None:
                logger.warning(f"Sector column '{sector}' not found under econometric parameters, returning sector data only")
                # Fallback to simple sector reading
                df = pd.read_excel(excel_path, sheet_name=sector)
                workbook.close()
                return {
                    'success': True,
                    'data': df.to_dict('records'),
                    'columns': df.columns.tolist()
                }

            # 2. Extract economic indicator names below this column
            indicators = []
            for row in range(headers_row + 1, main_sheet.max_row + 1):
                cell_value = main_sheet.cell(row=row, column=sector_column).value
                if not cell_value or cell_value == '':
                    break
                indicators.append(str(cell_value))

            # 3. Read sector sheet for Year & Electricity
            sector_sheet = find_sheet(workbook, sector)
            if not sector_sheet:
                workbook.close()
                return {'success': False, 'error': f"Sector sheet for '{sector}' not found"}

            # Convert sector sheet to list of dicts
            sector_data = []
            headers = [cell.value for cell in next(sector_sheet.iter_rows(min_row=1, max_row=1))]
            for row in sector_sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 0 and row[0] is not None:  # Skip empty rows
                    row_dict = dict(zip(headers, row))
                    sector_data.append(row_dict)

            # 4. Collect economic values from Economic_Indicators sheet
            econ_data = []
            econ_headers = [cell.value for cell in next(econ_sheet.iter_rows(min_row=1, max_row=1))]
            for row in econ_sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 0 and row[0] is not None:  # Skip empty rows
                    row_dict = dict(zip(econ_headers, row))
                    econ_data.append(row_dict)

            # 5. Merge data (case-insensitive column matching)
            merged = []
            for sector_row in sector_data:
                # Case-insensitive Year and Electricity extraction
                year = sector_row.get('Year') or sector_row.get('year')
                electricity = sector_row.get('Electricity') or sector_row.get('electricity')

                if year is None:
                    continue

                # Find matching economic data for this year
                econ_row = next(
                    (e for e in econ_data if (e.get('Year') or e.get('year')) == year),
                    {}
                )

                obj = {"Year": year, "Electricity": electricity}
                for key in indicators:
                    obj[key] = econ_row.get(key, None)

                merged.append(obj)

            workbook.close()

            if not merged:
                logger.warning(f"No merged data for sector {sector}")
                return {'success': False, 'error': 'No data found'}

            return {
                'success': True,
                'data': merged,
                'columns': list(merged[0].keys()) if merged else []
            }

        except Exception as e:
            logger.error(f"Error extracting sector data: {e}")
            return {'success': False, 'error': str(e)}

    def read_solar_share_data(self, project_path: str) -> Dict[str, float]:
        """
        Read solar share percentages for each sector from input_demand_file.xlsx.

        Looks for the ~Solar_share marker in the 'main' sheet and reads
        Sector and Percentage_share columns below it.

        Returns:
            Dictionary mapping sector name to percentage share (e.g., {"Agriculture": 5.5})
        """
        try:
            excel_path = os.path.join(project_path, 'inputs', 'input_demand_file.xlsx')

            if not os.path.exists(excel_path):
                logger.warning(f"input_demand_file.xlsx not found at: {excel_path}")
                return {}

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

            # Find 'main' sheet (case-insensitive)
            main_sheet = find_sheet(workbook, 'main')

            if not main_sheet:
                logger.warning("Sheet 'main' not found")
                workbook.close()
                return {}

            # Find ~Solar_share marker
            marker_row = None
            marker_col = None
            for row_idx, row in enumerate(main_sheet.iter_rows(values_only=True), start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if (isinstance(cell_value, str) and
                        cell_value.strip().lower() == '~solar_share'):
                        marker_row = row_idx
                        marker_col = col_idx
                        break
                if marker_row:
                    break

            if not marker_row:
                logger.info("Marker '~Solar_share' not found, returning empty dict")
                workbook.close()
                return {}

            # Read header row (should be 1 or 2 rows below marker)
            headers_row = marker_row + 1
            headers = [main_sheet.cell(row=headers_row, column=col).value
                      for col in range(1, main_sheet.max_column + 1)]

            # If first row after marker is blank, try next row
            if not any(headers):
                headers_row = marker_row + 2
                headers = [main_sheet.cell(row=headers_row, column=col).value
                          for col in range(1, main_sheet.max_column + 1)]

            # Find Sector and Percentage_share columns
            sector_col = None
            percentage_col = None

            for col_idx, header in enumerate(headers, start=1):
                if header:
                    header_lower = str(header).strip().lower()
                    if header_lower == 'sector':
                        sector_col = col_idx
                    elif 'percentage' in header_lower and 'share' in header_lower:
                        percentage_col = col_idx

            if not sector_col or not percentage_col:
                logger.warning(f"Required columns not found. Sector col: {sector_col}, Percentage col: {percentage_col}")
                workbook.close()
                return {}

            # Read data rows
            solar_shares = {}
            for row_idx in range(headers_row + 1, main_sheet.max_row + 1):
                sector_name = main_sheet.cell(row=row_idx, column=sector_col).value
                percentage_value = main_sheet.cell(row=row_idx, column=percentage_col).value

                if not sector_name or sector_name == '':
                    break  # Stop at first empty sector

                try:
                    percentage_float = float(percentage_value) if percentage_value else 0.0
                    solar_shares[str(sector_name).strip()] = percentage_float
                except (ValueError, TypeError):
                    logger.warning(f"Invalid percentage value for sector {sector_name}: {percentage_value}")
                    solar_shares[str(sector_name).strip()] = 0.0

            workbook.close()
            logger.info(f"Successfully loaded solar shares for {len(solar_shares)} sectors")
            return solar_shares

        except Exception as e:
            logger.error(f"Error reading solar share data: {e}")
            return {}

    # ==================== CONSOLIDATED VIEW ====================

    def get_consolidated_electricity(self, project_path: str, sectors: Optional[List[str]] = None) -> Dict:
        """
        Consolidate electricity consumption by sectors and years using openpyxl.

        Reads all sheets from input_demand_file.xlsx that contain 'Year' and 'Electricity' columns,
        then consolidates the data into a single table (matching FastAPI logic).
        """
        try:
            excel_path = os.path.join(project_path, 'inputs', 'input_demand_file.xlsx')

            if not os.path.exists(excel_path):
                return {'success': False, 'error': 'input_demand_file.xlsx not found'}

            # Get sectors if not provided
            if sectors is None:
                sectors_response = self.get_sectors(project_path)
                sectors = sectors_response.get('sectors', [])

            # Load workbook
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            year_wise = {}
            found_sectors = set()

            # Iterate through all sector sheets
            for sector_name in sectors:
                sector_sheet = find_sheet(workbook, sector_name)

                if not sector_sheet:
                    continue

                # Convert sheet to list of dicts
                headers = [cell.value for cell in next(sector_sheet.iter_rows(min_row=1, max_row=1))]
                data = []
                for row in sector_sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > 0 and row[0] is not None:
                        row_dict = dict(zip(headers, row))
                        data.append(row_dict)

                # Check if sheet has Year and Electricity columns (case-insensitive)
                has_year = any(h and str(h).lower() == 'year' for h in headers)
                has_electricity = any(h and str(h).lower() == 'electricity' for h in headers)

                if not has_year or not has_electricity:
                    continue

                sector = sector_name.strip()
                found_sectors.add(sector)

                # Process each row
                for row in data:
                    # Case-insensitive column matching
                    year = row.get('Year') or row.get('year')
                    electricity = row.get('Electricity') or row.get('electricity')

                    if not year or year == '':
                        continue

                    try:
                        numeric_year = safe_int(year)
                        if numeric_year == 0:
                            continue
                    except (ValueError, TypeError):
                        continue

                    if numeric_year not in year_wise:
                        year_wise[numeric_year] = {"Year": numeric_year}

                    year_wise[numeric_year][sector] = electricity

            workbook.close()

            # Maintain only valid sectors in the order received
            ordered_sectors = [s for s in sectors if s in found_sectors]

            # Format the output
            formatted_array = []
            for year in sorted(year_wise.keys()):
                result_row = {"Year": year}
                for sector in ordered_sectors:
                    result_row[sector] = year_wise[year].get(sector, '')
                formatted_array.append(result_row)

            if not formatted_array:
                return {'success': False, 'error': 'No valid data found'}

            return {
                'success': True,
                'data': formatted_array
            }

        except Exception as e:
            logger.error(f"Error getting consolidated data: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== COLOR SETTINGS ====================

    def get_color_settings(self, project_path: str) -> Dict:
        """Get color configuration - dynamic based on sectors in Excel"""
        try:
            color_file = os.path.join(project_path, 'color.json')

            if os.path.exists(color_file):
                with open(color_file, 'r') as f:
                    colors = json.load(f)
                return {'colors': colors}

            # Generate default colors dynamically based on sectors from Excel
            sectors_result = self.get_sectors(project_path)
            sectors = sectors_result.get('sectors', [])

            # Default color palette (professional colors for data visualization)
            default_palette = [
                '#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#ec4899',
                '#06b6d4', '#f43f5e', '#84cc16', '#6366f1', '#eab308',
                '#14b8a6', '#f97316', '#a855f7', '#22c55e', '#ef4444'
            ]

            default_colors = {}
            for idx, sector in enumerate(sectors):
                # Assign color from palette (cycle if more sectors than colors)
                default_colors[sector] = default_palette[idx % len(default_palette)]

            # Also add colors for forecast models
            default_colors['SLR'] = '#3b82f6'  # Blue
            default_colors['MLR'] = '#10b981'  # Green
            default_colors['WAM'] = '#f59e0b'  # Orange
            default_colors['Time Series'] = '#8b5cf6'  # Purple

            return {'colors': default_colors}

        except Exception as e:
            logger.error(f"Error getting colors: {e}")
            return {'colors': {}}

    def save_color_settings(self, project_path: str, colors: Dict) -> Dict:
        """Save color configuration"""
        try:
            color_file = os.path.join(project_path, 'color.json')

            with open(color_file, 'w') as f:
                json.dump(colors, f, indent=2)

            return {'success': True}

        except Exception as e:
            logger.error(f"Error saving colors: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== CORRELATION ====================

    def get_sector_correlation(self, project_path: str, sector: str) -> Dict:
        """Calculate correlation matrix for sector with economic indicators"""
        try:
            # Extract sector data (includes merged economic indicators)
            data_response = self.extract_sector_data(project_path, sector)

            if not data_response.get('success'):
                return {'success': False, 'error': 'Could not load sector data'}

            df = pd.DataFrame(data_response['data'])

            # Get numeric columns only (exclude Year)
            numeric_cols = [col for col in df.columns
                          if col.lower() not in ['year'] and
                          pd.api.types.is_numeric_dtype(df[col])]

            if len(numeric_cols) < 2:
                return {'success': False, 'error': 'Not enough numeric columns for correlation'}

            # Calculate correlation matrix
            corr_df = df[numeric_cols].corr()

            # Convert to format expected by heatmap
            drivers = corr_df.columns.tolist()
            values = corr_df.values.tolist()

            # Create driver-keyed correlation dictionary for easy access
            # Format: {'Electricity': {'GSDP': 0.928, 'Population': 0.85, ...}, ...}
            correlation_dict = {}
            for driver in drivers:
                correlation_dict[driver] = corr_df[driver].to_dict()

            # Extract top correlations (excluding self-correlations)
            top_correlations = []
            for i, driver1 in enumerate(drivers):
                for j, driver2 in enumerate(drivers):
                    if i < j:  # Only upper triangle (avoid duplicates)
                        corr_value = corr_df.iloc[i, j]

                        # Determine strength and badge color
                        abs_corr = abs(corr_value)
                        if abs_corr >= 0.7:
                            strength = 'Strong'
                            badge_color = 'success'
                        elif abs_corr >= 0.4:
                            strength = 'Moderate'
                            badge_color = 'warning'
                        else:
                            strength = 'Weak'
                            badge_color = 'secondary'

                        top_correlations.append({
                            'driver1': driver1,
                            'driver2': driver2,
                            'value': corr_value,
                            'strength': strength,
                            'badge_color': badge_color
                        })

            # Sort by absolute correlation value (strongest first)
            top_correlations.sort(key=lambda x: abs(x['value']), reverse=True)

            return {
                'success': True,
                'correlation_matrix': correlation_dict,  # NOW: Dict[driver, Dict[driver, correlation]]
                'correlation_matrix_raw': {
                    'values': values,
                    'drivers': drivers
                },
                'drivers': drivers,
                'top_correlations': top_correlations[:10]  # Top 10 correlations
            }

        except Exception as e:
            logger.error(f"Error calculating correlation: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== FORECASTING ====================

    def start_demand_forecast(self, project_path: str, config: Dict) -> Dict:
        """Start demand forecasting process with subprocess execution (matching FastAPI)"""
        try:
            logger.info(f"Starting demand forecast with config: {config}")

            # Create scenario results directory
            scenario_path = Path(project_path) / "results" / "demand_forecasts" / config['scenario_name']
            scenario_path.mkdir(parents=True, exist_ok=True)

            # Prepare configuration for Python script (matching FastAPI format)
            config_for_python = {
                "scenario_name": config['scenario_name'],
                "target_year": config.get('target_year', 2037),
                "exclude_covid": config.get('exclude_covid_years', False),
                "forecast_path": str(scenario_path),
                "sectors": {}
            }

            # Convert sectors to format expected by forecasting.py
            for sector in config.get('sectors', []):
                sector_name = sector.get('name', sector.get('sector_name', ''))
                config_for_python["sectors"][sector_name] = {
                    "enabled": True,
                    "models": sector.get('selected_methods', sector.get('models', [])),
                    "parameters": {
                        "MLR": {"independent_vars": sector.get('mlr_parameters', [])},
                        "WAM": {"window_size": sector.get('wam_window', 10)}
                    },
                    "data": sector.get('data', [])
                }

            # Write config to file
            config_path = scenario_path / "forecast_config.json"
            with open(config_path, 'w') as f:
                json.dump(config_for_python, f, indent=2)

            logger.info(f"Forecast configuration saved to: {config_path}")

            # Start subprocess in background thread
            process_id = f"forecast_{config['scenario_name']}"

            thread = threading.Thread(
                target=self._run_forecast_subprocess,
                args=(config_path, process_id, config['scenario_name'])
            )
            thread.daemon = True
            thread.start()

            # Track process
            global forecast_processes
            forecast_processes[process_id] = {
                'thread': thread,
                'status': 'running',
                'scenario': config['scenario_name'],
                'start_time': time.time()
            }

            logger.info(f"Forecast process started with ID: {process_id}")

            return {
                'success': True,
                'process_id': process_id,
                'message': 'Forecast process started successfully.'
            }

        except Exception as e:
            logger.error(f"Error starting forecast: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def _run_forecast_subprocess(self, config_path: Path, process_id: str, scenario_name: str):
        """
        Run forecasting subprocess and stream output to SSE queue.
        Matches FastAPI implementation in forecast_routes.py:182-290
        """
        global forecast_sse_queue, forecast_processes

        script_path = Path(__file__).parent.parent / "models" / "forecasting.py"
        logger.info(f"Forecast script path: {script_path}")

        try:
            # Verify script exists
            if not script_path.exists():
                error_msg = f"Forecasting script not found: {script_path}"
                logger.error(error_msg)
                forecast_sse_queue.put({
                    'type': 'end',
                    'status': 'failed',
                    'error': error_msg
                })
                forecast_processes[process_id]['status'] = 'failed'
                return

            # Verify config exists
            if not config_path.exists():
                error_msg = f"Config file not found: {config_path}"
                logger.error(error_msg)
                forecast_sse_queue.put({
                    'type': 'end',
                    'status': 'failed',
                    'error': error_msg
                })
                forecast_processes[process_id]['status'] = 'failed'
                return

            logger.info(f"Starting subprocess: python {script_path} --config {config_path}")

            # Start subprocess (matching FastAPI implementation)
            process = subprocess.Popen(
                ["python", str(script_path), "--config", str(config_path)],
                cwd=str(script_path.parent),  # Set working directory
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,  # Text mode for easier string handling
                bufsize=1   # Line buffered
            )

            logger.info(f"Subprocess started with PID: {process.pid}")
            forecast_processes[process_id]['pid'] = process.pid
            forecast_processes[process_id]['process'] = process  # Store process object for cancellation

            # Read stderr in a separate thread to capture errors
            def read_stderr():
                for line in iter(process.stderr.readline, ''):
                    line = line.strip()
                    if line:
                        logger.error(f"[Forecast STDERR]: {line}")
                        # Queue stderr as error logs
                        forecast_sse_queue.put({
                            'type': 'error',
                            'text': line,
                            'timestamp': time.strftime('%H:%M:%S')
                        })

            stderr_thread = threading.Thread(target=read_stderr)
            stderr_thread.daemon = True
            stderr_thread.start()

            # Read stdout line by line
            for line in iter(process.stdout.readline, ''):
                line = line.strip()
                if line:
                    logger.info(f"[Forecast STDOUT]: {line}")

                    # Parse progress lines (matching FastAPI format)
                    if line.startswith('PROGRESS:'):
                        try:
                            progress_data = json.loads(line[9:])  # Remove 'PROGRESS:' prefix
                            forecast_sse_queue.put(progress_data)
                            logger.debug(f"Progress event queued: {progress_data.get('type', 'unknown')}")
                        except json.JSONDecodeError as e:
                            logger.error(f"Failed to parse progress JSON: {e}")
                    else:
                        # Queue other output as info logs
                        forecast_sse_queue.put({
                            'type': 'log',
                            'text': line,
                            'timestamp': time.strftime('%H:%M:%S')
                        })

            # Wait for process completion
            process.wait()
            stderr_thread.join(timeout=5)  # Wait for stderr thread to finish

            logger.info(f"Forecast process completed with return code: {process.returncode}")

            # Send completion event
            if process.returncode == 0:
                forecast_sse_queue.put({
                    'type': 'end',
                    'status': 'completed',
                    'message': f'Forecast for scenario "{scenario_name}" completed successfully',
                    'scenario_name': scenario_name
                })
                forecast_processes[process_id]['status'] = 'completed'
            else:
                # Read stderr
                stderr_output = process.stderr.read()
                error_msg = f'Forecast process failed with code {process.returncode}'
                if stderr_output:
                    error_msg += f': {stderr_output[:500]}'  # Limit error message length

                logger.error(error_msg)
                forecast_sse_queue.put({
                    'type': 'end',
                    'status': 'failed',
                    'error': error_msg,
                    'scenario_name': scenario_name
                })
                forecast_processes[process_id]['status'] = 'failed'

        except Exception as e:
            logger.error(f"Forecast subprocess error: {e}")
            import traceback
            traceback.print_exc()
            forecast_sse_queue.put({
                'type': 'end',
                'status': 'failed',
                'error': str(e),
                'scenario_name': scenario_name
            })
            forecast_processes[process_id]['status'] = 'failed'

    def get_forecast_progress(self, project_path: str, process_id: str) -> Dict:
        """Get forecast progress from tracked processes"""
        global forecast_processes

        if process_id in forecast_processes:
            process_info = forecast_processes[process_id]
            return {
                'status': process_info.get('status', 'unknown'),
                'progress': 0 if process_info.get('status') == 'running' else 100,
                'current_task': f"Forecast {process_info.get('status', 'running')}",
                'logs': []
            }

        return {
            'status': 'not_found',
            'progress': 0,
            'current_task': 'Process not found',
            'logs': []
        }

    def get_forecast_status_url(self) -> str:
        """Get SSE URL for forecast progress (matching FastAPI)"""
        return '/api/forecast-progress'


    def get_generation_status_url(self) -> str:
        """Get SSE URL for load profile generation progress"""
        return '/api/generation-status'

    def get_solver_logs_url(self) -> str:
        """Get SSE URL for PyPSA solver logs"""
        return '/api/pypsa-solver-logs'

    def cancel_forecast(self, process_id: str) -> Dict:
        """
        Cancel forecasting process

        Args:
            process_id: Process identifier

        Returns:
            Dict with success status and message
        """
        global forecast_processes

        print("\n" + "="*80)
        print("[DEBUG] cancel_forecast method called")
        print(f"[DEBUG] process_id: {process_id}")
        print(f"[DEBUG] forecast_processes keys: {list(forecast_processes.keys())}")
        print("="*80 + "\n")

        if process_id not in forecast_processes:
            print(f"[DEBUG] Process ID {process_id} not found in forecast_processes")
            return format_error('process_not_found', f'Process ID: {process_id}')

        proc_info = forecast_processes[process_id]
        process = proc_info['process']
        print(f"[DEBUG] Found process info: {proc_info.keys()}")
        print(f"[DEBUG] Process object: {process}")
        print(f"[DEBUG] Process poll status: {process.poll()}")

        try:
            # Check if already finished
            if process.poll() is not None:
                # Already finished
                print(f"[DEBUG] Process already completed with poll={process.poll()}")
                cleanup_process(process_id, 'forecast')
                return {
                    'success': True,
                    'message': 'Process already completed'
                }

            logger.info(f"Cancelling forecast process: {process_id}")
            print(f"[DEBUG] Calling process.terminate() on PID: {process.pid}")

            # Try graceful termination first
            process.terminate()
            print(f"[DEBUG] process.terminate() called successfully")

            # Wait up to 5 seconds for graceful shutdown
            try:
                print(f"[DEBUG] Waiting up to 5 seconds for graceful shutdown...")
                process.wait(timeout=5)
                logger.info(f"Process {process_id} terminated gracefully")
                print(f"[DEBUG] Process terminated gracefully")
            except subprocess.TimeoutExpired:
                # Force kill if not terminated
                logger.warning(f"Process {process_id} did not terminate gracefully, forcing kill")
                print(f"[DEBUG] Graceful termination timed out, calling process.kill()")
                process.kill()
                process.wait()
                logger.info(f"Process {process_id} killed")
                print(f"[DEBUG] Process killed successfully")

            # Send cancellation event to SSE queue
            print(f"[DEBUG] Sending cancellation event to forecast_sse_queue")
            forecast_sse_queue.put({
                'type': 'end',
                'status': 'cancelled',
                'message': 'Forecasting cancelled by user',
                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
            })
            print(f"[DEBUG] SSE cancellation event sent")

            # Cleanup
            print(f"[DEBUG] Calling cleanup_process for {process_id}")
            cleanup_process(process_id, 'forecast')
            print(f"[DEBUG] cleanup_process completed")

            result = {
                'success': True,
                'message': 'Forecasting cancelled successfully'
            }
            print(f"[DEBUG] Returning success result: {result}")
            return result

        except Exception as e:
            logger.error(f"Error cancelling forecast: {e}")
            print(f"[ERROR] Exception in cancel_forecast: {e}")
            import traceback
            traceback.print_exc()
            return format_error('cancellation_failed', str(e))

    def cancel_profile_generation(self, process_id: str) -> Dict:
        """
        Cancel load profile generation process

        Args:
            process_id: Process identifier

        Returns:
            Dict with success status and message
        """
        global profile_processes

        print("\n" + "="*80)
        print("[DEBUG] cancel_profile_generation method called")
        print(f"[DEBUG] process_id: {process_id}")
        print(f"[DEBUG] profile_processes keys: {list(profile_processes.keys())}")
        print("="*80 + "\n")

        if process_id not in profile_processes:
            print(f"[DEBUG] Process ID {process_id} not found in profile_processes")
            return format_error('process_not_found', f'Process ID: {process_id}')

        proc_info = profile_processes[process_id]
        process = proc_info['process']
        print(f"[DEBUG] Found process info: {proc_info.keys()}")
        print(f"[DEBUG] Process poll status: {process.poll()}")

        try:
            # Check if already finished
            if process.poll() is not None:
                print(f"[DEBUG] Process already completed")
                cleanup_process(process_id, 'profile')
                return {
                    'success': True,
                    'message': 'Process already completed'
                }

            logger.info(f"Cancelling profile generation process: {process_id}")
            print(f"[DEBUG] Calling process.terminate() on PID: {process.pid}")

            # Try graceful termination
            process.terminate()
            print(f"[DEBUG] process.terminate() called successfully")

            try:
                print(f"[DEBUG] Waiting up to 5 seconds for graceful shutdown...")
                process.wait(timeout=5)
                logger.info(f"Process {process_id} terminated gracefully")
                print(f"[DEBUG] Process terminated gracefully")
            except subprocess.TimeoutExpired:
                print(f"[DEBUG] Graceful termination timed out, calling process.kill()")
                process.kill()
                process.wait()
                logger.info(f"Process {process_id} killed")
                print(f"[DEBUG] Process killed successfully")

            # Cleanup
            print(f"[DEBUG] Calling cleanup_process for {process_id}")
            cleanup_process(process_id, 'profile')
            print(f"[DEBUG] cleanup_process completed")

            result = {
                'success': True,
                'message': 'Profile generation cancelled successfully'
            }
            print(f"[DEBUG] Returning success result: {result}")
            return result

        except Exception as e:
            logger.error(f"Error cancelling profile generation: {e}")
            print(f"[ERROR] Exception in cancel_profile_generation: {e}")
            import traceback
            traceback.print_exc()
            return format_error('cancellation_failed', str(e))

    def cancel_pypsa_model(self, process_id: str) -> Dict:
        """
        Cancel PyPSA optimization process

        Args:
            process_id: Process identifier

        Returns:
            Dict with success status and message
        """
        global pypsa_solver_processes

        print("\n" + "="*80)
        print("[DEBUG] cancel_pypsa_model method called")
        print(f"[DEBUG] process_id: {process_id}")
        print(f"[DEBUG] pypsa_solver_processes keys: {list(pypsa_solver_processes.keys())}")
        print("="*80 + "\n")

        if process_id not in pypsa_solver_processes:
            print(f"[DEBUG] Process ID {process_id} not found in pypsa_solver_processes")
            return format_error('process_not_found', f'Process ID: {process_id}')

        proc_info = pypsa_solver_processes[process_id]
        process = proc_info['process']
        print(f"[DEBUG] Found process info: {proc_info.keys()}")
        print(f"[DEBUG] Process poll status: {process.poll()}")

        try:
            # Check if already finished
            if process.poll() is not None:
                print(f"[DEBUG] Process already completed")
                cleanup_process(process_id, 'pypsa')
                return {
                    'success': True,
                    'message': 'Process already completed'
                }

            logger.info(f"Cancelling PyPSA optimization process: {process_id}")
            print(f"[DEBUG] Calling process.terminate() on PID: {process.pid}")

            # Try graceful termination
            process.terminate()
            print(f"[DEBUG] process.terminate() called successfully")

            try:
                print(f"[DEBUG] Waiting up to 5 seconds for graceful shutdown...")
                process.wait(timeout=5)
                logger.info(f"Process {process_id} terminated gracefully")
                print(f"[DEBUG] Process terminated gracefully")
            except subprocess.TimeoutExpired:
                print(f"[DEBUG] Graceful termination timed out, calling process.kill()")
                process.kill()
                process.wait()
                logger.info(f"Process {process_id} killed")
                print(f"[DEBUG] Process killed successfully")

            # Send cancellation event to SSE queue
            print(f"[DEBUG] Sending cancellation event to pypsa_solver_sse_queue")
            pypsa_solver_sse_queue.put({
                'type': 'end',
                'status': 'cancelled',
                'message': 'PyPSA optimization cancelled by user',
                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
            })
            print(f"[DEBUG] SSE cancellation event sent")

            # Cleanup
            print(f"[DEBUG] Calling cleanup_process for {process_id}")
            cleanup_process(process_id, 'pypsa')
            print(f"[DEBUG] cleanup_process completed")

            result = {
                'success': True,
                'message': 'PyPSA optimization cancelled successfully'
            }
            print(f"[DEBUG] Returning success result: {result}")
            return result

        except Exception as e:
            logger.error(f"Error cancelling PyPSA model: {e}")
            print(f"[ERROR] Exception in cancel_pypsa_model: {e}")
            import traceback
            traceback.print_exc()
            return format_error('cancellation_failed', str(e))

    # ==================== SCENARIOS ====================

    def get_scenarios(self, project_path: str) -> Dict:
        """List all forecast scenarios"""
        try:
            scenarios_dir = os.path.join(project_path, 'results', 'demand_forecasts')

            if not os.path.exists(scenarios_dir):
                return {'scenarios': []}

            scenarios = [d for d in os.listdir(scenarios_dir)
                        if os.path.isdir(os.path.join(scenarios_dir, d))]

            return {'scenarios': scenarios}

        except Exception as e:
            logger.error(f"Error getting scenarios: {e}")
            return {'scenarios': []}

    def get_scenario_sectors(self, project_path: str, scenario_name: str) -> Dict:
        """
        Get sectors in scenario by listing Excel files
        Each sector has its own .xlsx file (e.g., Agriculture.xlsx, Domestic_lt.xlsx)
        """
        try:
            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)

            if not os.path.exists(scenario_dir):
                logger.warning(f"Scenario directory not found: {scenario_dir}")
                return {'sectors': []}

            # List all .xlsx files in scenario directory
            files = os.listdir(scenario_dir)
            excel_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('~$')]

            # Extract sector names (remove .xlsx extension)
            sectors = [os.path.splitext(f)[0] for f in excel_files]

            # Filter out Consolidated_Results
            sectors = [s for s in sectors if s != 'Consolidated_Results']

            logger.info(f"Found {len(sectors)} sectors in scenario '{scenario_name}': {sectors}")
            return {'sectors': sectors}

        except Exception as e:
            logger.error(f"Error getting scenario sectors: {e}")
            return {'sectors': []}

    def get_sector_data(self, project_path: str, scenario_name: str, sector_name: str,
                        start_year: int = None, end_year: int = None) -> Dict:
        """
        Get forecast data for specific sector from individual Excel file
        Reads from {sector_name}.xlsx in the scenario directory

        Returns data in format expected by demand_visualization page:
        {
            'years': [2006, 2007, ...],
            'forecastStartYear': 2023,
            'models': {
                'Historical': [val1, val2, ...],
                'MLR': [val1, val2, ...],
                ...
            }
        }
        """
        try:
            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)
            excel_path = os.path.join(scenario_dir, f'{sector_name}.xlsx')

            if not os.path.exists(excel_path):
                logger.error(f"Sector file not found: {excel_path}")
                return {'success': False, 'error': f'Results file not found for sector {sector_name}'}

            # Read the 'Results' sheet which contains the forecast data
            df = pd.read_excel(excel_path, sheet_name='Results')

            # Filter by year range if provided
            if start_year is not None and 'Year' in df.columns:
                df = df[df['Year'] >= start_year]
            if end_year is not None and 'Year' in df.columns:
                df = df[df['Year'] <= end_year]

            # Get forecast start year from 'Inputs' sheet
            forecast_start_year = None
            try:
                inputs_df = pd.read_excel(excel_path, sheet_name='Inputs')
                if 'Year' in inputs_df.columns and 'Electricity' in inputs_df.columns:
                    # Get max year that has electricity data
                    valid_rows = inputs_df[inputs_df['Electricity'].notna()]
                    if not valid_rows.empty:
                        forecast_start_year = int(valid_rows['Year'].max())
            except Exception as e:
                logger.warning(f"Could not determine forecast start year: {e}")
                pass

            # Transform data to expected format
            years = df['Year'].tolist() if 'Year' in df.columns else []

            # Get all model columns (exclude 'Year' column)
            model_columns = [col for col in df.columns if col != 'Year']

            # Build models dict with arrays of values for each model
            models = {}
            for model in model_columns:
                models[model] = df[model].tolist()

            logger.info(f"Loaded {len(df)} rows for sector '{sector_name}' from {excel_path}")
            logger.info(f"Available models: {list(models.keys())}")

            return {
                'years': years,
                'forecastStartYear': forecast_start_year,
                'models': models
            }

        except Exception as e:
            logger.error(f"Error getting sector data for '{sector_name}': {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    # ==================== LOAD PROFILES ====================

    def get_load_profiles(self, project_path: str) -> Dict:

        """
        List all load profile Excel files in results/load_profiles.

        Returns:
            {'profiles': ['ProfileA', 'ProfileB', ...]}
        """
        try:
            profiles_dir = Path(project_path) / "results" / "load_profiles"

            if not profiles_dir.exists() or not profiles_dir.is_dir():
                logger.info(f"Load profiles directory not found: {profiles_dir}")
                return {'profiles': []}

            excel_exts = {'.xlsx', '.xls', '.xlsm', '.xlsb', '.csv'}
            profiles = []

            for p in sorted(profiles_dir.iterdir()):
                # Skip hidden files (Unix: startswith '.'), and directories
                if p.name.startswith('.') or p.is_dir():
                    continue

                # Accept only known spreadsheet extensions
                if p.suffix.lower() not in excel_exts:
                    continue

                # Use filename without extension as profile name
                profile_name = p.stem.strip()
                if profile_name:
                    profiles.append(profile_name)

            return {'profiles': profiles}

        except Exception as e:
            logger.error(f"Error listing load profiles in {project_path}: {e}")
            return {'profiles': []}


    def get_load_profiles_with_meta(self, project_path: str) -> Dict:
        """
        Alternative richer listing with metadata:
            {'profiles': [{'name': 'ProfileA', 'file': 'ProfileA.xlsx', 'modified': '2025-11-15T12:34:56'}, ...]}

        Useful if frontend wants labels, original filenames, or modification times.
        """
        try:
            profiles_dir = Path(project_path) / "results" / "load_profiles"
            if not profiles_dir.exists() or not profiles_dir.is_dir():
                return {'profiles': []}

            out = []
            for p in sorted(profiles_dir.iterdir()):
                if p.name.startswith('.') or not p.is_dir():  # Skip hidden files and regular files, only process directories
                    continue

                try:
                    mtime = p.stat().st_mtime
                    modified_iso = pd.to_datetime(mtime, unit='s').isoformat()
                except Exception:
                    modified_iso = None

                out.append({
                    'name': p.name,  # Directory name is the profile name
                    'file': p.name,  # Directory path is the profile path
                    'modified': modified_iso
                })

            return {'profiles': out}
        except Exception as e:
            logger.error(f"Error listing load profiles with meta in {project_path}: {e}")
            return {'profiles': []}
    def generate_profile(self, config: Dict) -> Dict:
        """Start load profile generation with subprocess execution"""
        try:
            project_path = config['project_path']
            profile_config = config['profile_configuration']
            profile_name = profile_config['general']['profile_name']

            logger.info(f"Starting profile generation for: {profile_name}")

            # Create profile results directory
            profile_dir = Path(project_path) / "results" / "load_profiles"
            profile_dir.mkdir(parents=True, exist_ok=True)

            # Save config for subprocess
            config_path = profile_dir / f"{profile_name}_config.json"
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)

            logger.info(f"Profile configuration saved to: {config_path}")

            # Start subprocess in background thread
            process_id = f"profile_{profile_name}"

            thread = threading.Thread(
                target=self._run_profile_subprocess,
                args=(config_path, process_id, profile_name)
            )
            thread.daemon = True
            thread.start()

            # Track process
            global profile_processes
            profile_processes[process_id] = {
                'thread': thread,
                'status': 'running',
                'profile_name': profile_name,
                'start_time': time.time()
            }

            logger.info(f"Profile generation process started with ID: {process_id}")

            return {
                'success': True,
                'process_id': process_id,
                'message': 'Profile generation process started successfully.'
            }

        except Exception as e:
            logger.error(f"Error starting profile generation: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def _run_profile_subprocess(self, config_path: Path, process_id: str, profile_name: str):
        """
        Run profile generation subprocess and stream output to SSE queue.
        Similar to _run_forecast_subprocess but for load profiles.
        """
        global profile_sse_queue, profile_processes

        script_path = Path(__file__).parent.parent / "models" / "load_profile_generation.py"
        logger.info(f"Profile generation script path: {script_path}")

        try:
            # Verify script exists
            if not script_path.exists():
                error_msg = f"Profile generation script not found: {script_path}"
                logger.error(error_msg)
                profile_sse_queue.put({
                    'type': 'end',
                    'status': 'failed',
                    'error': error_msg
                })
                profile_processes[process_id]['status'] = 'failed'
                return

            # Verify config exists
            if not config_path.exists():
                error_msg = f"Config file not found: {config_path}"
                logger.error(error_msg)
                profile_sse_queue.put({
                    'type': 'end',
                    'status': 'failed',
                    'error': error_msg
                })
                profile_processes[process_id]['status'] = 'failed'
                return

            # Read config to get the full payload
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            config_json = json.dumps(config_data)

            logger.info(f"Starting subprocess: python {script_path} --config <json>")

            # Start subprocess with proper encoding to fix charmap error
            process = subprocess.Popen(
                ["python", str(script_path), "--config", config_json],
                cwd=str(script_path.parent),
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',  # Force UTF-8 encoding
                errors='replace',  # Replace invalid characters instead of crashing
                bufsize=1
            )

            logger.info(f"Profile generation subprocess started with PID: {process.pid}")
            profile_processes[process_id]['pid'] = process.pid
            profile_processes[process_id]['process'] = process

            # Read stderr in a separate thread
            def read_stderr():
                try:
                    for line in iter(process.stderr.readline, ''):
                        line = line.strip()
                        if line:
                            logger.error(f"[Profile Generation STDERR]: {line}")
                            # Check if it's a PROGRESS line mistakenly sent to stderr
                            if line.startswith('PROGRESS:'):
                                try:
                                    progress_data = json.loads(line[9:])
                                    profile_sse_queue.put(progress_data)
                                except json.JSONDecodeError as e:
                                    logger.error(f"Failed to parse progress JSON from stderr: {e}")
                            else:
                                profile_sse_queue.put({
                                    'type': 'error',
                                    'text': line,
                                    'timestamp': time.strftime('%H:%M:%S')
                                })
                except Exception as e:
                    logger.error(f"Error reading profile generation stderr: {e}")

            stderr_thread = threading.Thread(target=read_stderr)
            stderr_thread.daemon = True
            stderr_thread.start()

            # Read stdout line by line
            for line in iter(process.stdout.readline, ''):
                line = line.strip()
                if line:
                    logger.info(f"[Profile Generation STDOUT]: {line}")

                    # Parse progress lines
                    if line.startswith('PROGRESS:'):
                        try:
                            progress_data = json.loads(line[9:])
                            profile_sse_queue.put(progress_data)
                            logger.debug(f"Progress event queued: {progress_data.get('type', 'unknown')}")
                        except json.JSONDecodeError as e:
                            logger.error(f"Failed to parse progress JSON: {e}")
                    else:
                        # Queue other output as info logs
                        profile_sse_queue.put({
                            'type': 'log',
                            'text': line,
                            'timestamp': time.strftime('%H:%M:%S')
                        })

            # Wait for process completion
            process.wait()
            stderr_thread.join(timeout=5)

            logger.info(f"Profile generation process completed with return code: {process.returncode}")

            # Send completion event
            if process.returncode == 0:
                profile_sse_queue.put({
                    'type': 'end',
                    'status': 'completed',
                    'message': f'Profile "{profile_name}" generated successfully',
                    'profile_name': profile_name
                })
                profile_processes[process_id]['status'] = 'completed'
            else:
                error_msg = f'Profile generation failed with code {process.returncode}'
                logger.error(error_msg)
                profile_sse_queue.put({
                    'type': 'end',
                    'status': 'failed',
                    'error': error_msg,
                    'profile_name': profile_name
                })
                profile_processes[process_id]['status'] = 'failed'

        except Exception as e:
            logger.error(f"Profile generation subprocess error: {e}")
            import traceback
            traceback.print_exc()
            profile_sse_queue.put({
                'type': 'end',
                'status': 'failed',
                'error': str(e),
                'profile_name': profile_name
            })
            profile_processes[process_id]['status'] = 'failed'

    # ==================== PyPSA ====================

    def get_pypsa_scenarios(self, project_path: str) -> Dict:
        """List all PyPSA scenarios"""
        try:
            pypsa_dir = os.path.join(project_path, 'results', 'pypsa_optimization')

            if not os.path.exists(pypsa_dir):
                return {'scenarios': []}

            scenarios = [d for d in os.listdir(pypsa_dir)
                        if os.path.isdir(os.path.join(pypsa_dir, d))]

            return {'scenarios': scenarios}

        except Exception as e:
            logger.error(f"Error getting PyPSA scenarios: {e}")
            return {'scenarios': []}

    def run_pypsa_model(self, config: Dict) -> Dict:
        """Execute PyPSA model"""
        try:
            # Import PyPSA model executor
            import sys
            models_path = os.path.join(os.path.dirname(__file__), '..', 'models')
            if models_path not in sys.path:
                sys.path.insert(0, models_path)

            from pypsa_model_executor import run_pypsa_model_complete

            # Execute PyPSA optimization
            results = run_pypsa_model_complete(config)

            return {
                'success': True,
                'process_id': 'local_pypsa',
                'results': results
            }

        except Exception as e:
            logger.error(f"Error running PyPSA model: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== T&D LOSSES ====================

    def get_td_losses(self, project_path: str, scenario_name: str) -> Dict:
        """
        Get T&D loss configuration (time-varying points)
        Matches React backend structure: [{year: X, loss: Y}, ...]
        """
        try:
            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)
            td_losses_file = os.path.join(scenario_dir, 'td_losses.json')

            if os.path.exists(td_losses_file):
                with open(td_losses_file, 'r') as f:
                    loss_points = json.load(f)
                return {'success': True, 'data': loss_points}

            # Default: single point at current year with 15% loss
            from datetime import datetime
            return {'success': True, 'data': [{'year': datetime.now().year, 'loss': 15}]}

        except Exception as e:
            logger.error(f"Error getting T&D losses: {e}")
            from datetime import datetime
            return {'success': True, 'data': [{'year': datetime.now().year, 'loss': 15}]}

    def save_td_losses(self, project_path: str, scenario_name: str, loss_points: list) -> Dict:
        """
        Save T&D loss configuration (time-varying points)
        Args:
            loss_points: List of {year: int, loss: float} dicts
        """
        try:
            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)
            os.makedirs(scenario_dir, exist_ok=True)

            td_losses_file = os.path.join(scenario_dir, 'td_losses.json')

            # Remove any 'id' fields that might exist from frontend
            clean_points = [{'year': int(p['year']), 'loss': float(p['loss'])} for p in loss_points]

            with open(td_losses_file, 'w') as f:
                json.dump(clean_points, f, indent=2)

            return {'success': True, 'message': 'T&D losses saved successfully'}

        except Exception as e:
            logger.error(f"Error saving T&D losses: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== METADATA ====================

    def get_scenario_metadata(self, project_path: str, scenario_name: str) -> Dict:
        """Get scenario metadata from scenario_meta.json"""
        try:
            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)

            # Try scenario_meta.json first (actual filename), then metadata.json (fallback)
            meta_file = os.path.join(scenario_dir, 'scenario_meta.json')
            if not os.path.exists(meta_file):
                meta_file = os.path.join(scenario_dir, 'metadata.json')

            if os.path.exists(meta_file):
                with open(meta_file, 'r') as f:
                    metadata = json.load(f)
                logger.info(f"Loaded metadata for scenario '{scenario_name}': {metadata}")
                return {'success': True, 'meta': metadata}

            logger.warning(f"No metadata file found for scenario '{scenario_name}'")
            return {'success': False, 'error': 'Metadata file not found'}

        except Exception as e:
            logger.error(f"Error getting metadata: {e}")
            return {'success': False, 'error': str(e)}

    def get_available_models(self, project_path: str, scenario_name: str) -> Dict:
        """Get available forecast models in scenario"""
        response = self.get_scenario_metadata(project_path, scenario_name)
        if response.get('success'):
            metadata = response.get('meta', {})
            return {'models': metadata.get('models', [])}
        return {'models': []}

    def _read_solar_share_data(self, project_path: str) -> Dict[str, float]:
        """
        Read solar share percentages for each sector from input_demand_file.xlsx.

        Returns:
            Dictionary mapping sector name to percentage share (e.g., {"Agriculture": 5.5})
        """
        try:
            file_path = os.path.join(project_path, 'inputs', 'input_demand_file.xlsx')

            if not os.path.exists(file_path):
                logger.warning(f"[read_solar_share_data] Excel file not found at: {file_path}")
                return {}

            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Find 'main' sheet (case-insensitive)
            main_sheet = None
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == 'main':
                    main_sheet = workbook[sheet_name]
                    break

            if not main_sheet:
                logger.warning("[read_solar_share_data] Sheet 'main' not found")
                workbook.close()
                return {}

            # Find ~Solar_share marker
            marker_row = None
            for row_idx, row in enumerate(main_sheet.iter_rows(values_only=True), start=1):
                for cell_value in row:
                    if isinstance(cell_value, str) and cell_value.strip().lower() == '~solar_share':
                        marker_row = row_idx
                        break
                if marker_row:
                    break

            if not marker_row:
                logger.info("[read_solar_share_data] Marker '~Solar_share' not found")
                workbook.close()
                return {}

            # Read header row
            headers_row = marker_row + 1
            headers = [main_sheet.cell(row=headers_row, column=col).value for col in range(1, main_sheet.max_column + 1)]

            # Find Sector and Percentage columns
            sector_col = None
            percentage_col = None

            for col_idx, header in enumerate(headers, start=1):
                if header:
                    header_lower = str(header).strip().lower()
                    if header_lower == 'sector':
                        sector_col = col_idx
                    elif 'percentage' in header_lower and 'share' in header_lower:
                        percentage_col = col_idx

            if not sector_col or not percentage_col:
                workbook.close()
                return {}

            # Read data rows
            solar_shares = {}
            for row_idx in range(headers_row + 1, main_sheet.max_row + 1):
                sector_name = main_sheet.cell(row=row_idx, column=sector_col).value
                percentage_value = main_sheet.cell(row=row_idx, column=percentage_col).value

                if not sector_name or sector_name == '':
                    break

                try:
                    percentage_float = float(percentage_value) if percentage_value else 0.0
                    solar_shares[str(sector_name).strip()] = percentage_float
                except (ValueError, TypeError):
                    solar_shares[str(sector_name).strip()] = 0.0

            workbook.close()
            logger.info(f"[read_solar_share_data] Successfully loaded solar shares for {len(solar_shares)} sectors")
            return solar_shares

        except Exception as e:
            logger.error(f"[read_solar_share_data] Error: {e}")
            return {}

    def _is_solar_sector(self, sector_name: str) -> bool:
        """
        Check if sector is a solar rooftop generation sector.
        Must contain BOTH "solar" AND "rooftop" (case-insensitive).
        """
        if not sector_name:
            return False
        name_lower = sector_name.lower()
        return 'solar' in name_lower and 'rooftop' in name_lower

    def _calculate_td_loss_percentage(self, target_year: int, loss_points: List[Dict]) -> float:
        """
        Calculate T&D loss percentage for a given year using linear interpolation.

        Returns:
            Loss percentage as decimal (e.g., 0.10 for 10%)
        """
        DEFAULT_LOSS = 0.10  # 10%

        if not loss_points or len(loss_points) == 0:
            return DEFAULT_LOSS

        # Filter and sort points
        sorted_points = sorted(
            [p for p in loss_points if 'year' in p and 'loss' in p],
            key=lambda p: p['year']
        )

        if len(sorted_points) == 0:
            return DEFAULT_LOSS
        if len(sorted_points) == 1:
            return sorted_points[0]['loss'] / 100

        first_point = sorted_points[0]
        last_point = sorted_points[-1]

        # Extrapolation
        if target_year <= first_point['year']:
            return first_point['loss'] / 100
        if target_year >= last_point['year']:
            return last_point['loss'] / 100

        # Interpolation
        for i in range(len(sorted_points) - 1):
            p1 = sorted_points[i]
            p2 = sorted_points[i + 1]

            if target_year >= p1['year'] and target_year <= p2['year']:
                if p2['year'] - p1['year'] == 0:
                    return p1['loss'] / 100

                # Linear interpolation
                interpolated_loss = p1['loss'] + (target_year - p1['year']) * (p2['loss'] - p1['loss']) / (p2['year'] - p1['year'])
                return interpolated_loss / 100

        return DEFAULT_LOSS

    def calculate_consolidated(self, project_path: str, scenario_name: str, start_year: int,
                              end_year: int, model_selections: Dict[str, str],
                              demand_type: str = 'gross') -> Dict:
        """
        Calculate consolidated forecast from individual sector files.

        Args:
            project_path: Project root directory
            scenario_name: Scenario name
            start_year: Start year for data
            end_year: End year for data
            model_selections: Dict mapping sector names to selected model names
            demand_type: 'gross', 'net', or 'onGrid'

        Returns:
            Dict with 'success' and 'data' (list of records with Year and sector columns)
        """
        try:
            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)

            if not os.path.exists(scenario_dir):
                return {'success': False, 'error': 'Scenario folder not found'}

            # Get all Excel files except Consolidated_Results and temporary files
            excel_files = [
                f for f in os.listdir(scenario_dir)
                if f.endswith('.xlsx') and f != 'Consolidated_Results.xlsx' and not f.startswith('~$')
            ]
            sectors = [os.path.splitext(f)[0] for f in excel_files]

            # Create year range
            years = list(range(start_year, end_year + 1))

            # Initialize consolidated data
            consolidated_data = []
            data_map = {}
            for year in years:
                row_data = {"Year": year}
                for sector in sectors:
                    row_data[sector] = None
                consolidated_data.append(row_data)
                data_map[year] = row_data

            # Read data from each sector
            for sector in sectors:
                selected_model = model_selections.get(sector)
                if not selected_model:
                    continue

                file_path = os.path.join(scenario_dir, f"{sector}.xlsx")
                if not os.path.exists(file_path):
                    continue

                try:
                    df = pd.read_excel(file_path, sheet_name='Results')

                    for _, row in df.iterrows():
                        year = row.get('Year')
                        if year and int(year) in data_map:
                            value = row.get(selected_model)
                            if value is not None:
                                data_map[int(year)][sector] = value

                except Exception as e:
                    logger.error(f"Error reading sector {sector}: {e}")
                    continue

            # Load solar share data for net demand calculation
            solar_shares = {}
            if demand_type in ["net", "onGrid"]:
                solar_shares = self._read_solar_share_data(project_path)
                logger.info(f"[calculate_consolidated] Loaded solar shares: {solar_shares}")

            # Load T&D losses
            td_loss_points = []
            losses_file_path = os.path.join(scenario_dir, 'td_losses.json')
            if os.path.exists(losses_file_path):
                try:
                    with open(losses_file_path, 'r') as f:
                        td_loss_points = json.load(f)
                except Exception as e:
                    logger.error(f"Could not parse td_losses.json: {e}")

            # Calculate totals based on demand type
            for row in consolidated_data:
                year = row['Year']

                # Apply solar sector abs() transformation
                for sector in sectors:
                    if row[sector] is not None and self._is_solar_sector(sector):
                        row[sector] = abs(float(row[sector]))

                if demand_type == "gross":
                    # GROSS DEMAND: Sector values (original) + Total ONLY
                    # NO intermediate columns, NO T&D breakdown
                    gross_total = sum(
                        float(row[sector]) if row[sector] is not None else 0
                        for sector in sectors
                    )
                    row['Total'] = gross_total

                elif demand_type == "net":
                    # NET DEMAND: Subtract solar rooftop share from each sector
                    # Solar rooftop sector will be HIDDEN in frontend (already subtracted)
                    # Show: Sectors (solar-adjusted) + Total ONLY
                    net_total = 0

                    for sector in sectors:
                        if row[sector] is not None:
                            sector_value = float(row[sector])

                            # For non-solar sectors, subtract their solar share
                            if not self._is_solar_sector(sector):
                                solar_share_pct = solar_shares.get(sector, 0.0)
                                net_value = sector_value - (sector_value * solar_share_pct / 100.0)
                                row[sector] = net_value
                                net_total += net_value
                            else:
                                # For solar sectors, keep the absolute value (will be hidden in UI)
                                row[sector] = abs(sector_value)
                                net_total += abs(sector_value)

                    row['Total'] = net_total

                elif demand_type == "onGrid":
                    # ON GRID DEMAND: Net demand + T&D losses
                    # Solar rooftop sector will be HIDDEN in frontend
                    # Show: Sectors (solar-adjusted) + T&D Loss (%) + T&D Losses + Total
                    net_total = 0

                    for sector in sectors:
                        if row[sector] is not None:
                            sector_value = float(row[sector])

                            # For non-solar sectors, subtract their solar share
                            if not self._is_solar_sector(sector):
                                solar_share_pct = solar_shares.get(sector, 0.0)
                                net_value = sector_value - (sector_value * solar_share_pct / 100.0)
                                row[sector] = net_value
                                net_total += net_value
                            else:
                                # For solar sectors, keep the absolute value (will be hidden in UI)
                                row[sector] = abs(sector_value)
                                net_total += abs(sector_value)

                    # Apply T&D losses to net demand
                    td_percentage = self._calculate_td_loss_percentage(year, td_loss_points)
                    td_losses = net_total * (td_percentage / (1 - td_percentage))
                    final_total = net_total + td_losses

                    # Add T&D breakdown columns
                    row['T&D Loss (%)'] = td_percentage
                    row['T&D Losses'] = td_losses
                    row['Total'] = final_total

            return {'success': True, 'data': consolidated_data}

        except Exception as e:
            logger.error(f"Error calculating consolidated: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def save_consolidated_data(self, project_path: str, scenario_name: str, data: Dict) -> Dict:
        """Save consolidated results to Excel in scenario folder for load profile generation"""
        try:
            import openpyxl
            from openpyxl.styles import Font

            scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario_name)
            os.makedirs(scenario_dir, exist_ok=True)

            # IMPORTANT: Save as Consolidated_Results.xlsx (not forecast_results.xlsx)
            # This file is used by load profile generation
            excel_path = os.path.join(scenario_dir, 'Consolidated_Results.xlsx')

            # Process data: Convert T&D Loss (%) to percentage string
            processed_data = []
            for row in data:
                new_row = dict(row)
                if 'T&D Loss (%)' in new_row and new_row['T&D Loss (%)'] is not None:
                    new_row['T&D Loss (%)'] = f"{float(new_row['T&D Loss (%)']) * 100:.2f}%"
                processed_data.append(new_row)

            # Create workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Consolidated Data"

            # Write headers
            if processed_data:
                headers = list(processed_data[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)

                # Write data
                for row_idx, row_data in enumerate(processed_data, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

            # Save workbook
            workbook.save(excel_path)
            workbook.close()

            logger.info(f"✅ Consolidated results saved to: {excel_path}")

            return {
                'success': True,
                'message': 'File saved successfully!',
                'path': str(excel_path)
            }

        except Exception as e:
            logger.error(f"Error saving consolidated data: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    # ==================== PyPSA NETWORK METHODS ====================

    def get_pypsa_networks(self, project_path: str, scenario_name: str) -> Dict:
        """List all .nc network files in PyPSA scenario folder"""
        try:
            pypsa_dir = os.path.join(project_path, 'results', 'pypsa_optimization', scenario_name)

            if not os.path.exists(pypsa_dir):
                return {'networks': []}

            networks = [f for f in os.listdir(pypsa_dir) if f.endswith('.nc')]

            return {'networks': networks}

        except Exception as e:
            logger.error(f"Error getting PyPSA networks: {e}")
            return {'networks': []}

    def get_network_info(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Load PyPSA network and extract metadata"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            if not os.path.exists(network_path):
                return {'success': False, 'error': 'Network file not found'}

            # Load network with caching for 10-100x performance improvement
            network = load_network_cached(network_path)

            # Extract metadata
            info = {
                'success': True,
                'buses': len(network.buses),
                'generators': len(network.generators),
                'lines': len(network.lines),
                'loads': len(network.loads),
                'storage_units': len(network.storage_units),
                'snapshots': len(network.snapshots),
                'multi_period': False
            }

            # Check for multi-period
            if hasattr(network, 'investment_periods'):
                info['multi_period'] = len(network.investment_periods) > 0
                if info['multi_period']:
                    info['investment_periods'] = network.investment_periods.tolist()

            return info

        except Exception as e:
            logger.error(f"Error getting network info: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_buses(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get all buses from PyPSA network"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            return {
                'success': True,
                'buses': network.buses.reset_index().to_dict('records')
            }

        except Exception as e:
            logger.error(f"Error getting buses: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_generators(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get all generators from PyPSA network"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            return {
                'success': True,
                'generators': network.generators.reset_index().to_dict('records')
            }

        except Exception as e:
            logger.error(f"Error getting generators: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_storage_units(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get all storage units from PyPSA network"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            return {
                'success': True,
                'storage_units': network.storage_units.reset_index().to_dict('records')
            }

        except Exception as e:
            logger.error(f"Error getting storage units: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_lines(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get all transmission lines from PyPSA network"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            return {
                'success': True,
                'lines': network.lines.reset_index().to_dict('records')
            }

        except Exception as e:
            logger.error(f"Error getting lines: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_loads(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get all loads from PyPSA network"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            return {
                'success': True,
                'loads': network.loads.reset_index().to_dict('records')
            }

        except Exception as e:
            logger.error(f"Error getting loads: {e}")
            return {'success': False, 'error': str(e)}

    def get_comprehensive_analysis(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get comprehensive network analysis including dispatch, capacity, etc."""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            analysis = {
                'success': True,
                'network_name': network_file,
                'snapshots': len(network.snapshots),
                'components': {
                    'buses': len(network.buses),
                    'generators': len(network.generators),
                    'storage_units': len(network.storage_units),
                    'lines': len(network.lines),
                    'loads': len(network.loads)
                }
            }

            # Get generator dispatch (time series)
            if hasattr(network, 'generators_t') and hasattr(network.generators_t, 'p'):
                dispatch = network.generators_t.p
                analysis['dispatch'] = {
                    'timestamps': dispatch.index.tolist(),
                    'data': dispatch.to_dict()
                }

            # Get load time series
            if hasattr(network, 'loads_t') and hasattr(network.loads_t, 'p'):
                loads = network.loads_t.p
                analysis['loads_t'] = {
                    'timestamps': loads.index.tolist(),
                    'data': loads.to_dict()
                }

            # Get storage state of charge
            if hasattr(network, 'storage_units_t') and hasattr(network.storage_units_t, 'state_of_charge'):
                soc = network.storage_units_t.state_of_charge
                analysis['storage_soc'] = {
                    'timestamps': soc.index.tolist(),
                    'data': soc.to_dict()
                }

            return analysis

        except Exception as e:
            logger.error(f"Error in comprehensive analysis: {e}")
            return {'success': False, 'error': str(e)}

    def get_available_years(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get available years for multi-period networks"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            years = []
            if hasattr(network, 'investment_periods') and len(network.investment_periods) > 0:
                years = network.investment_periods.tolist()
            else:
                # Single period - extract year from snapshots
                if len(network.snapshots) > 0:
                    years = [network.snapshots[0].year]

            return {'years': years}

        except Exception as e:
            logger.error(f"Error getting available years: {e}")
            return {'years': []}

    def get_plot_availability(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Detect which plot types are available based on network data"""
        try:
            import pypsa

            network_path = os.path.join(project_path, 'results', 'pypsa_optimization',
                                       scenario_name, network_file)

            # Use cached network loading for 10-100x performance improvement
            network = load_network_cached(network_path)

            availability = {
                'dispatch': hasattr(network, 'generators_t') and hasattr(network.generators_t, 'p'),
                'capacity': len(network.generators) > 0,
                'storage': len(network.storage_units) > 0,
                'lines': len(network.lines) > 0,
                'loads': len(network.loads) > 0,
                'multi_period': hasattr(network, 'investment_periods') and len(network.investment_periods) > 0
            }

            return availability

        except Exception as e:
            logger.error(f"Error checking plot availability: {e}")
            return {}

    # ==================== ADVANCED PYPSA ANALYSIS METHODS ====================

    def _get_network_path(self, project_path: str, scenario_name: str, network_file: str) -> str:
        """Helper to construct network path"""
        return os.path.join(project_path, 'results', 'pypsa_optimization',
                           scenario_name, network_file)

    def analyze_pypsa_network(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """
        Comprehensive PyPSA network analysis using pypsa_analyzer

        Returns detailed analysis including:
        - Energy mix by carrier
        - Capacity factors
        - Renewable share
        - Emissions
        - System costs
        - Dispatch profiles
        - Storage operations
        """
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            # Run all available analyses
            analysis = analyzer.run_all_analyses()

            return {
                'success': True,
                'analysis': analysis
            }
        except Exception as e:
            logger.error(f"Error analyzing network: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_energy_mix(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get energy generation mix by carrier (solar, wind, hydro, etc.)"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            energy_mix = analyzer.get_energy_mix()

            return {
                'success': True,
                'energy_mix': energy_mix
            }
        except Exception as e:
            logger.error(f"Error getting energy mix: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_capacity_factors(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get capacity utilization factors (CUF) for all generators"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            capacity_factors = analyzer.get_capacity_factors()

            return {
                'success': True,
                'capacity_factors': capacity_factors
            }
        except Exception as e:
            logger.error(f"Error getting capacity factors: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_renewable_share(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get renewable energy penetration percentage"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            renewable_share = analyzer.get_renewable_share()

            return {
                'success': True,
                'renewable_share': renewable_share
            }
        except Exception as e:
            logger.error(f"Error getting renewable share: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_emissions(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get CO2 emissions by carrier and total"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            emissions = analyzer.get_emissions_tracking()

            return {
                'success': True,
                'emissions': emissions
            }
        except Exception as e:
            logger.error(f"Error getting emissions: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_system_costs(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get total system costs breakdown (capital, operational, fuel)"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            system_costs = analyzer.get_system_costs()

            return {
                'success': True,
                'system_costs': system_costs
            }
        except Exception as e:
            logger.error(f"Error getting system costs: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_dispatch(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get hourly dispatch profiles by generator"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            dispatch = analyzer.get_dispatch_data()

            return {
                'success': True,
                'dispatch': dispatch
            }
        except Exception as e:
            logger.error(f"Error getting dispatch: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_capacity(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get installed capacity by carrier"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Create analyzer instance for this network
            analyzer = PyPSASingleNetworkAnalyzer(network)
            capacities = analyzer.get_total_capacities()

            return {
                'success': True,
                'capacities': capacities
            }
        except Exception as e:
            logger.error(f"Error getting capacity: {e}")
            return {'success': False, 'error': str(e)}

    def get_pypsa_storage(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """Get storage units information"""
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Get storage units data directly from network
            storage_data = {}
            if len(network.storage_units) > 0:
                storage_data = {
                    'storage_units': network.storage_units.to_dict('records'),
                    'count': len(network.storage_units)
                }

                # Add time series if available
                if hasattr(network, 'storage_units_t') and hasattr(network.storage_units_t, 'state_of_charge'):
                    soc = network.storage_units_t.state_of_charge
                    storage_data['state_of_charge'] = {
                        'timestamps': soc.index.tolist(),
                        'data': soc.to_dict()
                    }

            return {
                'success': True,
                'storage': storage_data
            }
        except Exception as e:
            logger.error(f"Error getting storage: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== MULTI-PERIOD DETECTION METHODS ====================

    def detect_network_type(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """
        Detect if network is single-period or multi-period optimization

        Returns:
            - network_type: 'single-period' or 'multi-period'
            - periods: List of periods (for multi-period)
            - years: List of years (for multi-period)
            - snapshots_structure: Description of snapshot structure
        """
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            # Check if snapshots are MultiIndex (multi-period indicator)
            import pandas as pd
            is_multi_period = isinstance(network.snapshots, pd.MultiIndex)

            result = {
                'success': True,
                'network_type': 'multi-period' if is_multi_period else 'single-period',
                'total_snapshots': len(network.snapshots)
            }

            if is_multi_period:
                # Extract periods from MultiIndex level 0
                periods = network.snapshots.get_level_values(0).unique().tolist()
                result['periods'] = [str(p) for p in periods]
                result['num_periods'] = len(periods)

                # Try to extract years from periods
                years = []
                for period in periods:
                    try:
                        # Periods might be years or date-like objects
                        if hasattr(period, 'year'):
                            years.append(period.year)
                        else:
                            # Try converting to int (if period is year as string/int)
                            years.append(int(str(period)))
                    except:
                        pass

                if years:
                    result['years'] = sorted(list(set(years)))

                # Snapshots per period
                snapshots_per_period = network.snapshots.get_level_values(1).nunique()
                result['snapshots_per_period'] = snapshots_per_period

            else:
                # Single period - extract year from first snapshot
                if len(network.snapshots) > 0:
                    first_snapshot = network.snapshots[0]
                    if hasattr(first_snapshot, 'year'):
                        result['year'] = first_snapshot.year

            return result

        except Exception as e:
            logger.error(f"Error detecting network type: {e}")
            return {'success': False, 'error': str(e)}

    def get_multi_year_info(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """
        Extract multi-year information from multi-period networks

        Returns:
            - years: List of optimization years
            - periods: List of period identifiers
            - period_details: Detailed info about each period
        """
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            import pandas as pd
            is_multi_period = isinstance(network.snapshots, pd.MultiIndex)

            if not is_multi_period:
                return {
                    'success': False,
                    'error': 'Network is not multi-period',
                    'network_type': 'single-period'
                }

            # Extract period information
            periods = network.snapshots.get_level_values(0).unique()
            period_details = []

            for period in periods:
                # Get snapshots for this period
                period_snapshots = network.snapshots[network.snapshots.get_level_values(0) == period]
                period_timestamps = period_snapshots.get_level_values(1)

                detail = {
                    'period': str(period),
                    'num_snapshots': len(period_timestamps),
                    'start': str(period_timestamps[0]) if len(period_timestamps) > 0 else None,
                    'end': str(period_timestamps[-1]) if len(period_timestamps) > 0 else None
                }

                # Try to extract year
                try:
                    if hasattr(period, 'year'):
                        detail['year'] = period.year
                    else:
                        detail['year'] = int(str(period))
                except:
                    pass

                period_details.append(detail)

            # Extract unique years
            years = []
            for detail in period_details:
                if 'year' in detail:
                    years.append(detail['year'])

            return {
                'success': True,
                'network_type': 'multi-period',
                'periods': [str(p) for p in periods.tolist()],
                'years': sorted(list(set(years))) if years else [],
                'num_periods': len(periods),
                'period_details': period_details
            }

        except Exception as e:
            logger.error(f"Error getting multi-year info: {e}")
            return {'success': False, 'error': str(e)}

    def get_period_comparison(self, project_path: str, scenario_name: str, network_file: str) -> Dict:
        """
        Compare key metrics across periods in multi-period optimization

        Returns comparison of:
        - Total generation by period
        - Capacity by period
        - Costs by period
        - Emissions by period
        """
        try:
            network_path = self._get_network_path(project_path, scenario_name, network_file)
            network = load_network_cached(network_path)

            import pandas as pd
            is_multi_period = isinstance(network.snapshots, pd.MultiIndex)

            if not is_multi_period:
                return {
                    'success': False,
                    'error': 'Network is not multi-period'
                }

            periods = network.snapshots.get_level_values(0).unique()
            comparison = []

            for period in periods:
                period_data = {'period': str(period)}

                # Get period snapshots
                period_mask = network.snapshots.get_level_values(0) == period

                # Total generation
                if hasattr(network, 'generators_t') and hasattr(network.generators_t, 'p'):
                    period_dispatch = network.generators_t.p.loc[period_mask]
                    period_data['total_generation_mwh'] = period_dispatch.sum().sum()

                # Installed capacity (should be same or growing)
                if len(network.generators) > 0:
                    period_data['total_capacity_mw'] = network.generators['p_nom_opt'].sum()

                comparison.append(period_data)

            return {
                'success': True,
                'comparison': comparison,
                'num_periods': len(periods)
            }

        except Exception as e:
            logger.error(f"Error getting period comparison: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== LOAD PROFILE ANALYSIS METHODS ====================

    def get_analysis_data(self, project_path: str, profile_name: str) -> Dict:
        """Get monthly/seasonal analysis data for load profile"""
        try:
            profile_dir = os.path.join(project_path, 'results', 'load_profiles', profile_name)

            # Check for pre-computed statistics
            stats_file = os.path.join(profile_dir, 'statistics.json')
            if os.path.exists(stats_file):
                with open(stats_file, 'r') as f:
                    return json.load(f)

            # If not found, calculate from CSV
            csv_file = os.path.join(profile_dir, 'hourly_profile.csv')
            if not os.path.exists(csv_file):
                return {'success': False, 'error': 'Profile CSV not found'}

            df = pd.read_csv(csv_file, parse_dates=['Timestamp'] if 'Timestamp' in pd.read_csv(csv_file, nrows=1).columns else None)

            # Calculate monthly statistics
            if 'Timestamp' in df.columns:
                df['Month'] = pd.to_datetime(df['Timestamp']).dt.month
                monthly = df.groupby('Month')['Load_MW'].agg(['mean', 'max', 'min', 'sum']).to_dict('index')
            else:
                monthly = {}

            return {
                'success': True,
                'monthly': monthly,
                'overall': {
                    'peak': df['Load_MW'].max(),
                    'average': df['Load_MW'].mean(),
                    'min': df['Load_MW'].min()
                }
            }

        except Exception as e:
            logger.error(f"Error getting analysis data: {e}")
            return {'success': False, 'error': str(e)}

    def get_profile_years(self, project_path: str, profile_name: str) -> Dict:
        """Get fiscal years available in load profile"""
        try:
            csv_file = os.path.join(project_path, 'results', 'load_profiles', profile_name, 'hourly_profile.csv')

            if not os.path.exists(csv_file):
                return {'years': []}

            df = pd.read_csv(csv_file)

            years = []
            if 'FiscalYear' in df.columns:
                years = sorted(df['FiscalYear'].unique().tolist())
            elif 'Timestamp' in df.columns:
                df['Year'] = pd.to_datetime(df['Timestamp']).dt.year
                years = sorted(df['Year'].unique().tolist())

            return {'years': years}

        except Exception as e:
            logger.error(f"Error getting profile years: {e}")
            return {'years': []}

    def get_load_duration_curve(self, project_path: str, profile_name: str, fiscal_year: str) -> Dict:
        """Get load duration curve data (sorted loads)"""
        try:
            csv_file = os.path.join(project_path, 'results', 'load_profiles', profile_name, 'hourly_profile.csv')

            df = pd.read_csv(csv_file)

            # Filter by fiscal year if specified
            if fiscal_year and 'FiscalYear' in df.columns:
                df = df[df['FiscalYear'] == fiscal_year]

            # Sort loads in descending order
            loads_sorted = df['Load_MW'].sort_values(ascending=False).reset_index(drop=True)

            # Calculate cumulative percentage
            hours = list(range(len(loads_sorted)))
            percentages = [(h / len(loads_sorted)) * 100 for h in hours]

            return {
                'success': True,
                'hours': hours,
                'loads': loads_sorted.tolist(),
                'percentages': percentages
            }

        except Exception as e:
            logger.error(f"Error getting load duration curve: {e}")
            return {'success': False, 'error': str(e)}

    def get_full_load_profile(self, project_path: str, profile_name: str,
                              fiscal_year: Optional[str] = None,
                              month: Optional[str] = None,
                              season: Optional[str] = None) -> Dict:
        """Get full hourly load profile data with optional filters"""
        try:
            csv_file = os.path.join(project_path, 'results', 'load_profiles', profile_name, 'hourly_profile.csv')

            df = pd.read_csv(csv_file, parse_dates=['Timestamp'] if 'Timestamp' in pd.read_csv(csv_file, nrows=1).columns else None)

            # Apply filters
            if fiscal_year and 'FiscalYear' in df.columns:
                df = df[df['FiscalYear'] == fiscal_year]

            if month and 'Timestamp' in df.columns:
                df['Month'] = pd.to_datetime(df['Timestamp']).dt.month
                df = df[df['Month'] == int(month)]

            if season and 'Season' in df.columns:
                df = df[df['Season'] == season]

            return {
                'success': True,
                'data': df.to_dict('records')
            }

        except Exception as e:
            logger.error(f"Error getting full load profile: {e}")
            return {'success': False, 'error': str(e)}

    def get_load_profile_statistics(self, project_path: str, profile_name: str,
                                    fiscal_year: Optional[str] = None) -> Dict:
        """
        Get comprehensive load profile statistics

        Returns:
            - Peak demand (MW) and timestamp
            - Minimum demand (MW) and timestamp
            - Average demand (MW)
            - Median demand (MW)
            - Standard deviation
            - Load factor (average/peak)
            - Percentiles (p95, p75, p50, p25, p05)
            - Peak hour of day analysis
            - Total energy (MWh)
        """
        try:
            csv_file = os.path.join(project_path, 'results', 'load_profiles', profile_name, 'hourly_profile.csv')

            if not os.path.exists(csv_file):
                return {'success': False, 'error': 'Profile CSV not found'}

            df = pd.read_csv(csv_file, parse_dates=['Timestamp'] if 'Timestamp' in pd.read_csv(csv_file, nrows=1).columns else None)

            # Filter by fiscal year if specified
            if fiscal_year and 'FiscalYear' in df.columns:
                df = df[df['FiscalYear'] == fiscal_year]

            if len(df) == 0:
                return {'success': False, 'error': 'No data available'}

            loads = df['Load_MW']

            # Basic statistics
            peak_demand = loads.max()
            min_demand = loads.min()
            avg_demand = loads.mean()
            median_demand = loads.median()
            std_deviation = loads.std()

            # Load factor (ratio of average to peak demand)
            load_factor = (avg_demand / peak_demand) * 100 if peak_demand > 0 else 0

            # Percentiles
            p95 = loads.quantile(0.95)
            p75 = loads.quantile(0.75)
            p50 = loads.quantile(0.50)  # Same as median
            p25 = loads.quantile(0.25)
            p05 = loads.quantile(0.05)

            # Total energy (MWh) - assuming hourly data
            total_energy_mwh = loads.sum()

            # Peak and minimum timestamps
            peak_idx = loads.idxmax()
            min_idx = loads.idxmin()

            result = {
                'success': True,
                'peak_demand_mw': float(peak_demand),
                'min_demand_mw': float(min_demand),
                'avg_demand_mw': float(avg_demand),
                'median_demand_mw': float(median_demand),
                'std_deviation': float(std_deviation),
                'load_factor_percent': float(load_factor),
                'percentiles': {
                    'p95': float(p95),
                    'p75': float(p75),
                    'p50': float(p50),
                    'p25': float(p25),
                    'p05': float(p05)
                },
                'total_energy_mwh': float(total_energy_mwh),
                'num_hours': len(df)
            }

            # Add timestamps if available
            if 'Timestamp' in df.columns:
                result['peak_timestamp'] = str(df.loc[peak_idx, 'Timestamp'])
                result['min_timestamp'] = str(df.loc[min_idx, 'Timestamp'])

                # Peak hour of day analysis
                df['Hour'] = pd.to_datetime(df['Timestamp']).dt.hour
                hourly_avg = df.groupby('Hour')['Load_MW'].mean()
                peak_hour = hourly_avg.idxmax()
                result['peak_hour_of_day'] = int(peak_hour)
                result['hourly_average'] = {int(h): float(v) for h, v in hourly_avg.items()}

            return result

        except Exception as e:
            logger.error(f"Error getting load profile statistics: {e}")
            return {'success': False, 'error': str(e)}

    def get_seasonal_analysis(self, project_path: str, profile_name: str,
                             fiscal_year: Optional[str] = None) -> Dict:
        """
        Get seasonal analysis of load profile

        Seasons (India):
            - Monsoon: June - September
            - Post-monsoon: October - November
            - Winter: December - February
            - Summer: March - May

        Returns for each season:
            - Peak demand (MW)
            - Average demand (MW)
            - Minimum demand (MW)
            - Total energy (MWh)
            - Load factor (%)
            - Number of hours
        """
        try:
            csv_file = os.path.join(project_path, 'results', 'load_profiles', profile_name, 'hourly_profile.csv')

            if not os.path.exists(csv_file):
                return {'success': False, 'error': 'Profile CSV not found'}

            df = pd.read_csv(csv_file, parse_dates=['Timestamp'] if 'Timestamp' in pd.read_csv(csv_file, nrows=1).columns else None)

            # Filter by fiscal year if specified
            if fiscal_year and 'FiscalYear' in df.columns:
                df = df[df['FiscalYear'] == fiscal_year]

            if 'Timestamp' not in df.columns:
                return {'success': False, 'error': 'Timestamp column not found'}

            # Extract month
            df['Month'] = pd.to_datetime(df['Timestamp']).dt.month

            # Define seasons based on month
            def get_season(month):
                if month in [6, 7, 8, 9]:
                    return 'Monsoon'
                elif month in [10, 11]:
                    return 'Post-monsoon'
                elif month in [12, 1, 2]:
                    return 'Winter'
                else:  # 3, 4, 5
                    return 'Summer'

            df['Season'] = df['Month'].apply(get_season)

            # Calculate statistics for each season
            seasonal_stats = {}
            for season in ['Monsoon', 'Post-monsoon', 'Winter', 'Summer']:
                season_df = df[df['Season'] == season]

                if len(season_df) == 0:
                    continue

                loads = season_df['Load_MW']
                peak = loads.max()
                avg = loads.mean()
                min_load = loads.min()
                total_energy = loads.sum()
                load_factor = (avg / peak) * 100 if peak > 0 else 0

                seasonal_stats[season] = {
                    'peak_demand_mw': float(peak),
                    'avg_demand_mw': float(avg),
                    'min_demand_mw': float(min_load),
                    'total_energy_mwh': float(total_energy),
                    'load_factor_percent': float(load_factor),
                    'num_hours': len(season_df),
                    'months': sorted(season_df['Month'].unique().tolist())
                }

            # Overall comparison
            total_peak = df['Load_MW'].max()
            overall_stats = {
                'peak_season': max(seasonal_stats.items(), key=lambda x: x[1]['peak_demand_mw'])[0] if seasonal_stats else None,
                'lowest_season': min(seasonal_stats.items(), key=lambda x: x[1]['avg_demand_mw'])[0] if seasonal_stats else None,
                'annual_peak_mw': float(total_peak)
            }

            return {
                'success': True,
                'seasonal_stats': seasonal_stats,
                'overall': overall_stats
            }

        except Exception as e:
            logger.error(f"Error getting seasonal analysis: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== ADDITIONAL HELPER METHODS ====================

    def get_available_base_years(self, project_path: str) -> Dict:
        """
        Get available base years from load curve template.

        EXACT IMPLEMENTATION FROM FASTAPI:
        - File: inputs/load_curve_template.xlsx
        - Sheet: 'Past_Hourly_Demand' (case-insensitive)
        - Column: 'date' (case-insensitive)
        - Extract financial years from dates (Apr-Mar cycle)
        - Return format: ['FY2021', 'FY2022', ...]
        """
        try:
            import openpyxl
            from datetime import datetime

            file_path = os.path.join(project_path, 'inputs', 'load_curve_template.xlsx')

            if not os.path.exists(file_path):
                logger.warning(f"load_curve_template.xlsx not found at: {file_path}")
                return {'success': True, 'years': []}

            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Find sheet (case-insensitive) - EXACT MATCH TO FASTAPI
            sheet_name = next(
                (name for name in workbook.sheetnames if name.lower() == 'past_hourly_demand'),
                None
            )

            if not sheet_name:
                workbook.close()
                logger.error("Sheet 'Past_Hourly_Demand' not found in load_curve_template.xlsx")
                return {'success': False, 'years': []}

            worksheet = workbook[sheet_name]

            # Read data - EXACT MATCH TO FASTAPI
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                data.append(row_dict)

            if len(data) == 0:
                workbook.close()
                return {'success': True, 'years': []}

            # Find date column (case-insensitive) - EXACT MATCH TO FASTAPI
            date_header = next(
                (key for key in headers if key and key.lower() == 'date'),
                None
            )

            if not date_header:
                workbook.close()
                logger.error("Column 'date' not found in 'Past_Hourly_Demand' sheet")
                return {'success': False, 'years': []}

            # Extract financial years (Apr-Mar cycle) - EXACT MATCH TO FASTAPI
            financial_years = set()
            for row in data:
                date_value = row.get(date_header)
                if date_value:
                    # Handle both datetime objects and string dates
                    if isinstance(date_value, datetime):
                        parsed_date = date_value
                    else:
                        try:
                            parsed_date = datetime.fromisoformat(str(date_value))
                        except:
                            continue

                    if parsed_date:
                        # Calculate financial year (Apr-Mar)
                        year = parsed_date.year
                        month = parsed_date.month
                        financial_year = year + 1 if month >= 4 else year
                        financial_years.add(f"FY{financial_year}")

            workbook.close()

            # Sort years - EXACT MATCH TO FASTAPI
            sorted_years = sorted(financial_years, key=lambda x: int(x[2:]))

            logger.info(f"Found {len(sorted_years)} base years: {sorted_years}")
            return {'success': True, 'years': sorted_years}

        except Exception as e:
            logger.error(f"Error getting base years: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'years': []}

    def get_available_scenarios_for_profiles(self, project_path: str) -> Dict:
        """
        Get completed forecast scenarios for profile generation.

        EXACT IMPLEMENTATION FROM FASTAPI:
        - Path: results/demand_forecasts/{scenario}/Consolidated_Results.xlsx
        - Returns scenarios that have this file
        - Format: {'success': True, 'scenarios': [...]}
        """
        try:
            from pathlib import Path

            scenarios_parent_path = Path(project_path) / "results" / "demand_forecasts"

            if not scenarios_parent_path.exists():
                return {'success': True, 'scenarios': []}

            # Find directories with Consolidated_Results.xlsx - EXACT MATCH TO FASTAPI
            valid_scenarios = []
            for item in scenarios_parent_path.iterdir():
                if not item.is_dir():
                    continue

                expected_file = item / "Consolidated_Results.xlsx"
                if expected_file.exists():
                    valid_scenarios.append(item.name)

            logger.info(f"Found {len(valid_scenarios)} scenarios: {valid_scenarios}")
            return {'success': True, 'scenarios': sorted(valid_scenarios)}

        except Exception as e:
            logger.error(f"Error getting scenarios for profiles: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'scenarios': []}

    def check_profile_exists(self, project_path: str, profile_name: str) -> Dict:
        """
        Check if a load profile already exists.

        EXACT IMPLEMENTATION FROM FASTAPI:
        - Path: results/load_profiles/{profileName}.xlsx
        - Checks for file existence (not directory)
        """
        try:
            # FIXED: Check for .xlsx file (not directory)
            file_path = os.path.join(project_path, 'results', 'load_profiles', f"{profile_name}.xlsx")
            exists = os.path.exists(file_path)

            return {'exists': exists}

        except Exception as e:
            logger.error(f"Error checking profile exists: {e}")
            return {'exists': False}

    def save_model_config(self, project_path: str, config: Dict) -> Dict:
        """Save PyPSA model configuration"""
        try:
            config_dir = os.path.join(project_path, 'results', 'pypsa_optimization')
            os.makedirs(config_dir, exist_ok=True)

            config_file = os.path.join(config_dir, 'model_config.json')

            with open(config_file, 'w') as f:
                json.dump(config, f, indent=2)

            return {'success': True}

        except Exception as e:
            logger.error(f"Error saving model config: {e}")
            return {'success': False, 'error': str(e)}

    def get_cache_stats(self) -> Dict:
        """Get PyPSA network cache statistics"""
        try:
            stats = get_cache_stats()
            return {
                'success': True,
                'stats': stats
            }
        except Exception as e:
            logger.error(f"Error getting cache stats: {e}")
            return {'success': False, 'error': str(e)}

    def invalidate_cache(self, network_path: Optional[str] = None) -> Dict:
        """
        Invalidate network cache

        Args:
            network_path: Optional specific network path to invalidate.
                         If None, clears entire cache.
        """
        try:
            invalidate_network_cache(network_path)
            message = f"Cache cleared for {network_path}" if network_path else "Entire cache cleared"
            return {
                'success': True,
                'message': message
            }
        except Exception as e:
            logger.error(f"Error invalidating cache: {e}")
            return {'success': False, 'error': str(e)}

    # ==================== MODEL REGISTRY & VALIDATION ====================

    def get_available_models(self) -> Dict:
        """
        Get available forecasting models from registry.

        Returns:
            {
                'success': True,
                'models': [...],  # List of model configs
                'defaults': {...}  # Default configuration
            }
        """
        try:
            if not MODEL_REGISTRY_AVAILABLE:
                return {
                    'success': False,
                    'error': 'Model registry not available',
                    'models': [],
                    'defaults': {}
                }

            registry = export_model_registry()
            return {
                'success': True,
                **registry
            }

        except Exception as e:
            logger.error(f"Error getting available models: {e}")
            return {
                'success': False,
                'error': str(e),
                'models': [],
                'defaults': {}
            }

    def get_sector_forecast_metadata(self, project_path: str, sector: str) -> Dict:
        """
        Get comprehensive forecast metadata for a sector.

        Args:
            project_path: Project root path
            sector: Sector name

        Returns:
            Complete metadata including available models, parameters, recommendations
        """
        try:
            if not MODEL_REGISTRY_AVAILABLE:
                # Fallback to basic metadata
                return self._get_sector_metadata_legacy(project_path, sector)

            # Extract sector data
            data_response = self.extract_sector_data(project_path, sector)
            if not data_response.get('success'):
                return {
                    'success': False,
                    'error': f'Cannot load data for sector {sector}: {data_response.get("error")}'
                }

            data = data_response.get('data', [])
            row_count = len(data)

            # Get year range
            years = [d.get('Year') for d in data if d.get('Year')]
            year_range = f"{min(years)}-{max(years)}" if years else 'N/A'

            # Get correlation analysis for MLR
            corr_response = self.get_sector_correlation(project_path, sector)
            mlr_params = []
            if corr_response.get('success'):
                corr_data = corr_response.get('data', [])
                mlr_params = [
                    c['variable'] for c in corr_data
                    if abs(c.get('correlation', 0)) > 0.3
                ]

            # Get available models for this sector
            available_models = get_available_models_for_sector(row_count, len(mlr_params) > 0)

            # Get recommendations
            recommended = get_recommended_models(row_count, mlr_params)

            # WAM configuration
            wam_max = calculate_wam_max_window(row_count)

            return {
                'success': True,
                'sector': sector,
                'data_statistics': {
                    'row_count': row_count,
                    'year_range': year_range
                },
                'available_models': available_models,
                'recommended_models': recommended,
                'mlr_parameters': mlr_params,
                'wam_config': {
                    'min_window': 2,
                    'max_window': wam_max,
                    'default_window': DEFAULT_FORECAST_CONFIG['default_wam_window']
                }
            }

        except Exception as e:
            logger.error(f"Error getting sector forecast metadata: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def _get_sector_metadata_legacy(self, project_path: str, sector: str) -> Dict:
        """Legacy fallback for sector metadata when registry unavailable"""
        try:
            data_response = self.extract_sector_data(project_path, sector)
            if not data_response.get('success'):
                return {'success': False, 'error': 'Cannot load sector data'}

            data = data_response.get('data', [])
            row_count = len(data)

            corr_response = self.get_sector_correlation(project_path, sector)
            mlr_params = []
            if corr_response.get('success'):
                mlr_params = [c['variable'] for c in corr_response.get('data', [])]

            return {
                'success': True,
                'sector': sector,
                'data_statistics': {
                    'row_count': row_count
                },
                'available_models': [
                    {'id': 'SLR', 'available': row_count >= 3},
                    {'id': 'MLR', 'available': row_count >= 5 and len(mlr_params) > 0},
                    {'id': 'WAM', 'available': row_count >= 4}
                ],
                'recommended_models': ['SLR', 'MLR', 'WAM'][:2],
                'mlr_parameters': mlr_params,
                'wam_config': {
                    'min_window': 2,
                    'max_window': max(3, row_count - 2),
                    'default_window': 3
                }
            }

        except Exception as e:
            return {'success': False, 'error': str(e)}

    def validate_forecast_configuration(self, config: Dict, sector_metadata: Dict) -> Dict:
        """
        Validate forecast configuration before execution.

        Args:
            config: Forecast configuration dict
            sector_metadata: Metadata dict per sector

        Returns:
            {
                'valid': bool,
                'errors': List[str],
                'warnings': List[str],
                'suggestions': List[str]
            }
        """
        try:
            if not MODEL_REGISTRY_AVAILABLE:
                return {
                    'valid': True,
                    'errors': [],
                    'warnings': ['Validation skipped - registry not available'],
                    'suggestions': []
                }

            result = validate_forecast_config(config, sector_metadata)
            return result.to_dict()

        except Exception as e:
            logger.error(f"Error validating forecast config: {e}")
            return {
                'valid': False,
                'errors': [f'Validation error: {str(e)}'],
                'warnings': [],
                'suggestions': []
            }

    def get_sector_available_models(self, project_path: str, scenario: str) -> Dict:
        """
        Get available models for ALL sectors in a scenario - FULLY DYNAMIC.
        Reads actual Excel Results sheets to determine what models exist.

        Args:
            project_path: Project root path
            scenario: Scenario name

        Returns:
            {
                'success': True,
                'models': {
                    'Agriculture': ['Historical', 'MLR', 'SLR', 'WAM'],
                    'Domestic_lt': ['Historical', 'MLR', 'SLR', 'User Data'],
                    ...
                }
            }
        """
        try:
            # Get scenario sectors (from actual Excel files in results folder)
            sectors_response = self.get_scenario_sectors(project_path, scenario)
            sectors = sectors_response.get('sectors', [])

            if not sectors:
                return {
                    'success': False,
                    'error': 'No sectors found for scenario',
                    'models': {}
                }

            models_per_sector = {}

            for sector in sectors:
                try:
                    # Read the actual Excel file to see what models exist
                    scenario_dir = os.path.join(project_path, 'results', 'demand_forecasts', scenario)
                    excel_path = os.path.join(scenario_dir, f'{sector}.xlsx')

                    if not os.path.exists(excel_path):
                        logger.warning(f"Excel file not found: {excel_path}")
                        models_per_sector[sector] = []
                        continue

                    # Read Results sheet to get column names (= model names)
                    df = pd.read_excel(excel_path, sheet_name='Results')

                    # Get all columns except 'Year'
                    model_names = [col for col in df.columns if col.lower() != 'year']

                    # Filter out only models with actual data (not all NaN)
                    available_models = []
                    for model in model_names:
                        # Check if model column has at least one non-null value
                        if df[model].notna().any():
                            available_models.append(model)

                    models_per_sector[sector] = available_models

                except Exception as e:
                    logger.warning(f"Error reading models for sector {sector}: {e}")
                    # Fallback to empty list (will show "No models found" in UI)
                    models_per_sector[sector] = []

            return {
                'success': True,
                'models': models_per_sector
            }

        except Exception as e:
            logger.error(f"Error getting sector available models: {e}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'error': str(e),
                'models': {}
            }


# Global service instance
service = LocalService()
